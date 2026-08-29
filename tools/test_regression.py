#!/usr/bin/env python3
"""
Regression suite for defects found in the 2026-08-25 audit.

Each test pins a bug that was silent — the code ran, produced plausible output,
and was wrong. They are grouped by the defect they prevent from returning.

  python3 tools/test_regression.py
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from agents import runtime as rt
from tools import deal_profile as dp
from tools import minigraph as mg
from tools import grounding_gate as gg
from tools.deal_profile import DealProfile

PASS, FAIL, SKIP = [], [], []


def check(name: str, cond: bool, detail: str = "") -> None:
    (PASS if cond else FAIL).append(name)
    mark = "✓" if cond else "✗"
    print(f"  {mark} {name}" + (f"  [{detail}]" if detail else ""))


def skip(name: str, reason: str) -> None:
    SKIP.append(name)
    print(f"\n  ⊙ {name} skipped  [{reason}]")


# ── 1. Model-graph shape: the staleness cascade ──────────────────────────────
# model_graph.json is a flat adjacency map. Three call sites read it as
# {"nodes": [...], "dependencies": {...}}, so subject lookup returned None and
# the cascade walked zero edges — it marked 1 node instead of 40.

def test_cascade() -> None:
    print("\n1. Staleness cascade over a flat adjacency graph")
    flat = {
        "MN-QOE-EBITDA":  ["MN-BASE-MOIC", "MN-BASE-IRR"],
        "MN-BASE-MOIC":   ["MN-EXIT"],
        "MN-BASE-IRR":    [],
        "MN-EXIT":        [],
        "MN-FIRM-EBITDA": [],
        "stale_nodes":    ["MN-QOE-EBITDA"],      # marker, never a node
    }
    check("subject resolves on adjacency shape",
          rt._node_for_subject(flat, "QoE-normalized EBITDA") == "MN-QOE-EBITDA")
    check("generic EBITDA falls back to firm basis",
          rt._node_for_subject(flat, "EBITDA") == "MN-FIRM-EBITDA")
    check("'stale_nodes' is not treated as a model node",
          "stale_nodes" not in [n["model_node_id"] for n in
                                ([{"model_node_id": k} for k in flat
                                  if k not in rt._GRAPH_RESERVED_KEYS])])

    # richer shape must still work
    rich = {"nodes": [{"model_node_id": "MN-QOE-EBITDA", "name": "QoE EBITDA"}],
            "dependencies": {"MN-QOE-EBITDA": []}}
    check("richer {'nodes':[...]} shape still resolves",
          rt._node_for_subject(rich, "QoE EBITDA") == "MN-QOE-EBITDA")

    # cascade must walk transitively: QOE -> {MOIC, IRR} -> EXIT  == 4 nodes
    import tempfile
    with tempfile.TemporaryDirectory() as td:
        deal_dir = Path(td) / "deals" / "t" / "models"
        deal_dir.mkdir(parents=True)
        (deal_dir / "model_graph.json").write_text(json.dumps(
            {k: v for k, v in flat.items() if k != "stale_nodes"}))
        old_vault = rt.VAULT
        try:
            rt.VAULT = Path(td)
            g = json.loads((deal_dir / "model_graph.json").read_text())
            chain = rt._mark_model_node_stale("t", g, "MN-QOE-EBITDA")
        finally:
            rt.VAULT = old_vault
    check("cascade walks transitively, not just the trigger node",
          len(chain) == 4, f"marked {len(chain)}: {chain}")


# ── 2. Benchmark scorer field name ───────────────────────────────────────────
# The scorer read item["epistemic"]; extract_v2 emits "epistemic_class", so the
# epistemic dimension scored 0.0% by construction regardless of quality.

def test_benchmark_field() -> None:
    print("\n2. Benchmark reads the epistemic field the extractor writes")
    from tools.extraction_quality import score_e3
    e3 = {
        "claims": [{
            "claim_id": "synthetic-1",
            "period": "FY2025A",
            "perimeter": "Synthetic standalone",
            "epistemic_class": "attested",
            "locator": "fixture.xlsx::Inputs!1:4",
        }],
        "extraction_metadata": {"compiler_fields_per_claim": []},
    }
    score = score_e3(e3)
    check("scorer accepts epistemic_class", score["rates"]["epistemic"] == 1.0)
    check("scorer recognises cell-addressable Excel locators",
          score["rates"]["excel_locator"] == 1.0)


# ── 3. Deal profile: no cross-deal perimeter leakage ─────────────────────────
# bridge_v7 defaulted a missing perimeter to the literal "Alderstone standalone",
# so any second deal would carry Alderstone's economic scope.

def test_deal_profile() -> None:
    print("\n3. Deal profile never lends one deal's perimeter to another")
    empty = DealProfile("other-deal")
    check("unknown deal gets no perimeter, not a borrowed one",
          empty.cp_perimeter("CP-EBITDA-FIRM", None) == "")
    check("unknown deal records a warning instead of guessing",
          len(empty.warnings) > 0)
    check("'unknown' from the extractor is not accepted as a perimeter",
          empty.cp_perimeter("CP-REVENUE", "unknown") == "")
    check("a deal's own claim perimeter is still honoured",
          empty.cp_perimeter("CP-REVENUE", "Astrelia consolidated")
          == "Astrelia consolidated")
    check("no deal-specific literal remains in bridge_v7",
          "Alderstone" not in (ROOT / "tools" / "bridge_v7.py").read_text())

    synthetic = DealProfile("synthetic", {
        "cp_institutional": {
            "CP-EBITDA-FIRM": {
                "perimeter": "Synthetic standalone, firm underwriting definition",
                "unit": "$m/year",
            }
        }
    })
    check("versioned synthetic profile loads", synthetic.loaded)
    check("declared perimeter wins over claim ambiguity",
          synthetic.cp_perimeter("CP-EBITDA-FIRM", "unknown")
          == "Synthetic standalone, firm underwriting definition")
    check("declared time-frequency unit is preserved",
          synthetic.cp_unit("CP-EBITDA-FIRM", "$m") == "$m/year")

    try:
        dp.load_profile("no-such-deal", strict=True)
        check("strict mode refuses a missing profile", False)
    except FileNotFoundError:
        check("strict mode refuses a missing profile", True)


# ── 4. Grounding gate ────────────────────────────────────────────────────────
# The extractor bled a customer name into an IC-memo claim and typed it
# `attested`. The gate must route that to human review without adjudicating.

def test_grounding_gate() -> None:
    print("\n4. Grounding gate catches mis-scoped and unsourced claims")
    prof = DealProfile("synthetic", {
        "entity_aliases": ["Alderstone"],
        "counterparty_entities": {"Riverton": "customer"},
        "perimeter_vocabulary": [
            "Alderstone standalone",
            "Alderstone standalone, firm underwriting definition",
        ],
    })
    source_texts = {
        "keystone_ic_memo.md": (
            "Alderstone prices the Riverton transaction on $11.4m EBITDA. "
            "Firm-underwritten EBITDA at entry is $11.4m."
        )
    }

    riverton = {
        "claim_id": "t-1",
        "statement": "Alderstone prices the Riverton transaction on $11.4m EBITDA.",
        "perimeter": "Riverton Group",
        "epistemic_class": "attested",
        "value": "11.4",
        "locator": "keystone_ic_memo.md::## Concern block",
    }
    codes = {f["code"] for f in gg.check_claim(riverton, prof, source_texts)}
    check("counterparty-as-perimeter is caught",
          "PERIMETER_IS_COUNTERPARTY" in codes, str(sorted(codes)))
    check("that finding blocks admission",
          "PERIMETER_IS_COUNTERPARTY" in gg.BLOCKING)

    good = {
        "claim_id": "t-2",
        "statement": "Firm-underwritten EBITDA at entry is $11.4m.",
        "perimeter": "Alderstone standalone, firm underwriting definition",
        "epistemic_class": "attested",
        "value": "11.4",
        "locator": "keystone_ic_memo.md::## Concern block",
    }
    check("a well-scoped, sourced claim is clean",
          gg.check_claim(good, prof, source_texts) == [],
          str(gg.check_claim(good, prof, source_texts)))

    derived = {
        "claim_id": "t-3", "statement": "Total debt is $9.4m.",
        "perimeter": "Alderstone standalone", "epistemic_class": "derived",
        "value": "9.4", "derivation": "9.0 + 0.4",
        "locator": "keystone_ic_memo.md::## X",
    }
    codes = {f["code"] for f in gg.check_claim(derived, prof)}
    check("derived value is not required to appear verbatim in source",
          "VALUE_NOT_IN_SOURCE" not in codes)
    no_deriv = {**derived, "derivation": ""}
    check("derived without a derivation is blocked (invariant 3)",
          "DERIVATION_MISSING" in {f["code"] for f in gg.check_claim(no_deriv, prof)})

    # number grounding tolerances
    check("accounting negative $(0.20)m grounds -0.2",
          gg._number_in_text("-0.2", "a reserve of $(0.20)m was applied"))
    check("a rounded quote grounds against fuller precision",
          gg._number_in_text("30.7", "gross margin of 30.72%"))
    check("spelled-out counts ground",
          gg._number_in_text("8", "approximately eight years"))
    check("an invented midpoint does not ground",
          not gg._number_in_text("1.5", "capex is approximately 1%-2% of revenue"))
    check("a value unrelated to the statement does not ground",
          not gg._number_in_text("1092.0", "7 active staff members"))


# ── 5. minigraph: stdlib replacement for the networkx slice ──────────────────
# graph_store took a networkx dependency, breaking the stdlib-only invariant and
# making the UI unimportable on a clean checkout.

def test_minigraph() -> None:
    print("\n5. minigraph matches known-correct graph algorithms")
    g = mg.DiGraph()
    for a, b in [("A", "B"), ("B", "C"), ("C", "A")]:
        g.add_edge(a, b)
    pr = mg.pagerank(g)
    check("pagerank is uniform on a symmetric 3-cycle",
          all(abs(pr[n] - 1 / 3) < 1e-4 for n in "ABC"))
    check("pagerank sums to 1", abs(sum(pr.values()) - 1.0) < 1e-9)

    dangling = mg.DiGraph(); dangling.add_edge("A", "B")
    prd = mg.pagerank(dangling)
    check("dangling mass is conserved", abs(sum(prd.values()) - 1.0) < 1e-9)

    path = mg.DiGraph(); path.add_edge("A", "B"); path.add_edge("B", "C")
    bc = mg.betweenness_centrality(path, normalized=False)
    check("betweenness: only the middle node scores",
          abs(bc["B"] - 1.0) < 1e-9 and bc["A"] == 0.0 and bc["C"] == 0.0)
    check("betweenness normalises by (n-1)(n-2)",
          abs(mg.betweenness_centrality(path)["B"] - 0.5) < 1e-9)
    check("shortest_path follows edge direction",
          mg.shortest_path(path, "A", "C") == ["A", "B", "C"])
    try:
        mg.shortest_path(path, "C", "A")
        check("unreachable target raises", False)
    except mg.NetworkXNoPath:
        check("unreachable target raises", True)

    rt_g = mg.DiGraph(); rt_g.add_node("X", type="claim"); rt_g.add_edge("X", "Y", rel="supports")
    back = mg.node_link_graph(mg.node_link_data(rt_g))
    check("node_link round-trip preserves nodes, edges and attrs",
          back.number_of_nodes() == 2 and back.has_edge("X", "Y")
          and back.nodes["X"].get("type") == "claim")

    check("graph_store imports with no third-party graph library",
          "minigraph" in (ROOT / "tools" / "graph_store.py").read_text())




# ── 6. Excel compiler: cycles and cell semantics ─────────────────────────────
# extract_v3 reads formulas rather than cached values, and sheet_semantics must
# never borrow a header across a block boundary — that produced a confident
# wrong period on a table that carries none.

def test_excel() -> None:
    print("\n6. Excel: solver ciclico e semantica delle celle")
    from tools.extract_v3 import (solve_component, strongly_connected_components,
                                  parse_cell_node)

    check("parse_cell_node riconosce una cella vera",
          parse_cell_node("'[M.XLSX]Cashflow'!B3") == ("CASHFLOW", "B3"))
    check("parse_cell_node scarta un nodo di espressione",
          parse_cell_node("'[M.XLSX]Cashflow'!B3))") is None)

    succ = {"a": {"b"}, "b": {"c"}, "c": {"a"}, "d": {"a"}}
    comps = [sorted(c) for c in strongly_connected_components(
        ["a", "b", "c", "d"], succ) if len(c) > 1]
    check("Tarjan trova il ciclo a->b->c->a", comps == [["a", "b", "c"]], str(comps))

    # revolver loop with the draw binding, same closed form as the self-test
    e, t, r, m, n = 2.85, 42.8, 0.085 / 4, 1.0, 2.5

    def ev(cell, v):
        if cell == "I": return r * (t + v["R"])
        if cell == "C": return e - v["I"] - n
        if cell == "R": return max(0.0, m - v["C"])
        raise KeyError(cell)

    rep = solve_component(["I", "C", "R"], ev, tolerance=1e-10, max_iter=500)
    want_i = r * (t + m - e + n) / (1 - r)
    check("punto fisso converge", rep.converged, f"{rep.iterations} iterazioni")
    check("converge al valore in forma chiusa",
          abs(rep.values["I"] - want_i) < 1e-6,
          f"{rep.values['I']:.6f} vs {want_i:.6f}")
    check("un componente non convergente non viene spacciato per risolto",
          not solve_component(["x"], lambda c, v: v["x"] + 1.0,
                              tolerance=1e-12, max_iter=5).converged)

    from tools.sheet_semantics import classify_period, infer_unit
    check("periodo: FY2025A riconosciuto", classify_period("FY2025A") == "fiscal_year")
    check("periodo: 'Value' non è un periodo", classify_period("Value") is None)
    check("unità: number_format batte l'etichetta",
          infer_unit("Revenue", "0.0%") == ("%", "number_format"))
    check("unità: '(x)' letto come multiplo", infer_unit("Net leverage (x)", "General")[0] == "x")


# ── 7. Binding resolver (L3) ─────────────────────────────────────────────────
# Meaning is resolved over the whole deal, not cell by cell. On an
# over-constrained system the resolver must halt and offer a relaxation, never
# pick a winner itself.

def test_resolver() -> None:
    print("\n7. Resolver dei binding: vincoli globali, stop su sovra-vincolo")
    from tools.binding_resolver import Concept, Binding, resolve, granularity_of, norm_unit

    check("granularita: fine trimestre riconosciuta",
          granularity_of("2026-06-30") == "quarter")
    check("granularita: FY riconosciuto", granularity_of("FY2025A") == "fiscal_year")
    check("unita: $mm e $m sono la stessa cosa", norm_unit("$mm") == norm_unit("$m"))

    concepts = {
        "C-REV": Concept("C-REV", "Revenue", unit="$mm", granularity="quarter",
                         form="derived"),
        "C-IN":  Concept("C-IN", "Opening cash", unit="$mm", granularity="point",
                         form="input"),
    }
    source = {"cells": {
        "S!C5":  {"kind": "formula", "value": "=A1+A2", "precedents": ["S!A1", "S!A2"]},
        "S!C9":  {"kind": "number",  "value": 3.0, "precedents": []},
        "S!C7":  {"kind": "formula", "value": "=S!C7+1", "precedents": ["S!C7"]},
    }}

    # a clean binding is admitted
    ok = resolve([Binding("C-REV", "S!C5", "2026-06-30", "SB", "", "$mm", 0.9)],
                 concepts, source)
    check("un binding coerente viene ammesso",
          len(ok.admitted) == 1 and not ok.violations, ok.status)

    # annual cell must not bind to a quarterly concept
    r = resolve([Binding("C-REV", "S!C5", "FY2025A", "SB", "", "$mm", 0.9)], concepts, source)
    check("una serie annuale non si lega a un concetto trimestrale",
          any(v.code == "PERIOD_ALIGNMENT" for v in r.violations))

    # a declared input must not be produced by a formula
    r = resolve([Binding("C-IN", "S!C5", "2026-03-31", "SB", "", "$mm", 0.9)], concepts, source)
    check("un input dichiarato non puo' essere calcolato",
          any(v.code == "PRECEDENT_SHAPE" for v in r.violations))

    # unit mismatch is caught even at high confidence
    r = resolve([Binding("C-REV", "S!C5", "2026-06-30", "SB", "", "%", 0.99)], concepts, source)
    check("unita incoerente scartata anche ad alta confidenza",
          any(v.code == "UNIT_COHERENCE" for v in r.violations))

    # self-reference
    r = resolve([Binding("C-REV", "S!C7", "2026-06-30", "SB", "", "$mm", 0.9)], concepts, source)
    check("una cella che si legge da sola viene rifiutata",
          any(v.code == "NO_SELF_REFERENCE" for v in r.violations))

    # two cells competing for the same slot: halt, do not choose
    r = resolve([Binding("C-REV", "S!C5", "2026-06-30", "SB", "", "$mm", 0.9),
                 Binding("C-REV", "S!C9", "2026-06-30", "SB", "", "$mm", 0.9)],
                concepts, source)
    dup = [v for v in r.violations if v.code == "UNIQUE_BINDING"]
    check("due candidati per lo stesso slot sono un conflitto", len(dup) == 1)
    check("il resolver non ne sceglie uno da solo", len(r.admitted) == 0)
    check("emette comunque una proposta laterale",
          bool(dup and dup[0].relaxation), str(dup[0].relaxation.get("kind") if dup else None))
    check("la proposta riporta il margine, non solo il vincitore",
          dup[0].relaxation.get("margin") == 0.0)
    check("sovra-vincolato => HALTED", r.halted, r.status)

    # the same line item repeated once per scenario block is not a conflict
    two = resolve([Binding("C-REV", "S!C5", "2026-06-30", "SD", unit="$mm", confidence=0.9,
                           section="Standalone Base"),
                   Binding("C-REV", "S!C9", "2026-06-30", "SD", unit="$mm",
                           confidence=0.9, section="Standalone Downside")],
                  concepts, source)
    check("blocchi di scenario diversi non collidono",
          len(two.admitted) == 2 and not two.violations, two.status)
    same = resolve([Binding("C-REV", "S!C5", "2026-06-30", "SD", unit="$mm",
                            confidence=0.9, section="Standalone Base"),
                    Binding("C-REV", "S!C9", "2026-06-30", "SD", unit="$mm",
                            confidence=0.9, section="Standalone Base")],
                   concepts, source)
    check("nello stesso blocco resta un conflitto",
          any(v.code == "UNIQUE_BINDING" for v in same.violations))


# ── 8. Sheet classifier: stability by construction ───────────────────────────
# The model answers two questions layout cannot: what kind of sheet this is, and
# which headers are periods. Stability must not depend on the model behaving —
# a fingerprint that has been answered is never asked again.

def test_classifier() -> None:
    print("\n8. Classificatore dei fogli: stabilità per costruzione")
    from openpyxl import Workbook
    from tools.sheet_classifier import (fingerprint_sheet, veto, SHEET_KINDS,
                                        save_cache, load_cache)

    def sheet(vals):
        wb = Workbook(); ws = wb.active; ws.title = "M"
        for ref, v in vals.items():
            ws[ref] = v
        return ws

    base = {"A1": "Line", "B1": "FY2025A", "A2": "Revenue", "B2": 74.0,
            "A3": "EBITDA", "B3": 11.4}
    ws = sheet(base)
    fp_a, _ = fingerprint_sheet(ws)
    check("impronta stabile fra due letture", fp_a == fingerprint_sheet(ws)[0])

    # numbers are shape, not content: the cache must survive a re-forecast
    ws2 = sheet({**base, "B2": 999.0, "B3": 222.0})
    check("numeri diversi non invalidano la cache",
          fingerprint_sheet(ws2)[0] == fp_a)

    # a renamed line item is a different question
    ws3 = sheet({**base, "A2": "Turnover"})
    check("etichetta cambiata => impronta diversa",
          fingerprint_sheet(ws3)[0] != fp_a)

    # the veto: structure overrules the model, never the other way round
    check("model_sheet senza header viene respinto",
          veto("model_sheet", {"dims": [2, 2], "grid": ["nn", "nn"]})[0] == "unknown")
    check("record_table troppo corta viene respinta",
          veto("record_table", {"dims": [3, 4], "grid": ["tttt"]})[0] == "unknown")
    check("un tipo fuori enum viene respinto",
          veto("qualsiasi_cosa", {"dims": [9, 4], "grid": ["tttt"]})[0] == "unknown")
    check("un model_sheet ben formato passa",
          veto("model_sheet", {"dims": [3, 3], "grid": ["ttt", "tnf"]})[0] == "model_sheet")
    check("'unknown' è fra i tipi ammessi", "unknown" in SHEET_KINDS)

    # deal scoping is optional and must not invalidate what came before it
    fp_ks, _ = fingerprint_sheet(ws, deal="keystone")
    fp_as, _ = fingerprint_sheet(ws, deal="astrelia")
    check("senza slug l'impronta resta quella di prima",
          fingerprint_sheet(ws)[0] == fp_a)
    check("lo slug cambia l'impronta", fp_ks != fp_a)
    check("due deal non condividono il giudizio", fp_ks != fp_as)
    check("lo stesso deal e' stabile", fp_ks == fingerprint_sheet(ws, deal="keystone")[0])

    import tempfile
    with tempfile.TemporaryDirectory() as td:
        cp = Path(td) / "c.json"
        save_cache({fp_a: {"sheet": "M", "kind": "model_sheet",
                           "period_headers": {"P1": "sequence"}, "reason": "r"}}, cp)
        back = load_cache(cp)
        check("la cache sopravvive al round-trip",
              back.get(fp_a, {}).get("kind") == "model_sheet")

    # a judgment must lift the pessimistic default, never demote a positive one
    from tools.sheet_semantics import analyse_sheet
    fr = sheet({"A1": "Line", "B1": "P1", "C1": "P2",
                "A2": "Besoin brut", "B2": 10, "C2": 12,
                "A3": "Stock", "B3": 5, "C3": 6})
    without = analyse_sheet(fr)
    with_j = analyse_sheet(fr, {"kind": "model_sheet",
                                "period_headers": {"P1": "sequence", "P2": "sequence"}})
    check("senza giudizio il vocabolario ignoto non è un periodo",
          all(p.period_kind is None for p in without.proposals))
    check("col giudizio P1/P2 diventano periodi",
          any(p.period_kind == "sequence" for p in with_j.proposals))
    check("il giudizio alza la confidenza invece di abbassarla",
          max((p.confidence for p in with_j.proposals), default=0) >=
          max((p.confidence for p in without.proposals), default=0))

    # A record table identifies its values by (record key, field). Without the
    # key a cell is only "some Gross MOIC" and cannot be bound to anything.
    rt = sheet({"A1": "Case", "B1": "Sponsor Invested", "C1": "Gross MOIC",
                "A2": "Base", "B2": 62, "C2": 1.99,
                "A3": "Downside", "B3": 62, "C3": 1.28,
                "A4": "Upside", "B4": 62, "C4": 2.43})
    rep = analyse_sheet(rt, {"kind": "record_table"})
    by_cell = {p.cell.split("!")[1]: p for p in rep.proposals}
    check("la tabella di record porta la chiave di riga",
          by_cell["C2"].record_key == "Base" and by_cell["C3"].record_key == "Downside")
    check("la chiave non viene scambiata per un'etichetta di riga",
          all(not p.row_label for p in rep.proposals))
    check("chiave + campo identificano la cella: proposta sicura",
          not by_cell["C2"].needs_review)
    check("l'unità viene dall'intestazione di campo, non dalla riga",
          by_cell["C2"].unit == "x")
    check("la chiave è nell'evidenza",
          "A2" in by_cell["C2"].evidence and "C1" in by_cell["C2"].evidence)


# ── 9. The live store and the projection built from it ───────────────────────
# The projection served to the UI used to blend a bundle, the vault and the
# package fixture, so "is this real?" had to be answered section by section.
# These pin the property that replaced that: what is not in the store is not on
# screen, and an empty store produces an empty deal rather than a demo one.

def test_live() -> None:
    print("\n9. Store live: a schermo solo ciò che è stato estratto")
    import tempfile
    from tools import live_store as ls
    from tools import live_projection as lp

    with tempfile.TemporaryDirectory() as td:
        original, ls.LIVE = ls.LIVE, Path(td)
        try:
            st = ls.LiveStore("t")
            check("uno store nuovo è vuoto", st.is_empty)

            proj = lp.build(st)
            check("store vuoto => nessuno scenario",
                  "scenarioLab" in proj["deal"] and
                  not proj["deal"]["scenarioLab"]["scenarios"])
            check("store vuoto => nessuna room piena",
                  not proj["deal"]["rooms"]["foundations"]["sets"])
            check("store vuoto => ogni schermata è dichiarata assente",
                  len(proj["absent_views"]) == len(lp.SCREEN_VIEWS))
            check("lo scheletro ha tutte le chiavi che la UI legge",
                  all(k in proj["deal"] for k in _skel_keys()))

            src = Path(td) / "wb.xlsx"
            src.write_bytes(b"not really a workbook, only its digest matters here")
            st.add_workbook(src, {
                "L1_source_graph": {"cells": 12},
                "bindings": [{"concept_id": "C-GROSS-MOIC", "locator": "R!K4",
                              "period": "", "scenario": "Base", "section": "",
                              "unit": "x", "confidence": 0.8, "value": 2.0}],
                "records": [{"record": "Base", "fields": {
                    "moic": {"concept_id": "C-GROSS-MOIC", "value": 2.0,
                             "unit": "x", "locator": "R!K4"},
                    "irr": {"concept_id": "C-GROSS-XIRR", "value": 0.15,
                            "unit": "%", "locator": "R!L4"}}}],
            })
            proj = lp.build(st)
            lab = proj["deal"]["scenarioLab"]
            check("un record diventa uno scenario", len(lab["scenarios"]) == 1)
            s = lab["scenarios"][0]
            check("il MOIC arriva intatto", s["moic"] == 2.0)
            check("l'IRR è convertito in punti percentuali una volta sola",
                  s["irr"] == 15.0)
            check("ogni cifra porta la cella da cui viene",
                  all(f["locator"] for f in s["fields"]))
            check("lo scenario non è più dichiarato assente",
                  "scenario" not in proj["absent_views"])
            check("le schermate che nessun ingest alimenta restano assenti",
                  "foundations" in proj["absent_views"])

            # A second ingest of the same workbook must not double the case.
            st.add_workbook(src, {"L1_source_graph": {"cells": 12},
                                  "bindings": [], "records": [
                {"record": "Base", "fields": {"moic": {
                    "concept_id": "C-GROSS-MOIC", "value": 2.5, "unit": "x",
                    "locator": "R!K4"}}}]})
            proj = lp.build(st)
            check("re-ingerire lo stesso caso lo sostituisce, non lo duplica",
                  len(proj["deal"]["scenarioLab"]["scenarios"]) == 1 and
                  proj["deal"]["scenarioLab"]["scenarios"][0]["moic"] == 2.5)
            check("la stessa sorgente resta una riga sola",
                  len(st.manifest["sources"]) == 1)

            st.reset()
            check("reset riporta lo store a vuoto", ls.LiveStore("t").is_empty)
        finally:
            ls.LIVE = original


# ── 10. The gate on claims whose locator it cannot resolve ───────────────────
# extract_v2 cites 'p3:w250-291'; the gate resolved locators by filename only,
# so every PDF-extracted claim produced no source text — and the failure to look
# was reported under ENTITY_NOT_IN_SOURCE, the code for having looked and found
# nothing. G4 and G5 silently did not run on 31 of 31 claims.

def test_gate_sources() -> None:
    print("\n10. Gate: 'non ho potuto controllare' non è 'ho controllato'")
    prof = DealProfile("t", {
        "entity_aliases": ["Alderstone"],
        "counterparty_entities": {"Riverton": "customer"},
        "perimeter_vocabulary": ["Alderstone standalone"],
    })

    claim = {"claim_id": "c1", "statement": "Alderstone FY2025 EBITDA is $12.7m.",
             "perimeter": "Alderstone standalone", "locator": "p3:w250-291",
             "value": 12.7, "source_doc": "story.pdf"}

    codes = {f["code"] for f in gg.check_claim(claim, prof)}
    check("locator irrisolvibile => SOURCE_NOT_RESOLVED, non ENTITY_NOT_IN_SOURCE",
          "SOURCE_NOT_RESOLVED" in codes and "ENTITY_NOT_IN_SOURCE" not in codes)
    check("un claim non controllabile non è bloccante",
          not any(f["blocking"] for f in gg.check_claim(claim, prof)))

    src = {"story.pdf": "Alderstone FY2025 adjusted EBITDA is $12.7m."}
    codes = {f["code"] for f in gg.check_claim(claim, prof, src)}
    check("col testo della sorgente il controllo gira davvero", not codes)

    bad = {**claim, "value": 99.9}
    codes = {f["code"] for f in gg.check_claim(bad, prof, src)}
    check("un valore assente dalla sorgente ora viene visto",
          "VALUE_NOT_IN_SOURCE" in codes)

    # G2 split: a described scope and an unidentifiable one ask different
    # questions, and reporting both the same way made the queue uninformative.
    anch = {**claim, "perimeter": "Alderstone consolidated revenue"}
    check("perimetro ancorato a un'entità nota => OFF_VOCABULARY",
          "PERIMETER_OFF_VOCABULARY" in
          {f["code"] for f in gg.check_claim(anch, prof, src)})
    unan = {**claim, "perimeter": "Days sales outstanding"}
    check("perimetro che non nomina nulla di noto => UNANCHORED",
          "PERIMETER_UNANCHORED" in
          {f["code"] for f in gg.check_claim(unan, prof, src)})
    cp = {**claim, "perimeter": "Riverton account revenue"}
    check("un perimetro che nomina una controparte resta bloccante",
          any(f["blocking"] for f in gg.check_claim(cp, prof, src)))


# ── 11. The cell engine: what-if, range edges, honest verification ───────────
# L1 records a SUM's precedent as the range it sums. Walking precedents then
# reaches the string "SHEET!C17:C19", finds no cell with that key, and stops —
# so every dependency path through an aggregate was invisible. On this workbook
# expanding ranges recovered 6,910 edges and took one input's reach from 900
# cells to 4,280.

def test_cell_engine() -> None:
    print("\n11. Motore delle celle: what-if, archi dei range, verifica onesta")
    import tempfile
    from tools.cell_engine import CellEngine, expand_range

    check("un range si espande nelle sue celle",
          expand_range("S!C17:C19") == ["S!C17", "S!C18", "S!C19"])
    check("i simboli di ancoraggio non contano",
          expand_range("S!$C$17:$D$18") == ["S!C17", "S!C18", "S!D17", "S!D18"])
    check("una cella singola non è un range", expand_range("S!C17") is None)
    check("una colonna intera non viene espansa",
          expand_range("S!A1:XFD1048576") is None)

    import openpyxl
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "m.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "M"
        ws["A1"], ws["A2"], ws["A3"] = 10, 20, 30      # inputs
        ws["B1"] = "=SUM(A1:A3)"                        # the aggregate that broke
        ws["B2"] = "=B1*2"                              # only reachable through it
        ws["C1"] = 99                                   # nobody computes, nobody reads
        wb.save(p)

        eng = CellEngine.build(p, verify=True)

        check("la catena attraversa l'aggregato",
              "M!B2" in eng.dependents("M!A1"),
              f"{sorted(eng.dependents('M!A1'))}")
        check("i precedenti transitivi risalgono attraverso il range",
              {"M!A1", "M!A2", "M!A3", "M!B1"} <= eng.precedents("M!B2"))

        r = eng.reachability("M!B2")
        check("un derivato è raggiungibile e nomina i suoi driver",
              r.derivable and set(r.drivers) == {"M!A1", "M!A2", "M!A3"})
        check("un input è settabile ma non derivabile",
              eng.reachability("M!A1").role == "input"
              and not eng.reachability("M!A1").derivable)
        check("un numero che nessuno calcola e nessuno legge è isolato",
              eng.reachability("M!C1").role == "isolated")
        check("una cella inesistente non è raggiungibile",
              eng.reachability("M!Z9").role == "missing")

        res = eng.set_value("M!A1", 40, watch=["M!B1", "M!B2", "M!C1"])
        check("cambiare un input propaga attraverso la SUM",
              res.changed.get("M!B1", (None, None))[1] == 90
              and res.changed.get("M!B2", (None, None))[1] == 180,
              str(res.changed))
        check("una cella fuori dal cono non viene toccata",
              "M!C1" not in res.changed and any("M!C1" in u for u in res.unreachable))

        eng.reset()
        check("reset riporta il workbook ai valori originali",
              eng.value("M!B1") == 60 and not eng.overrides)

        res = eng.set_value("M!B1", 999)
        check("scrivere su una cella con formula viene rifiutato",
              not res.changed and any("è una formula" in s for s in res.skipped))

        # This workbook was written by openpyxl, so it carries no cached values.
        check("senza valori in cache il verdetto non è 'verificato'",
              eng.verification.verdict == "UNVERIFIABLE_NO_CACHE",
              eng.verification.verdict)

        st = CellEngine.structure(p)
        check("in sola struttura la raggiungibilità funziona",
              st.reachability("M!B2").derivable)
        check("in sola struttura il ricalcolo si dichiara indisponibile",
              not st.can_evaluate
              and any("sola struttura" in s
                      for s in st.set_value("M!A1", 1).skipped))


def _skel_keys() -> list[str]:
    """The keys V17's views read off `deal` without checking they exist."""
    return ["objective", "branches", "morning_delta", "next_best_work",
            "command_suggestions", "rooms", "scenarioLab", "decisionRoom",
            "executionRoom", "replay", "registry", "case_id"]


def main() -> int:
    print("=" * 62)
    print("REGRESSION SUITE — 2026-08-25 audit")
    print("=" * 62)
    tests = (
        (test_cascade, None),
        (test_benchmark_field, None),
        (test_deal_profile, None),
        (test_grounding_gate, None),
        (test_minigraph, None),
        (test_excel, None),
        (test_resolver, ROOT / "tools" / "binding_resolver.py"),
        (test_classifier, None),
        (test_live, None),
        (test_gate_sources, None),
        (test_cell_engine, ROOT / "tools" / "cell_engine.py"),
    )
    for test, private_dependency in tests:
        if private_dependency is not None and not private_dependency.exists():
            skip(test.__name__, f"private optional module not installed: {private_dependency.name}")
            continue
        test()
    print()
    print("=" * 62)
    print(f"  {len(PASS)} passed, {len(FAIL)} failed, {len(SKIP)} skipped")
    if FAIL:
        for f in FAIL:
            print(f"    FAILED: {f}")
    print("=" * 62)
    return 1 if FAIL else 0


if __name__ == "__main__":
    raise SystemExit(main())
