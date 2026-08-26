#!/usr/bin/env python3
"""
sheet_semantics — infer what a cell *means*, with the evidence for it.

extract_v3 gives a computable graph of cells. `CASHFLOW!B3` is not a concept
though: nothing downstream can bind it to quarterly interest expense. This
module builds that bridge.

Design rule
-----------
Deterministic first, model only where there is genuine ambiguity. Everything
here is deterministic: labels come from the sheet's own text, units from number
formats and label suffixes, periods from column headers. Each proposal carries
the cells its evidence came from, so a human can check the reasoning rather than
the conclusion — the same stance as the grounding gate on claims.

What this does NOT do is decide. A proposal below the confidence floor is
routed to review, not silently bound. An LLM pass over the low-confidence
remainder is the intended next layer, not a replacement for this one.

Layout assumption, stated openly
--------------------------------
Financial models are overwhelmingly laid out as a label column on the left and
a period header row on top. That is an assumption, not a law, so
`infer_orientation` measures it per sheet and reports what it found instead of
taking it for granted.

    python3 tools/sheet_semantics.py --workbook model.xlsx
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass, field, asdict
from datetime import datetime, date
from pathlib import Path
from typing import Any

# ── period recognition ───────────────────────────────────────────────────────

_PERIOD_PATTERNS = [
    (re.compile(r"^FY\s?(\d{4})\s?([AEF])?$", re.I),        "fiscal_year"),
    (re.compile(r"^(\d{4})\s?([AEF])$", re.I),              "fiscal_year"),
    (re.compile(r"^Q([1-4])\s?(FY)?\s?(\d{2,4})?$", re.I),  "quarter"),
    (re.compile(r"^(LTM|NTM)\b", re.I),                     "trailing"),
    (re.compile(r"^\d{4}-\d{2}-\d{2}$"),                    "date"),
    (re.compile(r"^(Opening|Closing|Entry|Exit)$", re.I),   "anchor"),
]


def classify_period(text: str) -> str | None:
    t = (text or "").strip()
    if not t:
        return None
    for pattern, kind in _PERIOD_PATTERNS:
        if pattern.match(t):
            return kind
    return None


# ── unit recognition ─────────────────────────────────────────────────────────

_UNIT_FROM_LABEL = [
    (re.compile(r"\bmultiple\b|\(x\)|\bx\b\s*$", re.I), "x"),
    (re.compile(r"\brate\b|\bmargin\b|\bgrowth\b|%", re.I), "%"),
    (re.compile(r"\bdays\b|\bDSO\b|\bDPO\b", re.I), "days"),
    (re.compile(r"\bIRR\b", re.I), "%"),
    (re.compile(r"\bMOIC\b", re.I), "x"),
]


def infer_unit(label: str, number_format: str | None) -> tuple[str, str]:
    """(unit, where it came from). Number format beats label wording."""
    fmt = (number_format or "").lower()
    if "%" in fmt:
        return "%", "number_format"
    if "$" in fmt or "usd" in fmt:
        return "$", "number_format"
    if "€" in fmt or "eur" in fmt:
        return "€", "number_format"
    for pattern, unit in _UNIT_FROM_LABEL:
        if pattern.search(label or ""):
            return unit, "label"
    return "", "unknown"


# ── proposals ────────────────────────────────────────────────────────────────

@dataclass
class ConceptProposal:
    cell: str                      # SHEET!B3
    concept: str                   # human-readable name
    row_label: str = ""
    col_header: str = ""
    unit: str = ""
    unit_source: str = ""
    period_kind: str | None = None
    is_formula: bool = False
    confidence: float = 0.0
    evidence: list[str] = field(default_factory=list)   # cells the label came from
    issues: list[str] = field(default_factory=list)

    @property
    def needs_review(self) -> bool:
        return self.confidence < 0.6 or bool(self.issues)


@dataclass
class SheetReport:
    sheet: str
    orientation: str                      # labels_left | labels_top | unclear
    label_column: str | None = None
    header_row: int | None = None
    kind: str = "unknown"          # model_sheet | record_table | unknown
    proposals: list[ConceptProposal] = field(default_factory=list)

    @property
    def confident(self) -> list[ConceptProposal]:
        return [p for p in self.proposals if not p.needs_review]

    @property
    def review(self) -> list[ConceptProposal]:
        return [p for p in self.proposals if p.needs_review]


# ── inference ────────────────────────────────────────────────────────────────

def _is_text(v: Any) -> bool:
    return isinstance(v, str) and bool(v.strip()) and not v.startswith("=")


# Models date their columns with real dates, not strings, so a header row of
# datetimes was being skipped entirely by a text-only scan.
def _is_header_like(v: Any) -> bool:
    from datetime import datetime, date
    return _is_text(v) or isinstance(v, (datetime, date))


def _header_text(v: Any) -> str:
    from datetime import datetime, date
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


# A column that holds units for the whole sheet ($mm, %, x, days) is not a
# label column. Locally "$mm" and "Revenue" are both text to the left of a
# value; only this sheet-wide fact separates them.
_UNIT_TOKEN = re.compile(r"^\s*(\$?(mm|m|k|bn)|%|x|days|ratio|units?|#)\s*$", re.I)


def find_unit_columns(ws, probe_rows: int = 60) -> set[int]:
    max_r = min(ws.max_row, probe_rows)
    max_c = min(ws.max_column, 20)
    found = set()
    for c in range(1, max_c + 1):
        vals = [ws.cell(row=r, column=c).value for r in range(1, max_r + 1)]
        texts = [v for v in vals if _is_text(v)]
        if len(texts) >= 3 and sum(1 for t in texts if _UNIT_TOKEN.match(t)) / len(texts) >= 0.6:
            found.add(c)
    return found


def infer_orientation(ws, probe_rows: int = 30, probe_cols: int = 20) -> tuple[str, str | None, int | None]:
    """
    Measure where the labels are instead of assuming.

    A label column is a column that is mostly text while the columns to its
    right are mostly numeric; a header row is the mirror image.
    """
    max_r = min(ws.max_row, probe_rows)
    max_c = min(ws.max_column, probe_cols)
    if max_r < 2 or max_c < 2:
        return "unclear", None, None

    # A title cell is text and nothing else; a header spans the sheet. Requiring
    # coverage as well as text share is what separates "Project Keystone" in A1
    # from the real dated header on row 3.
    min_span = max(3, int(max_c * 0.5))

    def score_col(c: int) -> tuple[float, int]:
        vals = [ws.cell(row=r, column=c).value for r in range(1, max_r + 1)]
        seen = [v for v in vals if v is not None]
        share = (sum(1 for v in seen if _is_text(v)) / len(seen)) if seen else 0.0
        return share, len(seen)

    def score_row(r: int) -> tuple[float, int]:
        vals = [ws.cell(row=r, column=c).value for c in range(1, max_c + 1)]
        seen = [v for v in vals if v is not None]
        share = (sum(1 for v in seen if _is_header_like(v)) / len(seen)) if seen else 0.0
        return share, len(seen)

    label_col = next((c for c in range(1, max_c + 1)
                      if score_col(c)[0] >= 0.7 and score_col(c)[1] >= 3), None)
    header_row = next((r for r in range(1, max_r + 1)
                       if score_row(r)[0] >= 0.7 and score_row(r)[1] >= min_span), None)

    from openpyxl.utils import get_column_letter
    if label_col and header_row:
        return "labels_left", get_column_letter(label_col), header_row
    if label_col:
        return "labels_left", get_column_letter(label_col), None
    if header_row:
        return "labels_top", None, header_row
    return "unclear", None, None


def _row_is_blank(ws, row: int, max_col: int) -> bool:
    return all(ws.cell(row=row, column=c).value is None
               for c in range(1, max_col + 1))


def _col_is_blank(ws, col: int, max_row: int) -> bool:
    return all(ws.cell(row=r, column=col).value is None
               for r in range(1, max_row + 1))


def _scan_left(ws, row: int, col: int, limit: int = 12,
               unit_cols: set[int] | None = None) -> tuple[str, str]:
    """Nearest label to the left, skipping unit columns, stopping at the block."""
    max_row = min(ws.max_row, 200)
    unit_cols = unit_cols or set()
    for c in range(col - 1, max(1, col - limit) - 1, -1):
        if c in unit_cols:
            continue
        # A fully empty column separates blocks: a label beyond it belongs to
        # a different table and must not be borrowed.
        if _col_is_blank(ws, c, max_row):
            return "", ""
        v = ws.cell(row=row, column=c).value
        if _is_text(v):
            return v.strip(), ws.cell(row=row, column=c).coordinate
    return "", ""


def _scan_up(ws, row: int, col: int, limit: int = 12) -> tuple[str, str]:
    """
    Nearest header above, stopping at the block boundary.

    Without this the scan crosses a blank row into the previous block and
    attaches, say, FY2024A to a covenant table that carries no period at all —
    a wrong answer delivered at high confidence, which is worse than none.
    """
    max_col = min(ws.max_column, 200)
    for r in range(row - 1, max(1, row - limit) - 1, -1):
        if _row_is_blank(ws, r, max_col):
            return "", ""
        v = ws.cell(row=r, column=col).value
        if _is_header_like(v):
            return _header_text(v), ws.cell(row=r, column=col).coordinate
    return "", ""


def classify_sheet(ws, label_col: int | None, header_row: int | None,
                   probe_cols: int = 24) -> str:
    """
    Tell a model sheet from a record table by what the columns are.

    In a model the columns are periods: SB_Base runs quarter-end dates across
    row 3. In a record table they are fields: Employees runs Name, Location,
    Hire date. That is the distinction that matters downstream, because a
    period column binds to a concept over time while a field column does not.

    Uniqueness of the label column was tried first and does not survive contact
    with real workbooks: an employee roster has perfectly unique ids in column A
    and is emphatically not a model, while Scenario_Drivers repeats "Revenue"
    once per scenario block and emphatically is one.
    """
    if not header_row:
        return "unknown"
    headers = [ws.cell(row=header_row, column=c).value
               for c in range(1, min(ws.max_column, probe_cols) + 1)]
    seen = [h for h in headers if h is not None]
    if len(seen) < 3:
        return "unknown"
    periodish = sum(1 for h in seen
                    if isinstance(h, (datetime, date)) or classify_period(str(h)))
    if periodish / len(seen) >= 0.5:
        return "model_sheet"

    # Everything else is treated as a record table, which discards the
    # left-hand text rather than trusting it.
    #
    # This is deliberately the pessimistic branch. Four classifiers were tried
    # and each traded one error for another: label-column uniqueness calls an
    # employee roster a model because its ids are unique; requiring positive
    # evidence to demote lets a flat data table come back 100% confident with
    # field values passing as concept labels. Only this direction fails by
    # losing coverage instead of inventing meaning.
    #
    # The cost is real and measured: a French model heading its columns P1, P2
    # lands here and drops to zero, because period vocabulary is unbounded and
    # this recognises only what it has met. That is the boundary of what layout
    # alone can decide.
    return "record_table"


def analyse_sheet(ws, judgment: dict | None = None) -> SheetReport:
    orientation, label_col, header_row = infer_orientation(ws)
    unit_cols = find_unit_columns(ws)
    from openpyxl.utils import column_index_from_string
    label_idx = column_index_from_string(label_col) if label_col else None
    kind = classify_sheet(ws, label_idx, header_row)
    # A cached model judgment resolves what layout alone cannot: whether this is
    # a model or a table of records, and which headers are periods. It only ever
    # replaces "record_table", the pessimistic default — it is not allowed to
    # demote a sheet the deterministic layer positively recognised.
    model_periods: dict[str, str] = {}
    if judgment:
        model_periods = judgment.get("period_headers") or {}
        jk = judgment.get("kind")
        if jk in ("model_sheet", "record_table", "scalar_block") and kind == "record_table":
            kind = jk
    report = SheetReport(sheet=ws.title.upper(), orientation=orientation,
                         label_column=label_col, header_row=header_row, kind=kind)

    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v is None:
                continue
            is_formula = isinstance(v, str) and v.startswith("=")
            # A cell that holds only text is a label, not a quantity.
            if _is_text(v) and not is_formula:
                continue

            if cell.column in unit_cols:
                continue      # a unit cell is metadata about its row, not a quantity
            if header_row and cell.row <= header_row:
                continue      # the header row describes the sheet, it is not data
            row_label, row_src = _scan_left(ws, cell.row, cell.column,
                                            unit_cols=unit_cols)
            declared_unit = ""
            for uc in sorted(unit_cols, reverse=True):
                if uc < cell.column:
                    uv = ws.cell(row=cell.row, column=uc).value
                    if _is_text(uv):
                        declared_unit = uv.strip()
                        break
            col_header, col_src = _scan_up(ws, cell.row, cell.column)
            if not col_header and header_row:
                # A sheet-wide header row governs every block beneath it. The
                # local upward scan stops at the blank rows that merely space
                # sections apart, which is right between two tables and wrong
                # here — this is the sheet-level fact that settles it.
                hv = ws.cell(row=header_row, column=cell.column).value
                if _is_header_like(hv):
                    col_header = _header_text(hv)
                    col_src = ws.cell(row=header_row, column=cell.column).coordinate

            evidence = [c for c in (row_src, col_src) if c]
            issues: list[str] = []
            if kind == "record_table" and row_label:
                # The column header names the field; the value to the left is
                # another field of the same record, not a label for this one.
                row_label, row_src = "", ""
                evidence = [c for c in (col_src,) if c]
                issues.append("tabella di record: la voce a sinistra è un altro campo, "
                              "non un'etichetta")
            if not row_label and not col_header:
                issues.append("nessuna etichetta trovata né a sinistra né sopra")

            period_kind = classify_period(col_header)
            if period_kind is None and col_header in model_periods:
                # vocabulary the parser has never met — P1, T1, "Anno 1"
                period_kind = model_periods[col_header]
            if declared_unit:
                # The sheet says what the unit is; nothing should override that.
                unit, unit_source = declared_unit, "unit_column"
            else:
                unit, unit_source = infer_unit(row_label, cell.number_format)

            # Confidence is built from what was actually found, so a reader can
            # see why a proposal is weak rather than trusting a bare number.
            confidence = 0.0
            if row_label:
                confidence += 0.55
            if col_header:
                confidence += 0.25
            if period_kind:
                confidence += 0.10
            if unit_source in ("number_format", "unit_column"):
                confidence += 0.10
            confidence = round(min(confidence, 1.0), 2)

            concept = " · ".join(p for p in (row_label, col_header) if p) or "?"
            report.proposals.append(ConceptProposal(
                cell=f"{ws.title.upper()}!{cell.coordinate}",
                concept=concept, row_label=row_label, col_header=col_header,
                unit=unit, unit_source=unit_source, period_kind=period_kind,
                is_formula=is_formula, confidence=confidence,
                evidence=evidence, issues=issues,
            ))
    return report


def load_judgments(cache_path: Path | None) -> dict[str, dict]:
    """Judgments keyed by sheet name, from the classifier's durable cache."""
    if not cache_path or not cache_path.exists():
        return {}
    try:
        js = json.loads(cache_path.read_text(encoding="utf-8")).get("judgments", {})
    except Exception:
        return {}
    return {v.get("sheet", "").upper(): v for v in js.values() if v.get("sheet")}


def analyse_workbook(path: Path, judgments: dict[str, dict] | None = None) -> list[SheetReport]:
    try:
        import openpyxl
    except ImportError:
        sys.exit("serve openpyxl")
    # data_only=False so formula cells are visible as formulas, not stale values.
    wb = openpyxl.load_workbook(str(path), data_only=False)
    judgments = judgments or {}
    return [analyse_sheet(ws, judgments.get(ws.title.upper())) for ws in wb]


def main() -> int:
    ap = argparse.ArgumentParser(description="Infer cell meaning from sheet layout")
    ap.add_argument("--workbook", type=Path, required=True)
    ap.add_argument("--out", type=Path)
    ap.add_argument("--show", type=int, default=6, help="righe da mostrare per foglio")
    ap.add_argument("--judgments", type=Path, default=None,
                    help="cache di sheet_classifier (vault/policy/sheet_classifications.json)")
    a = ap.parse_args()

    reports = analyse_workbook(a.workbook, load_judgments(a.judgments))
    total = sum(len(r.proposals) for r in reports)
    conf = sum(len(r.confident) for r in reports)

    print(f"[sheet_semantics] {a.workbook.name}")
    for r in reports:
        print(f"\n  {r.sheet}  ({r.kind}, {r.orientation}"
              f"{', etichette in ' + r.label_column if r.label_column else ''}"
              f"{', header riga ' + str(r.header_row) if r.header_row else ''})")
        print(f"    proposte {len(r.proposals)} · sicure {len(r.confident)} · da rivedere {len(r.review)}")
        for p in r.proposals[:a.show]:
            mark = " " if not p.needs_review else "?"
            unit = f" [{p.unit}]" if p.unit else ""
            per = f" ({p.period_kind})" if p.period_kind else ""
            print(f"      {mark} {p.cell:16} {p.concept}{unit}{per}  conf={p.confidence}")
        if len(r.proposals) > a.show:
            print(f"        … altre {len(r.proposals) - a.show}")

    print(f"\n  totale: {conf}/{total} sicure "
          f"({round(100 * conf / total) if total else 0}%), "
          f"{total - conf} in coda di revisione")

    if a.out:
        a.out.mkdir(parents=True, exist_ok=True)
        payload = [{"sheet": r.sheet, "orientation": r.orientation,
                    "label_column": r.label_column, "header_row": r.header_row,
                    "proposals": [asdict(p) for p in r.proposals]} for r in reports]
        (a.out / "cell_semantics.json").write_text(
            json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"  → {a.out / 'cell_semantics.json'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
