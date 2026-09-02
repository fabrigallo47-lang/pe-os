#!/usr/bin/env python3
"""
extract_v2.py — E3 extraction pipeline for PANTA.

Architecture:
  L1  Document parser    → deterministic chunks with stable locators
  L2  LLM annotator      → one call per chunk, tool_use schema-constrained
  L3  Claim validator    → deterministic normalization + stable_id
  L4  Graph assembler    → deterministic dedup + conflict detection

Claim schema aligned to PANTA_Keystone_Canonical_Investment_Case_v1.1.json
(CAP-003). Does NOT touch extract.py.

Manifest modes prevent temporal leakage between knowledge snapshots:
  K-PRE   CIM, DR, IA, QL, QoE  — state of knowledge before IC memo
  K-IC    K-PRE + IC memo        — as-of IC gate (2026-03-10)
  K-LIVE  K-IC + board packs     — post-close, event-by-event

Usage:
  python3 tools/extract_v2.py --manifest K-IC --deal keystone --dry-run

  export ANTHROPIC_API_KEY=sk-...  # or OPENROUTER_API_KEY with PEOS_LLM_PROVIDER=openrouter
  python3 tools/extract_v2.py --manifest K-PRE --deal keystone \\
      --output pipeline_out/e3/k_pre

  python3 tools/extract_v2.py --source vault/inbox/keystone_ic_memo.md \\
      --deal keystone --output pipeline_out/e3/single

  python3 tools/extract_v2.py --manifest K-IC --deal keystone \\
      --compare pipeline_out/keystone_full_story/graph.json
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import mailbox
import os
import re
import sys
import textwrap
import time
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from email import policy as email_policy
from email.parser import BytesParser
from email.utils import parsedate_to_datetime
from html.parser import HTMLParser
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from tools.llm_provider import (  # noqa: E402
    anthropic_client_kwargs,
    configured_api_key,
    configured_model,
    missing_key_message,
    openrouter_extra_body,
)
from tools.archetype_pack import load_pack, workstream_ids  # noqa: E402
from tools.object_identity import claim_id as canonical_claim_id  # noqa: E402
from tools.source_envelope import extractor_source_record  # noqa: E402
from tools.source_capabilities import (  # noqa: E402
    CAPABILITY_SCHEMA,
    capability_failure,
    resolve_source_capability,
)

VAULT_INBOX = ROOT / "vault" / "inbox"
MODEL = configured_model("claude-haiku-4-5-20251001")
MAX_TOKENS = int(os.environ.get("PEOS_EXTRACT_V2_MAX_TOKENS", "4096"))
MAX_PROVIDER_RETRIES = int(os.environ.get("PEOS_LLM_CHUNK_RETRIES", "3"))

# ─────────────────────────────────────────────────────────────────────────────
# Source registry — maps vault/inbox filenames to canonical SRC-xxx IDs.
# Aligned with the sources dict in the benchmark CIC v1.1.
# known_at = when the Firm first received this document.
# ─────────────────────────────────────────────────────────────────────────────

SOURCE_REGISTRY: dict[str, dict] = {
    # filename (without extension or with) → canonical source record
    "keystone_seller_cim": {
        "source_id": "SRC-CIM",
        "name": "Project Keystone CIM",
        "party": "Alderstone management and Hawthorne Capital Markets",
        "doc_type": "CIM",
        "effective_date": "2025-10-27",
        "known_at": "2025-10-27",
        "manifest": ["K-PRE", "K-IC", "K-LIVE"],
    },
    "keystone_data_room_extract": {
        "source_id": "SRC-DR",
        "name": "Keystone data room extract",
        "party": "Alderstone management / VDR export",
        "doc_type": "Data Room",
        "effective_date": "2025-11-15",
        "known_at": "2025-11-15",
        "manifest": ["K-PRE", "K-IC", "K-LIVE"],
    },
    "keystone_firm_initial_assessment": {
        "source_id": "SRC-IA",
        "name": "Firm initial assessment",
        "party": "The Firm investment team",
        "doc_type": "Internal",
        "effective_date": "2025-12-01",
        "known_at": "2025-12-01",
        "manifest": ["K-PRE", "K-IC", "K-LIVE"],
    },
    "keystone_question_list": {
        "source_id": "SRC-QL",
        "name": "Keystone question list",
        "party": "The Firm diligence team",
        "doc_type": "Internal",
        "effective_date": "2026-01-10",
        "known_at": "2026-01-10",
        "manifest": ["K-PRE", "K-IC", "K-LIVE"],
    },
    "keystone_qoe_report": {
        "source_id": "SRC-QOE",
        "name": "Quality of Earnings report",
        "party": "Independent financial diligence provider",
        "doc_type": "QoE Report",
        "effective_date": "2026-02-20",
        "known_at": "2026-02-20",
        "manifest": ["K-PRE", "K-IC", "K-LIVE"],
    },
    "keystone_firm_model_summary": {
        "source_id": "SRC-MODEL-SUM",
        "name": "Firm model summary",
        "party": "The Firm investment team",
        "doc_type": "LBO Model",
        "effective_date": "2026-03-05",
        "known_at": "2026-03-05",
        "manifest": ["K-PRE", "K-IC", "K-LIVE"],
    },
    "keystone_lbo_model_working": {
        "source_id": "SRC-MODEL",
        "name": "Keystone LBO model",
        "party": "The Firm investment team",
        "doc_type": "LBO Model",
        "effective_date": "2026-03-05",
        "known_at": "2026-03-05",
        "manifest": ["K-PRE", "K-IC", "K-LIVE"],
    },
    "keystone-model-part1": {
        "source_id": "SRC-MODEL",
        "name": "Keystone LBO model (part 1)",
        "party": "The Firm investment team",
        "doc_type": "LBO Model",
        "effective_date": "2026-03-05",
        "known_at": "2026-03-05",
        "manifest": ["K-PRE", "K-IC", "K-LIVE"],
    },
    "keystone-model-part2": {
        "source_id": "SRC-MODEL",
        "name": "Keystone LBO model (part 2)",
        "party": "The Firm investment team",
        "doc_type": "LBO Model",
        "effective_date": "2026-03-05",
        "known_at": "2026-03-05",
        "manifest": ["K-PRE", "K-IC", "K-LIVE"],
    },
    "keystone-model-part3": {
        "source_id": "SRC-MODEL",
        "name": "Keystone LBO model (part 3)",
        "party": "The Firm investment team",
        "doc_type": "LBO Model",
        "effective_date": "2026-03-05",
        "known_at": "2026-03-05",
        "manifest": ["K-PRE", "K-IC", "K-LIVE"],
    },
    "keystone-model-part4": {
        "source_id": "SRC-MODEL",
        "name": "Keystone LBO model (part 4)",
        "party": "The Firm investment team",
        "doc_type": "LBO Model",
        "effective_date": "2026-03-05",
        "known_at": "2026-03-05",
        "manifest": ["K-PRE", "K-IC", "K-LIVE"],
    },
    "keystone_ic_memo": {
        "source_id": "SRC-IC",
        "name": "IC memo 2026-03-10",
        "party": "The Firm investment team / Investment Committee",
        "doc_type": "IC Memo",
        "effective_date": "2026-03-10",
        "known_at": "2026-03-10",
        "manifest": ["K-IC", "K-LIVE"],  # not in K-PRE
    },
    "keystone_monitoring_boardpack1_dec2026": {
        "source_id": "SRC-BP1",
        "name": "Board pack Dec 2026",
        "party": "Alderstone management / board package",
        "doc_type": "Board Pack",
        "effective_date": "2026-12-31",
        "known_at": "2026-12-31",
        "manifest": ["K-LIVE"],
    },
    "keystone_monitoring_boardpack2_mar2027": {
        "source_id": "SRC-BP2",
        "name": "Board pack Mar 2027",
        "party": "Alderstone management / board package",
        "doc_type": "Board Pack",
        "effective_date": "2027-03-31",
        "known_at": "2027-03-31",
        "manifest": ["K-LIVE"],
    },
    "keystone_monitoring_junecompliance2027": {
        "source_id": "SRC-JUN27",
        "name": "June 2027 compliance certificate",
        "party": "Alderstone management / board package",
        "doc_type": "Board Pack",
        "effective_date": "2027-06-30",
        "known_at": "2027-06-30",
        "manifest": ["K-LIVE"],
    },
    "keystone_monitoring_augustamendment2027": {
        "source_id": "SRC-AUG27",
        "name": "August 2027 amendment",
        "party": "Alderstone management / board package",
        "doc_type": "Amendment",
        "effective_date": "2027-08-15",
        "known_at": "2027-08-15",
        "manifest": ["K-LIVE"],
    },
    "keystone_monitoring_recovery_exit2031": {
        "source_id": "SRC-EXIT31",
        "name": "Exit 2031 package",
        "party": "Alderstone management / board package",
        "doc_type": "Board Pack",
        "effective_date": "2031-06-01",
        "known_at": "2031-06-01",
        "manifest": ["K-LIVE"],
    },
}

# Sources per manifest — derived from registry
MANIFEST_SOURCES: dict[str, list[str]] = {
    "K-PRE": [k for k, v in SOURCE_REGISTRY.items() if "K-PRE" in v["manifest"]],
    "K-IC": [k for k, v in SOURCE_REGISTRY.items() if "K-IC" in v["manifest"]],
    "K-LIVE": [k for k, v in SOURCE_REGISTRY.items() if "K-LIVE" in v["manifest"]],
    "ALL": list(SOURCE_REGISTRY.keys()),
}

def _source_record(path: Path) -> dict:
    """Return the source registry record for this file, or a synthetic fallback."""
    stem = path.stem
    # Try exact match first, then strip suffixes like _part1, _extract
    record = SOURCE_REGISTRY.get(stem)
    if not record:
        for key in SOURCE_REGISTRY:
            if stem.startswith(key) or key.startswith(stem):
                record = SOURCE_REGISTRY[key]
                break
    if not record:
        record = {
            "source_id": f"SRC-{stem.upper()[:12]}",
            "name": stem,
            "party": "unknown",
            "doc_type": "Other",
            "effective_date": "",
            "known_at": "",
            "manifest": ["ALL"],
        }
    return record


# ─────────────────────────────────────────────────────────────────────────────
# Schema definitions (bounded enums)
# ─────────────────────────────────────────────────────────────────────────────

METRIC_ENUM: list[str] = [
    "Revenue", "Recurring Revenue", "Revenue Growth",
    "Gross Profit", "Gross Margin",
    "EBITDA", "EBITDA Margin", "EBITDA Add-back", "EBITDA Adjustment",
    "EBIT", "Net Income", "Free Cash Flow", "Operating Cash Flow",
    "Capex", "Working Capital", "DSO", "DPO", "Inventory Days",
    "Earnings Quality Risk", "Revenue Quality", "Adjustment Supportability",
    "Customer Concentration", "Customer Count", "Active Billing Accounts",
    "Customer Retention", "Contract Terms", "Market Position", "Market Size",
    "Enterprise Value", "Equity Value", "Entry Multiple", "Exit Multiple", "Exit EV",
    "Net Debt", "Gross Debt", "Leverage", "Interest Coverage",
    "Sponsor Equity", "Seller Rollover", "First-Lien Debt",
    "Revolver Capacity", "DDTL Availability", "Covenant EBITDA",
    "Covenant Threshold", "Covenant Headroom",
    "MOIC", "IRR", "Exit Horizon", "Supported Price",
    "Net Working Capital", "Net Working Capital Target", "Net Working Capital Adjustment",
    "Headcount", "Team Tenure", "Acquisition Count",
    "Systems Integration Risk", "Integration Risk", "Operational Risk",
    "Key Person Risk", "Regulatory Risk", "Competition Risk",
    "IC Conditions", "IC Vote", "Decision Coherence",
    "Customer Contract Terms",
]
_seen: set[str] = set()
METRIC_ENUM = [m for m in METRIC_ENUM if not (m in _seen or _seen.add(m))]  # type: ignore

UNIT_ENUM: list[str | None] = [
    None, "", "$m", "£m", "€m", "$m/year", "$m/quarter",
    "%", "x", "bps", "days", "headcount", "turns", "$", "£", "€",
]

PERIOD_MAP: dict[str, str] = {
    "FY2025A": "2025-12-31", "FY2025": "2025-12-31",
    "FY2026E": "2026-12-31", "FY2026": "2026-12-31",
    "FY2027E": "2027-12-31", "FY2027": "2027-12-31",
    "FY2028E": "2028-12-31", "FY2028": "2028-12-31",
    "FY2029E": "2029-12-31", "FY2029": "2029-12-31",
    "FY2030E": "2030-12-31", "FY2030": "2030-12-31",
    "FY2031E": "2031-12-31", "FY2031": "2031-12-31",
    "OPENING": "2026-03-31",
    "LTM": "LTM",
    "Q1 2026": "2026-06-30", "Q2 2026": "2026-09-30",
    "Q3 2026": "2026-12-31", "Q4 2026": "2027-03-31",
    "Q1 2027": "2027-06-30", "Q2 2027": "2027-09-30",
}
# Period is now free-text — no enum constraint.
# PERIOD_MAP is kept for period_iso normalization only (not for constraining the field).

EPISTEMIC_CLASS_ENUM = ["asserted", "observed", "derived", "attested"]
DIRECTION_ENUM = ["supports", "contradicts", "context"]

# What kind of assertion a claim is. The distinction that matters is the last
# one: a sentence with no number earns a claim only if someone could later
# confirm or refute it. "No material litigation as of 30 June" is checkable;
# "a scaled regional platform" is not, and admitting it puts a seller's adjective
# into the case as though it were evidence.
#
# This criterion used to be implicit, and that is exactly where extraction
# wobbled: across runs of one paragraph the quantified claims were identical
# every time, while a valueless MarketPosition claim from the phrase "scaled
# regional platform" appeared in some runs and not others.
CLAIM_KIND_ENUM = [
    "QUANTITATIVE",     # carries a number or a measured quantity
    "DEFINITION",       # states how something is defined or calculated
    "CONDITION",        # states a requirement, contingency or covenant
    "ATTRIBUTION",      # states who said, did or decided something
    "NEGATIVE",         # explicitly asserts an absence — checkable, so admissible
    "CHARACTERISATION", # a descriptor with no checkable content — DO NOT EMIT
]

# How to read the number. "More than 600 accounts" recorded as 600 EXACT makes a
# later, entirely consistent "640" look like a contradiction. The bound is not
# part of identity — it describes the value, not the quantity being measured —
# but comparison must respect it.
BOUND_ENUM = ["EXACT", "AT_LEAST", "AT_MOST", "APPROXIMATE", "RANGE", "NONE"]
ARCHETYPE_PACK = load_pack()
# "Use workstream/concept families to select schemas. Category never creates identity or contradiction."
# Topic is extraction routing metadata only; identity and conflict logic deliberately ignore it.
TOPIC_ENUM = workstream_ids(ARCHETYPE_PACK) + ["OTHER"]


def _topic_description(pack: dict[str, Any]) -> str:
    """Give the model the pack's governing questions beside each route ID."""
    workstreams = pack["workstreams"]
    routes = "\n".join(
        f"- {workstream_id}: {workstreams[workstream_id]['governing_question']}"
        for workstream_id in workstream_ids(pack)
    )
    return (
        "Choose the workstream whose governing question best routes this claim; "
        "do not infer identity or contradiction from topic. Use OTHER only when no "
        "workstream applies.\nCanonical workstreams:\n"
        f"{routes}"
    )


TOPIC_DESCRIPTION = _topic_description(ARCHETYPE_PACK)

# ── Identity dimensions (bounded) ─────────────────────────────────────────────
# `perimeter` stays prose because a human reads it. These carry the machine
# identity instead: an exact-match table cannot normalize unbounded language, so
# the dimensions that decide whether two claims are comparable are asked for
# separately and constrained here. See tools/object_identity.py.
BASIS_ENUM = [
    "SellerView", "QoEView", "FirmView", "CovenantView", "ReportedView", "unspecified",
]
SCOPE_ENUM = ["consolidated", "standalone", "customer", "segment", "unspecified"]
SCENARIO_ENUM = ["base", "management", "seller", "upside", "downside", "unspecified"]

# Canonical period grammar. Anything outside it is "none" — an honest gap beats
# an extraction timestamp masquerading as a period (the failure that put
# "as of <ingest date>" on 227 vault claims).
PERIOD_CANONICAL_PATTERN = (
    r"^(FY\d{4}[AE]?|LTM|ExitLTM|Opening|EntryToExit|CrossPeriod"
    r"|\d{4}-\d{2}-\d{2}|none)$"
)

# Known definition IDs — if LLM flags a claim as definition-linked, map to DEF-xxx.
# These align with the definitions[] array in the canonical benchmark.
DEFINITION_ENUM: list[str | None] = [
    None,
    "DEF-FIRM-EBITDA", "DEF-QOE-EBITDA", "DEF-COV-EBITDA",
    "DEF-SELLER-EBITDA", "DEF-NWC", "DEF-NWC-TARGET",
    "DEF-RECURRING-REVENUE", "DEF-CONCENTRATION",
    "DEF-MOIC", "DEF-IRR", "DEF-EV",
]

# ─────────────────────────────────────────────────────────────────────────────
# Tool schema (tool_use — enforced at LLM decoding level)
# ─────────────────────────────────────────────────────────────────────────────

CLAIM_TOOL = {
    "name": "emit_claims",
    "description": (
        "Emit 0-20 financial claims from this document fragment. "
        "Extract only what is explicitly stated. Empty list if no financial claims. "
        "epistemic_class: asserted=seller/mgmt claim, observed=third-party real-time, "
        "attested=formally certified (QoE conclusion, IC decision), derived=YOU computed it."
    ),
    "input_schema": {
        "type": "object",
        "required": ["claims"],
        "properties": {
            "claims": {
                "type": "array",
                # A 250-word Excel chunk can contain more than four distinct
                # financial rows. The old cap deterministically dropped later
                # rows such as PAN-37's DSO assumption.
                "maxItems": 20,
                "items": {
                    "type": "object",
                    "required": [
                        "metric", "value", "unit", "period", "perimeter",
                        "entity", "period_canonical", "scope", "measurement", "basis", "scenario",
                        "claim_kind", "bound",
                        "epistemic_class", "direction", "topic",
                        "statement", "locator_hint",
                    ],
                    "additionalProperties": False,
                    "properties": {
                        "metric": {
                            "type": "string",
                            "enum": METRIC_ENUM,
                        },
                        "value": {
                            "type": ["number", "string", "null"],
                        },
                        "unit": {
                            "type": ["string", "null"],
                            "enum": UNIT_ENUM,
                        },
                        "period": {
                            "type": "string",
                            "minLength": 1,
                            "pattern": r".*\S.*",
                            "description": (
                                "The time period or vintage this claim refers to, in the document's own language. "
                                "Use the richest description available. Examples: "
                                "'FY2025A', 'FY2025A / FY2025E seller presentation', 'As of 2025-10-27', "
                                "'LTM Sep-25', '2020 to 2025-10-27', 'FY2026E–FY2030E forecast'. "
                                "Use 'cross-period' only if the claim is genuinely timeless."
                            ),
                        },
                        "perimeter": {
                            "type": "string",
                            "minLength": 1,
                            "pattern": r".*\S.*",
                            "description": (
                                "The precise economic scope — entity + metric definition + any adjustments. "
                                "Write the full descriptive string, not a shorthand. Examples: "
                                "'Alderstone consolidated revenue', "
                                "'Alderstone consolidated EBITDA under seller adjustment perimeter', "
                                "'Alderstone consolidated EBITDA under independent QoE adjustment perimeter', "
                                "'Alderstone consolidated EBITDA under Firm valuation, leverage and returns perimeter', "
                                "'Alderstone customer revenue measured by individual billing account', "
                                "'Alderstone accounts-receivable ledger'. "
                                "Use 'unknown' only if scope is truly unspecified."
                            ),
                        },
                        "entity": {
                            "type": "string",
                            "minLength": 1,
                            "description": (
                                "The bare proper name of the company or counterparty this claim "
                                "measures — nothing else. 'Keystone', 'Riverton', 'Apex'. "
                                "NOT a sentence, NOT the metric, NOT the perimeter. If the text "
                                "says 'the Company' or 'the Target', resolve it to the deal's "
                                "proper name when the document makes it unambiguous, otherwise "
                                "write 'unspecified'."
                            ),
                        },
                        "period_canonical": {
                            "type": "string",
                            "pattern": PERIOD_CANONICAL_PATTERN,
                            "description": (
                                "The same period as the `period` field, reduced to one canonical "
                                "token: FY2025A, FY2025E, FY2024A, LTM, ExitLTM, Opening, "
                                "EntryToExit, CrossPeriod, an ISO date (2026-03-31), or exactly "
                                "'none'. Write 'none' when the source states no period — never "
                                "substitute today's date, and never guess a fiscal year the "
                                "document does not give."
                            ),
                        },
                        "scope": {
                            "type": "string",
                            "enum": SCOPE_ENUM,
                            "description": (
                                "Economic boundary: consolidated=whole group; standalone=one "
                                "entity; customer=one customer/account; segment=a division."
                            ),
                        },
                        "measurement": {
                            "type": "string",
                            "minLength": 1,
                            "description": (
                                "WHICH SLICE of the quantity this figure covers, in the source's "
                                "own words — a service line, department, customer account, "
                                "product, region, or cost category. Write exactly 'total' when "
                                "the figure is the whole, undivided quantity.\n"
                                "This is what separates a breakdown from a disagreement. Three "
                                "service lines reporting 30.3, 20.0 and 14.1 are components of "
                                "one revenue, not three conflicting claims about it — but only "
                                "if each names its slice. Leaving this blank on a component "
                                "makes it collide with the total and with its siblings.\n"
                                "Examples: 'total', 'EHS compliance service line', "
                                "'field inspection', 'Riverton account', 'engineering headcount'."
                            ),
                        },
                        "basis": {
                            "type": "string",
                            "enum": BASIS_ENUM,
                            "description": (
                                "Which party's adjustments the figure is stated under. This is "
                                "what makes seller EBITDA and QoE EBITDA two legitimate numbers "
                                "rather than a contradiction, so do not guess: "
                                "SellerView=seller/management adjusted; QoEView=independent "
                                "quality-of-earnings; FirmView=our own underwriting basis; "
                                "CovenantView=credit-agreement definition; ReportedView=statutory "
                                "or unadjusted; unspecified=the source does not say."
                            ),
                        },
                        "scenario": {
                            "type": "string",
                            "enum": SCENARIO_ENUM,
                            "description": (
                                "Which case the figure belongs to. Base/upside/downside of the "
                                "same metric are NOT contradictions, so label them: "
                                "management=management's own forecast; seller=seller case; "
                                "base=our base case or an actual historical figure; "
                                "unspecified=the source does not distinguish cases."
                            ),
                        },
                        "claim_kind": {
                            "type": "string",
                            "enum": CLAIM_KIND_ENUM,
                            "description": (
                                "What kind of assertion this is. The test for a sentence with no "
                                "number is whether someone could later CONFIRM OR REFUTE it:\n"
                                "  QUANTITATIVE    carries a number or measured quantity\n"
                                "  DEFINITION      states how something is defined or calculated\n"
                                "  CONDITION       states a requirement, contingency or covenant\n"
                                "  ATTRIBUTION     states who said, did or decided something\n"
                                "  NEGATIVE        explicitly asserts an absence — checkable\n"
                                "  CHARACTERISATION a descriptor with no checkable content\n"
                                "Label honestly; what happens to each kind is not your "
                                "concern. 'A scaled regional platform', 'low capital "
                                "expenditure' and 'strong market position' are "
                                "CHARACTERISATION — the seller's adjectives, with nothing to "
                                "check. 'No material litigation as of 30 June' is NEGATIVE: an "
                                "absence someone can verify. Do not reach for NEGATIVE when a "
                                "phrase is merely favourable."
                            ),
                        },
                        "bound": {
                            "type": "string",
                            "enum": BOUND_ENUM,
                            "description": (
                                "How to read the number.\n"
                                "  EXACT       the figure as stated\n"
                                "  AT_LEAST    'more than 600', 'over 72%', 'at least 4'\n"
                                "  AT_MOST     'up to 15%', 'no more than 3x', 'below 2%'\n"
                                "  APPROXIMATE 'around 11.4', 'circa', 'roughly'\n"
                                "  RANGE       '1%-2%' — put the lower figure in value\n"
                                "  NONE        the claim carries no number\n"
                                "Get this right or a later consistent figure reads as a conflict: "
                                "'more than 600' stored as EXACT 600 makes a subsequent 640 look "
                                "like a contradiction when the two agree."
                            ),
                        },
                        "epistemic_class": {
                            "type": "string",
                            "enum": EPISTEMIC_CLASS_ENUM,
                            "description": (
                                "asserted=seller/management; observed=third-party real-time; "
                                "attested=QoE conclusion or IC decision; derived=computed by you"
                            ),
                        },
                        "direction": {
                            "type": "string",
                            "enum": DIRECTION_ENUM,
                            "description": "supports=positive thesis signal; contradicts=negative signal; context=neutral",
                        },
                        "topic": {
                            "type": "string",
                            "enum": TOPIC_ENUM,
                            "description": TOPIC_DESCRIPTION,
                        },
                        "definition_id": {
                            "type": ["string", "null"],
                            "enum": DEFINITION_ENUM,
                            "description": "If claim uses a specific defined term, reference its DEF-xxx ID.",
                        },
                        "statement": {
                            "type": "string",
                            "maxLength": 280,
                            "description": "One complete sentence stating the claim with full numeric and contextual detail.",
                        },
                        "locator_hint": {
                            "type": "string",
                            "minLength": 1,
                            "pattern": r".*\S.*",
                            "maxLength": 120,
                            "description": "Section heading, table name, slide, or line reference where this appears.",
                        },
                        "derivation": {
                            "type": ["string", "null"],
                            "description": "Required when epistemic_class=derived. State the computation.",
                        },
                        "author": {
                            "type": ["string", "null"],
                            "description": "Party making the claim (e.g. management, QoE provider, IC).",
                        },
                    },
                },
            }
        },
    },
}

SYSTEM_PROMPT = textwrap.dedent("""
    You are a financial claim extractor for a private equity firm (PANTA system).
    Extract only claims explicitly stated in the fragment. Never infer or interpolate.
    Return an empty list when the fragment contains no financial claims.

    EPISTEMIC CLASS — apply strictly by document source and claim type:
    - attested: ANY claim from the IC memo, QoE report conclusion, firm underwriting, or
      formal buyer analysis. This includes the firm's EBITDA view, QoE findings,
      covenant analysis, and any value the buyer's team formally concluded.
      When in doubt for IC memo or QoE document fragments → use attested.
    - asserted: seller or management stated claims with no third-party verification.
      CIM numbers, management presentations, seller-deck projections → asserted.
    - observed: a third-party measured or witnessed something directly in real time
      (e.g. QoE workpaper data room observation, call transcript quote, site visit).
    - derived: YOU computed this claim from two or more stated values — requires derivation field.
      Only use when you performed arithmetic (ratio, sum, subtraction).

    Source → class mapping (use this every time):
      IC memo         → attested  (the firm's own formal conclusion)
      QoE report      → attested  (third-party certifies)
      Auditor opinion → attested  (formally certified conclusion)
      Initial assessment → attested  (firm's own preliminary underwriting)
      Data room management documents → asserted
      Data room transactional/workpaper observations → observed
      Meeting notes / call transcript / DDQ → observed
      Seller CIM / IM → asserted  (seller's marketing claims)
      Management presentation → asserted
      Computed by you → derived

    PERIOD EXTRACTION — mandatory for every emitted claim:
    - Never leave period blank. Read the workbook column header, table header,
      section title, page heading, or source date when it is not repeated in the row.
    - Preserve the source language: FY2024, FY2025A, LTM Sep-25, Q2 2025,
      Opening Balance Sheet, Budget 2026, or 'as of 2025-10-27'.
    - If the only available time reference is the source effective date, use
      'as of {source effective date}' rather than leaving period blank.
    - Use 'cross-period' only for an explicitly time-invariant fact.

    PERIMETER INFERENCE — mandatory when the source determines scope:
    - Use exactly one canonical evidence-view label from document type:
      Seller View (CIM/management report), QoE View (QoE report),
      Firm View (firm internal memo), or Statutory (audited accounts).
    - Combine that view with the entity and structural scope (standalone,
      consolidated, deal level) rather than returning only the generic label.
    - For concentration, distinguish billing-account from ultimate-parent scope.
    - Use 'unknown' only when neither the fragment nor source metadata supports a scope.

    LOCATOR HINT:
    - The deterministic fragment locator is mandatory, never blank, and always
      retained by the pipeline with section and page/slide metadata when available.
    - Add locator_hint only when the row, cell, table, or subsection inside the
      fragment can be identified. Accepted hints include 'slide 12', 'page 7',
      'Sheet1!D42', 'section "Revenue Analysis"', 'table row 3', and
      'timestamp 00:14:22'; never invent a page or cell address.

    Identity rule — write the FULL perimeter and FULL period, not a shorthand:
    - period: use the document's own language including context, e.g.
        'FY2025A / FY2025E seller presentation'  not just 'FY2025A'
        'As of 2025-10-27'  not just '2025-10-27'
        'FY2026E–FY2030E forecast'  not just 'FY2026E'
    - perimeter: write the complete scope string, e.g.
        'Alderstone consolidated EBITDA under seller adjustment perimeter'
        'Alderstone consolidated EBITDA under independent QoE adjustment perimeter'
        'Alderstone customer revenue measured by individual billing account'
      NOT just 'Alderstone consolidated' or 'Alderstone standalone'.

    Do not collapse different definitions:
    - Firm EBITDA ≠ QoE EBITDA ≠ Covenant EBITDA ≠ Seller EBITDA
    - account-level concentration ≠ parent-level concentration
    - FY2025A ≠ LTM ≠ FY2026E

    No source statement may be invented. If the fragment is ambiguous, omit the claim.
""").strip()


# ─────────────────────────────────────────────────────────────────────────────
# L1 — Document parser (deterministic)
# ─────────────────────────────────────────────────────────────────────────────

CHUNK_WORDS = 250


@dataclass
class Chunk:
    chunk_id: str
    locator: str
    body: str
    source_path: str
    source_type: str
    source_record: dict
    word_count: int
    section_heading: str | None = None
    page_or_slide_number: int | None = None
    provenance: dict[str, Any] = field(default_factory=dict)
    period_context: dict[str, Any] = field(default_factory=dict)


class UnsupportedSourceError(ValueError):
    """Raised with a machine-readable response when V2 cannot parse safely."""

    def __init__(self, message: str, response: dict[str, Any] | None = None):
        super().__init__(message)
        self.response = response or {
            "schema": CAPABILITY_SCHEMA,
            "status": "REJECTED",
            "code": "UNSUPPORTED_SOURCE",
            "detail": message,
        }

    def to_dict(self) -> dict[str, Any]:
        return dict(self.response)


def _chunk_hash(body: str) -> str:
    return "ch-" + hashlib.sha256(body.encode()).hexdigest()[:12]


def _split_words(text: str, max_words: int, locator_prefix: str,
                 source_path: str, source_type: str, source_record: dict,
                 section_heading: str | None = None,
                 page_or_slide_number: int | None = None) -> list[Chunk]:
    words = text.split()
    chunks = []
    for i in range(0, len(words), max_words):
        body = " ".join(words[i: i + max_words])
        locator = f"{locator_prefix}:w{i}-{i + len(body.split())}"
        chunks.append(Chunk(
            chunk_id=_chunk_hash(body),
            locator=locator,
            body=body,
            source_path=source_path,
            source_type=source_type,
            source_record=source_record,
            word_count=len(body.split()),
            section_heading=section_heading,
            page_or_slide_number=page_or_slide_number,
        ))
    return chunks


def _reject_source(
    path: Path,
    code: str,
    detail: str,
    *,
    capability_id: str | None = None,
    action: str | None = None,
) -> UnsupportedSourceError:
    response = capability_failure(
        path,
        code,
        capability_id=capability_id,
        action=action,
        detail=detail,
    )
    message = detail
    if response.get("action"):
        message += f" Action: {response['action']}"
    return UnsupportedSourceError(message, response)


def _decorate_chunks(
    chunks: list[Chunk],
    path: Path,
    source_record: dict[str, Any],
    capability_id: str,
) -> list[Chunk]:
    """Attach source identity and honest period semantics to every fragment."""
    envelope = source_record.get("source_envelope") or {}
    for chunk in chunks:
        chunk.provenance = {
            **chunk.provenance,
            "source_id": source_record.get("source_id"),
            "source_version_id": envelope.get("source_version_id"),
            "case_id": envelope.get("case_id"),
            "original_filename": envelope.get("original_filename") or path.name,
            "locator": chunk.locator,
            "parser_capability": capability_id,
            "capability_contract": CAPABILITY_SCHEMA,
        }
        specific = dict(chunk.period_context)
        specific.update({
            "effective_date": envelope.get("effective_date") or source_record.get("effective_date") or None,
            "known_at": envelope.get("known_at") or source_record.get("known_at") or None,
            "semantics": "DECLARED_ONLY",
            "content_period_policy": "preserve source label; never infer at L1",
        })
        chunk.period_context = specific
    return chunks


class _VisibleHTML(HTMLParser):
    """Small deterministic HTML-to-text reader; scripts/styles are excluded."""

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self._hidden_depth = 0
        self.parts: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() in {"script", "style", "noscript"}:
            self._hidden_depth += 1
        elif not self._hidden_depth and tag.lower() in {"p", "div", "br", "li", "tr", "h1", "h2", "h3"}:
            self.parts.append("\n")

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() in {"script", "style", "noscript"} and self._hidden_depth:
            self._hidden_depth -= 1
        elif not self._hidden_depth and tag.lower() in {"p", "div", "li", "tr", "h1", "h2", "h3"}:
            self.parts.append("\n")

    def handle_data(self, data: str) -> None:
        if not self._hidden_depth:
            self.parts.append(data)

    def text(self) -> str:
        return "\n".join(line.strip() for line in "".join(self.parts).splitlines() if line.strip())


def _chunks_from_numbered_lines(
    lines: list[str],
    path: Path,
    max_words: int,
    source_type: str,
    source_record: dict[str, Any],
    locator_label: str = "lines",
) -> list[Chunk]:
    chunks: list[Chunk] = []
    pending: list[str] = []
    start = end = 0

    def flush() -> None:
        nonlocal pending, start, end
        if not pending:
            return
        body = "\n".join(pending)
        chunks.append(Chunk(
            chunk_id=_chunk_hash(body),
            locator=f"{path.name}::{locator_label}:{start}-{end}",
            body=body,
            source_path=str(path),
            source_type=source_type,
            source_record=source_record,
            word_count=len(body.split()),
        ))
        pending, start, end = [], 0, 0

    for line_number, line in enumerate(lines, 1):
        if not line.strip():
            continue
        projected = len((" ".join(pending + [line])).split())
        if pending and projected > max_words:
            flush()
        if not pending:
            start = line_number
        end = line_number
        pending.append(line.strip())
    flush()
    return chunks


def parse_markdown(path: Path, max_words: int = CHUNK_WORDS,
                   source_record: dict | None = None) -> list[Chunk]:
    src = source_record or _source_record(path)
    text = path.read_text(encoding="utf-8")
    parts = re.split(r"(?=^#{2,3} )", text, flags=re.MULTILINE)
    chunks: list[Chunk] = []
    for part in parts:
        if not part.strip():
            continue
        header_match = re.match(r"^(#{2,3} .+)", part)
        section_label = header_match.group(1)[:50].strip() if header_match else "section"
        words = part.split()
        if len(words) <= max_words:
            body = part.strip()
            chunks.append(Chunk(
                chunk_id=_chunk_hash(body),
                locator=f"{path.name}::{section_label}",
                body=body,
                source_path=str(path),
                source_type="markdown",
                source_record=src,
                word_count=len(words),
                section_heading=section_label,
            ))
        else:
            sub = _split_words(part, max_words, f"{path.name}::{section_label}",
                               str(path), "markdown", src,
                               section_heading=section_label)
            chunks.extend(sub)
    return chunks


def parse_plain_text(path: Path, max_words: int = CHUNK_WORDS,
                     source_record: dict | None = None,
                     source_type: str = "text") -> list[Chunk]:
    src = source_record or _source_record(path)
    try:
        text = path.read_text(encoding="utf-8-sig")
    except UnicodeDecodeError as exc:
        raise _reject_source(
            path,
            "TEXT_ENCODING_UNSUPPORTED",
            f"{path.name} is not valid UTF-8 text ({exc}).",
            action="Export the artifact as UTF-8 without replacing source characters.",
        ) from exc
    return _chunks_from_numbered_lines(text.splitlines(), path, max_words, source_type, src)


def parse_html(path: Path, max_words: int = CHUNK_WORDS,
               source_record: dict | None = None) -> list[Chunk]:
    src = source_record or _source_record(path)
    try:
        markup = path.read_text(encoding="utf-8-sig")
    except UnicodeDecodeError as exc:
        raise _reject_source(
            path,
            "TEXT_ENCODING_UNSUPPORTED",
            f"{path.name} is not valid UTF-8 HTML ({exc}).",
            action="Export the page as UTF-8 HTML or plain text.",
        ) from exc
    reader = _VisibleHTML()
    reader.feed(markup)
    return _chunks_from_numbered_lines(
        reader.text().splitlines(), path, max_words, "html", src, "visible-lines"
    )


def parse_csv(path: Path, max_words: int = CHUNK_WORDS,
              source_record: dict | None = None) -> list[Chunk]:
    src = source_record or _source_record(path)
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.reader(handle))
    except (UnicodeDecodeError, csv.Error) as exc:
        raise _reject_source(
            path,
            "CSV_INVALID",
            f"Cannot parse {path.name} as UTF-8 CSV: {exc}.",
            action="Export a UTF-8 CSV with a single consistent delimiter and header row.",
        ) from exc
    lines = [" | ".join(f"c{column}={value}" for column, value in enumerate(row, 1)) for row in rows]
    return _chunks_from_numbered_lines(lines, path, max_words, "csv", src, "rows")


def _read_openxml_root(path: Path, member: str, capability_id: str) -> ElementTree.Element:
    try:
        with zipfile.ZipFile(path) as archive:
            payload = archive.read(member)
        return ElementTree.fromstring(payload)
    except (OSError, KeyError, zipfile.BadZipFile, ElementTree.ParseError) as exc:
        raise _reject_source(
            path,
            "OPENXML_INVALID",
            f"{path.name} is not a valid {capability_id.upper()} package: {exc}.",
            capability_id=capability_id,
            action=f"Open the artifact in its authoring application and save a valid {path.suffix.upper()} copy.",
        ) from exc


def parse_docx(path: Path, max_words: int = CHUNK_WORDS,
               source_record: dict | None = None) -> list[Chunk]:
    """Extract addressable DOCX paragraphs using only the Open XML package."""
    src = source_record or _source_record(path)
    root = _read_openxml_root(path, "word/document.xml", "docx")
    namespace = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    paragraphs: list[str] = []
    for paragraph in root.iter(namespace + "p"):
        text = "".join(node.text or "" for node in paragraph.iter(namespace + "t")).strip()
        if text:
            paragraphs.append(text)
    return _chunks_from_numbered_lines(paragraphs, path, max_words, "docx", src, "paragraphs")


def parse_pptx(path: Path, max_words: int = CHUNK_WORDS,
               source_record: dict | None = None) -> list[Chunk]:
    """Extract text per slide; speaker notes and embedded objects are not invented."""
    src = source_record or _source_record(path)
    try:
        with zipfile.ZipFile(path) as archive:
            slide_members = sorted(
                (
                    name for name in archive.namelist()
                    if re.fullmatch(r"ppt/slides/slide\d+\.xml", name)
                ),
                key=lambda name: int(re.search(r"slide(\d+)\.xml$", name).group(1)),  # type: ignore[union-attr]
            )
            payloads = [(name, archive.read(name)) for name in slide_members]
    except (OSError, zipfile.BadZipFile) as exc:
        raise _reject_source(
            path,
            "OPENXML_INVALID",
            f"{path.name} is not a valid PPTX package: {exc}.",
            capability_id="pptx",
            action="Open the deck in PowerPoint and save a valid .pptx copy.",
        ) from exc
    chunks: list[Chunk] = []
    drawing_text = "{http://schemas.openxmlformats.org/drawingml/2006/main}t"
    for member, payload in payloads:
        slide_number = int(re.search(r"slide(\d+)\.xml$", member).group(1))  # type: ignore[union-attr]
        try:
            root = ElementTree.fromstring(payload)
        except ElementTree.ParseError as exc:
            raise _reject_source(
                path, "OPENXML_INVALID", f"Slide {slide_number} XML is malformed: {exc}.",
                capability_id="pptx",
                action="Open the deck in PowerPoint and save a repaired .pptx copy.",
            ) from exc
        body = "\n".join(node.text.strip() for node in root.iter(drawing_text) if node.text and node.text.strip())
        if not body:
            continue
        chunks.extend(_split_words(
            body, max_words, f"{path.name}::slide:{slide_number}", str(path), "pptx", src,
            page_or_slide_number=slide_number,
        ))
    return chunks


def parse_transcript(path: Path, max_words: int = CHUNK_WORDS,
                     source_record: dict | None = None) -> list[Chunk]:
    src = source_record or _source_record(path)
    if path.suffix.lower() == ".txt":
        return parse_plain_text(path, max_words, src, "transcript")
    try:
        text = path.read_text(encoding="utf-8-sig").replace("\r\n", "\n")
    except UnicodeDecodeError as exc:
        raise _reject_source(
            path, "TEXT_ENCODING_UNSUPPORTED", f"Transcript is not valid UTF-8: {exc}.",
            capability_id="transcript_export",
            action="Export the transcript as UTF-8 WebVTT, SRT, or plain text.",
        ) from exc
    blocks = [block.strip() for block in re.split(r"\n\s*\n", text) if block.strip()]
    chunks: list[Chunk] = []
    cue_number = 0
    for block in blocks:
        lines = [line.strip() for line in block.splitlines() if line.strip()]
        if not lines or lines[0].upper() == "WEBVTT":
            continue
        timing_index = next((i for i, line in enumerate(lines) if "-->" in line), None)
        if timing_index is None:
            continue
        cue_number += 1
        timecode = lines[timing_index]
        body = "\n".join(lines[timing_index + 1:]).strip()
        if not body:
            continue
        locator = f"{path.name}::cue:{cue_number}:{timecode.replace(' ', '')}"
        chunks.extend(_split_words(body, max_words, locator, str(path), "transcript", src))
    return chunks


def _email_body(message: Any) -> str:
    plain: list[str] = []
    html_parts: list[str] = []
    parts = message.walk() if message.is_multipart() else [message]
    for part in parts:
        if part.is_multipart() or part.get_content_disposition() == "attachment":
            continue
        content_type = part.get_content_type()
        if content_type not in {"text/plain", "text/html"}:
            continue
        try:
            content = part.get_content()
        except (LookupError, UnicodeDecodeError):
            payload = part.get_payload(decode=True) or b""
            content = payload.decode("utf-8", errors="replace")
        if content_type == "text/plain":
            plain.append(str(content))
        else:
            reader = _VisibleHTML()
            reader.feed(str(content))
            html_parts.append(reader.text())
    return "\n".join(plain or html_parts).strip()


def _email_date(message: Any) -> str | None:
    raw = message.get("Date")
    if not raw:
        return None
    try:
        return parsedate_to_datetime(raw).isoformat()
    except (TypeError, ValueError, OverflowError):
        return None


def parse_email(path: Path, max_words: int = CHUNK_WORDS,
                source_record: dict | None = None) -> list[Chunk]:
    src = source_record or _source_record(path)
    if path.suffix.lower() == ".eml":
        try:
            messages = [BytesParser(policy=email_policy.default).parsebytes(path.read_bytes())]
        except (OSError, ValueError) as exc:
            raise _reject_source(
                path, "EMAIL_INVALID", f"Cannot parse RFC 822 email: {exc}.",
                capability_id="email_export",
                action="Export the message as RFC 822 .eml with complete headers.",
            ) from exc
    else:
        box = mailbox.mbox(path, create=False)
        try:
            messages = [message for message in box]
        finally:
            box.close()
    chunks: list[Chunk] = []
    for message_number, message in enumerate(messages, 1):
        body = _email_body(message)
        if not body:
            continue
        attachment_count = sum(
            1 for part in message.walk()
            if part.get_content_disposition() == "attachment"
        )
        message_chunks = _split_words(
            body, max_words, f"{path.name}::message:{message_number}:body",
            str(path), "email", src,
            section_heading=str(message.get("Subject") or "(no subject)"),
        )
        document_date = _email_date(message)
        for chunk in message_chunks:
            chunk.provenance = {
                "excluded_attachments": attachment_count,
                "attachment_policy": "SEPARATE_SOURCE_ENVELOPE_REQUIRED",
            }
            chunk.period_context = {
                "document_date": document_date,
                "document_date_source": "email-header:Date" if document_date else None,
            }
        chunks.extend(message_chunks)
    return chunks


def parse_pdf(path: Path, max_words: int = CHUNK_WORDS,
              source_record: dict | None = None) -> list[Chunk]:
    try:
        import pdfplumber
    except ImportError as exc:
        raise _reject_source(
            path,
            "READER_UNAVAILABLE",
            "Native PDF parsing is unavailable because the approved local reader is not installed.",
            capability_id="native_pdf",
        ) from exc
    src = source_record or _source_record(path)
    chunks: list[Chunk] = []
    try:
        with pdfplumber.open(path) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                text = page.extract_text() or ""
                if not text.strip():
                    continue
                words = text.split()
                if len(words) <= max_words:
                    body = text.strip()
                    chunks.append(Chunk(
                        chunk_id=_chunk_hash(body),
                        locator=f"p{page_num}",
                        body=body,
                        source_path=str(path),
                        source_type="pdf",
                        source_record=src,
                        word_count=len(words),
                        page_or_slide_number=page_num,
                    ))
                else:
                    sub = _split_words(text, max_words, f"p{page_num}",
                                       str(path), "pdf", src,
                                       page_or_slide_number=page_num)
                    chunks.extend(sub)
    except UnsupportedSourceError:
        raise
    except Exception as exc:
        raise _reject_source(
            path,
            "PDF_INVALID",
            f"The native PDF reader could not open {path.name}: {exc}.",
            capability_id="native_pdf",
            action="Repair the PDF or export it as searchable PDF or UTF-8 text.",
        ) from exc
    if not chunks:
        raise _reject_source(
            path,
            "OCR_REQUIRED",
            f"{path.name} contains no extractable text; PANTA will not pretend that image pixels were parsed.",
            capability_id="scanned_pdf_ocr",
        )
    return chunks


_XLSX_COLOR_ROLES = {
    "0000FF": "input",
    "008000": "cross_sheet_link",
    "00B050": "cross_sheet_link",
    "800080": "external_link",
    "7030A0": "external_link",
    "FF0000": "external_link",
    "C00000": "external_link",
}


def _xlsx_color_hex(color: Any) -> str | None:
    """Return a six-digit RGB value when an Open XML color is resolvable."""
    if color is None:
        return None
    color_type = str(getattr(color, "type", "") or "")
    raw: Any = None
    if color_type == "rgb":
        raw = getattr(color, "rgb", None)
    elif color_type == "indexed":
        indexed = getattr(color, "indexed", None)
        try:
            from openpyxl.styles.colors import COLOR_INDEX

            raw = COLOR_INDEX[int(indexed)]
        except (ImportError, IndexError, TypeError, ValueError):
            return None
    if not isinstance(raw, str):
        return None
    normalized = raw.strip().lstrip("#").upper()
    if len(normalized) == 8:
        normalized = normalized[-6:]
    return normalized if re.fullmatch(r"[0-9A-F]{6}", normalized) else None


def _xlsx_cell_role(cell: Any, raw: Any) -> str:
    """Classify a cell from formula structure plus declared font/fill color."""
    formula = isinstance(raw, str) and raw.startswith("=")
    expression = raw.upper() if formula else ""
    colors = {
        value
        for value in (
            _xlsx_color_hex(getattr(getattr(cell, "font", None), "color", None)),
            _xlsx_color_hex(
                getattr(getattr(cell, "fill", None), "fgColor", None)
                if getattr(getattr(cell, "fill", None), "patternType", None)
                else None
            ),
        )
        if value
    }
    color_roles = {_XLSX_COLOR_ROLES[color] for color in colors if color in _XLSX_COLOR_ROLES}
    if formula:
        if "[" in expression or "external_link" in color_roles:
            return "external_link"
        if "!" in expression or "cross_sheet_link" in color_roles:
            return "cross_sheet_link"
        return "formula"
    if "input" in color_roles:
        return "input"
    # Non-formula green/red/purple cells are still declared with the workbook's
    # link convention.  The value is preserved as-is; the role is context, not
    # a claim that PANTA resolved the target.
    if "external_link" in color_roles:
        return "external_link"
    if "cross_sheet_link" in color_roles:
        return "cross_sheet_link"
    return "unclassified"


def _xlsx_merged_ranges(path: Path, worksheet_path: str) -> list[tuple[int, int, int, int]]:
    """Read merged ranges without loading a large worksheet into memory.

    ``ReadOnlyWorksheet`` intentionally omits ``merged_cells``.  The merge
    declarations are tiny metadata in the worksheet XML, so reading them
    directly preserves the streaming workbook path used for real LBO models.
    """
    try:
        from openpyxl.utils.cell import range_boundaries

        with zipfile.ZipFile(path) as archive:
            with archive.open(worksheet_path.lstrip("/")) as worksheet_xml:
                references = []
                namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}mergeCell"
                for _event, node in ElementTree.iterparse(worksheet_xml, events=("end",)):
                    if node.tag == namespace and node.attrib.get("ref"):
                        references.append(node.attrib["ref"])
                    node.clear()
        ranges = [tuple(range_boundaries(reference)) for reference in references]
    except (
        ImportError,
        OSError,
        KeyError,
        TypeError,
        ValueError,
        zipfile.BadZipFile,
        ElementTree.ParseError,
    ) as exc:
        raise _reject_source(
            path,
            "WORKBOOK_INVALID",
            f"Cannot read merged-cell metadata from {path.name}: {exc}.",
            capability_id="openxml_workbook",
            action="Open the workbook in Excel and save a valid .xlsx/.xlsm copy.",
        ) from exc
    return ranges


def _xlsx_header_token(value: Any) -> str | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, str):
        token = value.strip()
        return token if token and not token.startswith("=") else None
    # Dates are labels in model grids; ordinary numeric values are not.  A
    # four-digit year is the one useful numeric header that can be identified
    # without guessing from number format or neighboring economics.
    if hasattr(value, "isoformat"):
        return str(value)
    if isinstance(value, int) and 1900 <= value <= 2200:
        return str(value)
    return None


def _xlsx_header_path(
    row_number: int,
    column_number: int,
    raw: Any,
    context_values: dict[tuple[int, int], Any],
) -> list[str]:
    """Build a deterministic column-header + row-label path for one cell."""
    current = _xlsx_header_token(raw)
    column_headers = []
    for prior_row in range(row_number - 1, 0, -1):
        token = _xlsx_header_token(context_values.get((prior_row, column_number)))
        if token and token != current and token not in column_headers:
            column_headers.append(token)
            if len(column_headers) == 3:
                break
    column_headers.reverse()

    row_labels = []
    for prior_column in range(1, column_number):
        token = _xlsx_header_token(context_values.get((row_number, prior_column)))
        if token and token != current and token not in row_labels:
            row_labels.append(token)
    row_labels = row_labels[-3:]

    path = []
    for token in column_headers + row_labels:
        if token not in path:
            path.append(token)
    return path


def parse_xlsx(path: Path, max_words: int = CHUNK_WORDS,
               source_record: dict | None = None) -> list[Chunk]:
    """Create reproducible, cell-addressable chunks from an Excel workbook.

    A workbook is not prose: formulas and their cached outputs carry different
    meanings.  The chunk body keeps both, while the locator names the exact
    sheet and cell range so a reviewer can verify any extracted claim.
    """
    try:
        import openpyxl
        from openpyxl.utils.cell import get_column_letter
    except ImportError as exc:
        raise _reject_source(
            path,
            "READER_UNAVAILABLE",
            "Open XML workbook parsing is unavailable because openpyxl is not installed.",
            capability_id="openxml_workbook",
        ) from exc
    try:
        formulas = openpyxl.load_workbook(path, data_only=False, read_only=True)
        values = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        raise _reject_source(
            path,
            "WORKBOOK_INVALID",
            f"Cannot read workbook {path.name}: {exc}.",
            capability_id="openxml_workbook",
            action="Open the workbook in Excel and save a valid .xlsx/.xlsm copy without flattening formulas.",
        ) from exc
    src = source_record or _source_record(path)
    chunks: list[Chunk] = []
    workbook_has_formulas = any(
        isinstance(cell.value, str) and cell.value.startswith("=")
        for ws in formulas.worksheets for row in ws.iter_rows() for cell in row
    )
    effective_max_words = max(max_words, 1200) if workbook_has_formulas else max_words
    for sheet_name in formulas.sheetnames:
        ws, value_ws = formulas[sheet_name], values[sheet_name]
        merged_ranges = _xlsx_merged_ranges(path, str(ws._worksheet_path))
        merged_anchors = {
            (row_number, column_number): (min_row, min_column)
            for min_column, min_row, max_column, max_row in merged_ranges
            for row_number in range(min_row, max_row + 1)
            for column_number in range(min_column, max_column + 1)
        }
        merged_values: dict[tuple[int, int], tuple[Any, Any, str]] = {}
        context_values: dict[tuple[int, int], Any] = {}
        # Models can contain thousands of raw input/output rows.  Their
        # complete deterministic cell graph is captured separately; L2
        # should see formula-bearing rows plus their labels, headers and stated
        # inputs, not every bare numeric row.  Previously every formula-free
        # row on a computing sheet was dropped, including section titles,
        # period headers and input rows needed to interpret the formulas.
        sheet_has_formulas = any(
            isinstance(cell.value, str) and cell.value.startswith("=")
            for row in ws.iter_rows() for cell in row
        )
        pending: list[str] = []
        start_row = end_row = None
        def flush() -> None:
            nonlocal pending, start_row, end_row
            if not pending or start_row is None or end_row is None:
                return
            body = f"Workbook: {path.name}\nSheet: {sheet_name}\n" + "\n".join(pending)
            chunks.append(Chunk(
                chunk_id=_chunk_hash(body),
                locator=f"{path.name}::{sheet_name}!{start_row}:{end_row}",
                body=body, source_path=str(path), source_type="xlsx",
                source_record=src, word_count=len(body.split()),
                section_heading=sheet_name,
            ))
            pending, start_row, end_row = [], None, None
        for row_number, (row, cached_row) in enumerate(zip(ws.iter_rows(), value_ws.iter_rows()), start=1):
            for column_number, (cell, cached_cell) in enumerate(zip(row, cached_row), start=1):
                anchor = merged_anchors.get((row_number, column_number))
                if anchor == (row_number, column_number) and cell.value is not None:
                    merged_values[anchor] = (
                        cell.value,
                        cached_cell.value,
                        _xlsx_cell_role(cell, cell.value),
                    )
            cells: list[str] = []
            row_roles: list[str] = []
            row_raw_values: list[Any] = []
            for column_number, (cell, cached_cell) in enumerate(zip(row, cached_row), start=1):
                raw = cell.value
                cached = cached_cell.value
                role = _xlsx_cell_role(cell, raw)
                anchor = merged_anchors.get((row_number, column_number))
                if raw is None and anchor in merged_values:
                    raw, cached, role = merged_values[anchor]
                if raw is None:
                    continue
                context_values[(row_number, column_number)] = raw
                row_roles.append(role)
                row_raw_values.append(raw)
                is_data_cell = (
                    isinstance(raw, str) and raw.startswith("=")
                ) or (
                    not isinstance(raw, str) and not hasattr(raw, "isoformat")
                )
                header_path = (
                    _xlsx_header_path(
                        row_number,
                        column_number,
                        raw,
                        context_values,
                    )
                    if is_data_cell
                    else []
                )
                context = []
                if header_path:
                    context.append(f"header_path={' > '.join(header_path)}")
                context.append(f"role={role}")
                suffix = f" [{'; '.join(context)}]"
                coordinate = f"{get_column_letter(column_number)}{row_number}"
                if isinstance(raw, str) and raw.startswith("="):
                    cells.append(f"{coordinate}=FORMULA({raw}); cached={cached!r}{suffix}")
                else:
                    cells.append(f"{coordinate}={raw}{suffix}")
            if not cells:
                continue
            has_formula = any(
                isinstance(value, str) and value.startswith("=")
                for value in row_raw_values
            )
            has_label_or_date = any(_xlsx_header_token(value) for value in row_raw_values)
            has_declared_input = "input" in row_roles
            if sheet_has_formulas and not (has_formula or has_label_or_date or has_declared_input):
                continue
            line = " | ".join(cells)
            projected = len((" ".join(pending + [line])).split())
            if pending and projected > effective_max_words:
                flush()
            if start_row is None:
                start_row = row_number
            end_row = row_number
            pending.append(line)
        flush()
    return chunks


def parse_source(path: Path, max_words: int = CHUNK_WORDS,
                 source_record: dict | None = None) -> list[Chunk]:
    src = source_record or _source_record(path)
    envelope = src.get("source_envelope") or {}
    capability = resolve_source_capability(path, {
        "document_type": envelope.get("document_type") or src.get("doc_type"),
        "parser_route": envelope.get("parser_route"),
    })
    capability_id = str(capability["capability_id"])
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        chunks = parse_pdf(path, max_words, src)
    elif suffix in (".md", ".markdown"):
        chunks = parse_markdown(path, max_words, src)
    elif suffix == ".txt" and capability_id == "transcript_export":
        chunks = parse_transcript(path, max_words, src)
    elif suffix == ".txt":
        chunks = parse_plain_text(path, max_words, src)
    elif suffix in (".html", ".htm"):
        chunks = parse_html(path, max_words, src)
    elif suffix == ".csv":
        chunks = parse_csv(path, max_words, src)
    elif suffix == ".docx":
        chunks = parse_docx(path, max_words, src)
    elif suffix == ".pptx":
        chunks = parse_pptx(path, max_words, src)
    elif suffix in (".srt", ".vtt"):
        chunks = parse_transcript(path, max_words, src)
    elif suffix in (".eml", ".mbox"):
        chunks = parse_email(path, max_words, src)
    elif suffix in (".xlsx", ".xlsm"):
        chunks = parse_xlsx(path, max_words, src)
    elif suffix == ".xls":
        raise _reject_source(
            path,
            "LEGACY_EXCEL_UNSUPPORTED",
            "Legacy .xls is not supported; convert the workbook to .xlsx before extraction.",
            capability_id="legacy_excel",
        )
    elif suffix == ".msg":
        raise _reject_source(
            path,
            "OUTLOOK_MSG_UNSUPPORTED",
            "Binary Outlook .msg is not parsed because doing so would require an unapproved decoder.",
            capability_id="outlook_msg",
        )
    else:
        raise _reject_source(
            path,
            "UNSUPPORTED_SOURCE",
            f"Unsupported source type: {suffix or '<no extension>'}.",
        )
    if not chunks:
        raise _reject_source(
            path,
            "NO_EXTRACTABLE_TEXT",
            f"{path.name} contains no safely extractable text.",
            capability_id=capability_id,
            action="Verify the export contains visible text and upload it again; attachments require their own SourceEnvelope.",
        )
    return _decorate_chunks(chunks, path, src, capability_id)


def load_manifest(
    manifest: str,
    deal: str,
    source_dir: Path = VAULT_INBOX,
) -> list[Path]:
    """Return ordered list of source files for the given manifest."""
    keys = MANIFEST_SOURCES.get(manifest, [])
    paths = []
    for key in keys:
        # Try stem match in vault/inbox
        for ext in (
            ".md", ".markdown", ".txt", ".html", ".htm", ".csv", ".pdf",
            ".docx", ".pptx", ".xlsx", ".xlsm", ".srt", ".vtt", ".eml", ".mbox",
        ):
            candidate = source_dir / f"{key}{ext}"
            if candidate.exists():
                paths.append(candidate)
                break
    return paths


# ─────────────────────────────────────────────────────────────────────────────
# L2 — LLM annotator (schema-constrained via tool_use)
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class RawClaim:
    metric: str
    value: Any
    unit: str | None
    period: str | None
    perimeter: str | None
    epistemic_class: str
    direction: str
    topic: str
    definition_id: str | None
    statement: str
    locator: str
    source_id: str
    source_path: str
    known_at: str
    source_version_id: str | None = None
    derivation: str | None = None
    author: str | None = None
    # Identity dimensions. Defaulted so a cached raw_claims run predating these
    # fields still loads — it simply scores as unresolvable until re-extracted.
    entity: str = "unspecified"
    period_canonical: str = "none"
    scope: str = "unspecified"
    # Which slice of the quantity. "total" means the whole; blank collides a
    # component with its own total, which reads as a contradiction.
    measurement: str = "total"
    claim_kind: str = "QUANTITATIVE"
    bound: str = "EXACT"
    basis: str = "unspecified"
    scenario: str = "unspecified"


def _is_fatal_provider_error(exc: Exception) -> bool:
    """Return True when retrying more chunks with the same key cannot succeed."""
    status_code = getattr(exc, "status_code", None)
    message = str(exc).lower()
    # OpenRouter uses HTTP 402 both for an exhausted account and for its
    # temporary in-flight spending cap.  The latter explicitly asks callers to
    # wait for outstanding requests to settle, so treating every 402 as fatal
    # throws away an otherwise resumable extraction batch.
    if status_code == 402 and "in_flight_budget_exhausted" in message:
        return False
    if status_code in {401, 402, 403}:
        return True
    return any(
        marker in message
        for marker in (
            "billing_error",
            "permission_error",
            "key limit exceeded",
            "invalid api key",
        )
    )


def _provider_retry_delay(exc: Exception, attempt: int) -> float | None:
    """Return a bounded retry delay for transient provider failures."""
    status_code = getattr(exc, "status_code", None)
    message = str(exc)
    transient_in_flight_cap = (
        status_code == 402
        and "in_flight_budget_exhausted" in message.lower()
    )
    if status_code not in {408, 409, 429, 500, 502, 503, 504, 529} and not transient_in_flight_cap:
        return None

    response = getattr(exc, "response", None)
    headers = getattr(response, "headers", {}) if response is not None else {}
    raw_delay = headers.get("Retry-After") if headers else None
    if raw_delay is None:
        match = re.search(
            r"retry_after_seconds(?:_raw)?['\"\s:]+([0-9]+(?:\.[0-9]+)?)",
            message,
            re.IGNORECASE,
        )
        if match is None:
            match = re.search(
                r"retry-after['\"\s:]+['\"]?([0-9]+(?:\.[0-9]+)?)",
                message,
                re.IGNORECASE,
            )
        raw_delay = match.group(1) if match else None
    try:
        delay = float(raw_delay) if raw_delay is not None else 2 ** attempt
    except (TypeError, ValueError):
        delay = 2 ** attempt
    return max(1.0, min(delay, 30.0))


def annotate_chunk(
    chunk: Chunk,
    client,
    deal: str,
    rate_limit_delay: float = 0.25,
    *,
    raise_errors: bool = False,
) -> list[RawClaim]:
    src = chunk.source_record
    prompt = (
        f"DEAL: {deal}\n"
        f"SOURCE: {src['source_id']} ({src['name']}) — {src['doc_type']}\n"
        f"EFFECTIVE DATE: {src.get('effective_date') or 'not available'}\n"
        f"KNOWN AT: {src['known_at']}\n"
        f"FRAGMENT LOCATOR: {chunk.locator}\n\n"
        f"SECTION HEADING: {chunk.section_heading or 'not available'}\n"
        f"PAGE OR SLIDE: {chunk.page_or_slide_number or 'not available'}\n\n"
        f"{chunk.body}"
    )
    try:
        request = {
            "model": MODEL,
            # Match the larger schema capacity: Excel chunks commonly contain
            # more than four claim-bearing rows and need room for tool JSON.
            "max_tokens": MAX_TOKENS,
            "system": SYSTEM_PROMPT,
            "tools": [CLAIM_TOOL],
            "tool_choice": {"type": "tool", "name": "emit_claims"},
            "messages": [{"role": "user", "content": prompt}],
        }
        # Sampling temperature. SDK 1.0.0 dropped `temperature` from the typed
        # signature of messages.create, but the API still honours it, and the SDK
        # forwards extra_body verbatim — so this is the only route to it here.
        #
        # It is not a nicety. Measured on one 92-word Keystone chunk at the API
        # default of 1.0: the claim count was stable at 11 across four runs, yet
        # all four identity fingerprints differed and not one of the 33
        # identities recurred in every run — the same EBITDA was FirmView twice
        # and something else the other two times.
        #
        # Extraction varying in *what* it finds would be visible. Varying in how
        # it *classifies* what it found is invisible and worse: the same quantity
        # lands under different identities, so grouping and contradiction
        # detection quietly stop working.
        extra_body = {"temperature": 0}
        provider_extra = openrouter_extra_body()
        if provider_extra:
            extra_body.update(provider_extra)
        request["extra_body"] = extra_body
        resp = client.messages.create(**request)
        time.sleep(rate_limit_delay)
    except Exception as e:
        if raise_errors:
            raise
        print(f"  [L2 ERROR] {chunk.chunk_id}: {e}", file=sys.stderr)
        return []

    raw_claims: list[RawClaim] = []
    for block in resp.content:
        if block.type == "tool_use" and block.name == "emit_claims":
            for c in block.input.get("claims", []):
                period = _non_empty_l2_text(c.get("period"))
                if period is None:
                    effective_date = _non_empty_l2_text(src.get("effective_date"))
                    period = f"as of {effective_date}" if effective_date else "unknown"

                # There is no reliable economic-scope inference when the model
                # omitted the perimeter. Preserve that absence explicitly rather
                # than inventing an entity, consolidation basis, or adjustment view.
                perimeter = _non_empty_l2_text(c.get("perimeter")) or "unknown"

                hint = _non_empty_l2_text(c.get("locator_hint"))
                if hint is None:
                    hint = _non_empty_l2_text(chunk.section_heading)
                if hint is None and chunk.page_or_slide_number is not None:
                    hint = f"page or slide {chunk.page_or_slide_number}"
                if hint is None:
                    # The fragment locator is deterministic L1 metadata. Reusing it
                    # does not add a fabricated location and the branch below avoids
                    # appending the locator to itself.
                    hint = chunk.locator
                # The model sometimes echoes the fragment locator it was given
                # instead of naming a position inside it, which produced
                # "file.md::## Heading:file.md::## Heading". Only append a hint
                # that adds something.
                if hint and hint not in chunk.locator and chunk.locator not in hint:
                    locator = f"{chunk.locator}:{hint}"
                else:
                    locator = chunk.locator
                raw_claims.append(RawClaim(
                    metric=c["metric"],
                    value=c.get("value"),
                    unit=c.get("unit"),
                    period=period,
                    perimeter=perimeter,
                    epistemic_class=c.get("epistemic_class", "asserted"),
                    direction=c.get("direction", "context"),
                    topic=c.get("topic", "OTHER"),
                    definition_id=c.get("definition_id"),
                    statement=c.get("statement", ""),
                    locator=locator,
                    source_id=src["source_id"],
                    source_path=chunk.source_path,
                    known_at=src["known_at"],
                    source_version_id=(
                        (src.get("source_envelope") or {}).get("source_version_id")
                    ),
                    derivation=c.get("derivation"),
                    author=c.get("author"),
                    entity=c.get("entity") or "unspecified",
                    period_canonical=c.get("period_canonical") or "none",
                    scope=c.get("scope") or "unspecified",
                    measurement=c.get("measurement") or "total",
                    claim_kind=c.get("claim_kind") or "QUANTITATIVE",
                    bound=c.get("bound") or "EXACT",
                    basis=c.get("basis") or "unspecified",
                    scenario=c.get("scenario") or "unspecified",
                ))
    return raw_claims


# ─────────────────────────────────────────────────────────────────────────────
# L3 — Validator / normalizer (deterministic)
# ─────────────────────────────────────────────────────────────────────────────

def _parse_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", "").replace("%", "").strip())
    except (TypeError, ValueError):
        return None


def _non_empty_l2_text(value: Any) -> str | None:
    """Return trimmed non-empty L2 text without coercing non-string values."""
    if not isinstance(value, str):
        return None
    value = value.strip()
    return value or None


def _normalize_period(raw: str | None) -> str:
    if not raw:
        return "unknown"
    if raw.upper().strip() in PERIOD_MAP:
        return PERIOD_MAP[raw.upper().strip()]
    if re.match(r"^\d{4}-\d{2}-\d{2}$", raw):
        return raw
    return f"RAW:{raw}"


@dataclass
class CanonicalClaim:
    # --- CAP-003 required fields (aligned to benchmark v1.1) ---
    claim_id: str           # canonical claim:<sha256> content-addressed ID
    statement: str
    source_id: str          # SRC-CIM, SRC-QOE, etc.
    locator: str
    epistemic_class: str    # asserted / observed / derived / attested
    value: float | None
    value_raw: Any
    unit: str | None
    definition_id: str | None
    period: str             # raw period label from source
    period_iso: str         # normalized ISO date or "LTM"
    perimeter: str
    ground_truth_flag: bool  # True only for benchmark validation claims
    validation_only: bool    # True only for quarantined benchmark claims
    notes: str | None
    # --- compiler metadata (NOT in frozen CIC schema) ---
    metric: str             # from METRIC_ENUM — for dedup and reporting
    source_path: str
    known_at: str
    direction: str
    topic: str
    derivation: str | None
    author: str | None
    source_version_id: str | None = None
    # Identity dimensions — see tools/object_identity.py for how they key a claim.
    entity: str = "unspecified"
    period_canonical: str = "none"
    scope: str = "unspecified"
    measurement: str = "total"
    claim_kind: str = "QUANTITATIVE"
    bound: str = "EXACT"
    basis: str = "unspecified"
    scenario: str = "unspecified"
    validation_errors: list[str] = field(default_factory=list)


def validate(raw: RawClaim) -> CanonicalClaim:
    errors: list[str] = []
    if raw.metric not in METRIC_ENUM:
        errors.append(f"unknown metric: '{raw.metric}'")
    value = _parse_float(raw.value)
    period_iso = _normalize_period(raw.period)
    # RAW: prefix means period didn't match PERIOD_MAP — store as-is, no longer reject.
    # period_iso carries the raw label for non-standard periods.
    perimeter = raw.perimeter or "unknown"
    ec = raw.epistemic_class if raw.epistemic_class in EPISTEMIC_CLASS_ENUM else "asserted"
    if raw.epistemic_class not in EPISTEMIC_CLASS_ENUM:
        errors.append(f"invalid epistemic_class: '{raw.epistemic_class}'")
    if ec == "derived" and not (raw.derivation or "").strip():
        errors.append("derived claim missing derivation field")
    # The schema tells the model not to emit a CHARACTERISATION, and the model
    # labels one correctly and emits it anyway — observed on "low capital
    # expenditure", classified CHARACTERISATION and returned regardless.
    #
    # So the rule is enforced here instead of asked for. A deterministic filter
    # is the right home for it in any case: it cannot drift between runs, and the
    # claim lands in rejected_claims.json with a reason rather than silently
    # never existing. A seller's adjective kept out of the case is a decision
    # somebody can review; one that was never extracted is invisible.
    if str(raw.claim_kind or "").upper() == "CHARACTERISATION":
        errors.append(
            "characterisation without checkable content — a descriptor, not evidence"
        )
    claim_id = canonical_claim_id({
        "entity": raw.entity,
        "metric": raw.metric,
        "period": raw.period,
        "period_canonical": raw.period_canonical,
        "scope": raw.scope,
        "basis": raw.basis,
        "measurement": raw.measurement,
        "scenario": raw.scenario,
        "unit": raw.unit,
        "source_id": raw.source_id,
        "source_version_id": raw.source_version_id,
        "locator": raw.locator,
        "epistemic_class": ec,
        # Hash the value exactly as E3 will persist it. Non-numeric bounded or
        # textual quantities (for example "30-60 days") keep ``value_raw``;
        # hashing None here made their emitted IDs impossible to reproduce from
        # e3_claims.json.
        "value": value if value is not None else raw.value,
        "perimeter": perimeter,
    })
    return CanonicalClaim(
        claim_id=claim_id,
        statement=raw.statement,
        source_id=raw.source_id,
        locator=raw.locator,
        epistemic_class=ec,
        value=value,
        value_raw=raw.value,
        unit=raw.unit,
        definition_id=raw.definition_id,
        period=raw.period or "",
        period_iso=period_iso,
        perimeter=perimeter,
        ground_truth_flag=False,
        validation_only=False,
        notes=None,
        metric=raw.metric,
        source_path=raw.source_path,
        known_at=raw.known_at,
        source_version_id=raw.source_version_id,
        direction=raw.direction,
        topic=raw.topic,
        derivation=raw.derivation,
        author=raw.author,
        entity=raw.entity or "unspecified",
        period_canonical=raw.period_canonical or "none",
        scope=raw.scope or "unspecified",
        measurement=raw.measurement or "total",
        claim_kind=raw.claim_kind or "QUANTITATIVE",
        bound=raw.bound or "EXACT",
        basis=raw.basis or "unspecified",
        scenario=raw.scenario or "unspecified",
        validation_errors=errors,
    )


# ─────────────────────────────────────────────────────────────────────────────
# L4 — Graph assembler (deterministic)
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class SubGraph:
    claims: list[CanonicalClaim]
    conflicts: list[dict]
    rejected: list[dict]
    admitted_count: int
    rejected_count: int
    conflict_count: int


def assemble(claims: list[CanonicalClaim]) -> SubGraph:
    admitted: dict[str, CanonicalClaim] = {}
    conflicts: list[dict] = []
    rejected: list[dict] = []
    for c in claims:
        if c.validation_errors:
            rejected.append({
                "claim_id": c.claim_id,
                "metric": c.metric,
                "errors": c.validation_errors,
                "statement": c.statement[:80],
            })
            continue
        sid = c.claim_id
        if sid in admitted:
            existing = admitted[sid]
            if existing.value != c.value:
                conflicts.append({
                    "claim_id": sid,
                    "metric": c.metric,
                    "source_a": existing.source_id,
                    "value_a": existing.value,
                    "source_b": c.source_id,
                    "value_b": c.value,
                    "note": "Same stable_id, different values — check perimeter or period.",
                })
        else:
            admitted[sid] = c
    return SubGraph(
        claims=list(admitted.values()),
        conflicts=conflicts,
        rejected=rejected,
        admitted_count=len(admitted),
        rejected_count=len(rejected),
        conflict_count=len(conflicts),
    )


# ─────────────────────────────────────────────────────────────────────────────
# Output — E3 manifest format + graph.json for comparison
# ─────────────────────────────────────────────────────────────────────────────

def _to_e3_manifest(graph: SubGraph, deal: str, manifest: str,
                    sources_used: list[dict],
                    workbook_graph_summary: list[dict[str, Any]] | None = None) -> dict:
    """
    Produce an E3 extraction manifest.
    Claims are in CAP-003 format: only the frozen fields, no compiler metadata.
    Compiler metadata lives in the adjacent extraction_metadata section.
    """
    claims_output = []
    for c in graph.claims:
        claims_output.append({
            "claim_id": c.claim_id,
            "statement": c.statement,
            "source_id": c.source_id,
            "locator": c.locator,
            "epistemic_class": c.epistemic_class,
            "value": str(c.value) if c.value is not None else c.value_raw,
            "unit": c.unit,
            "definition_id": c.definition_id,
            "period": c.period,
            "perimeter": c.perimeter,
            "ground_truth_flag": c.ground_truth_flag,
            "validation_only": c.validation_only,
            "notes": c.notes,
        })
    return {
        "schema_version": "e3-1.0",
        "manifest_id": manifest,
        "deal": deal,
        "extractor": "extract_v2",
        "sources_included": [s["source_id"] for s in sources_used],
        "claims": claims_output,
        "conflicts": graph.conflicts,
        "extraction_metadata": {
            "admitted_count": graph.admitted_count,
            "rejected_count": graph.rejected_count,
            "conflict_count": graph.conflict_count,
            "rejected": graph.rejected,
            "compiler_fields_per_claim": [
                {
                    "claim_id": c.claim_id,
                    "metric": c.metric,
                    "known_at": c.known_at,
                    "source_version_id": c.source_version_id,
                    "direction": c.direction,
                    "topic": c.topic,
                    "derivation": c.derivation,
                    "author": c.author,
                    "entity": c.entity,
                    "period_canonical": c.period_canonical,
                    "scope": c.scope,
                    "measurement": c.measurement,
                    "claim_kind": c.claim_kind,
                    "bound": c.bound,
                    "basis": c.basis,
                    "scenario": c.scenario,
                }
                for c in graph.claims
            ],
            # Formula cells are captured deterministically at L1 and written
            # beside this manifest.  Keep the E3 lightweight, but make the
            # immutable workbook graph discoverable by the product/compiler.
            "workbook_formula_graphs": workbook_graph_summary or [],
        },
    }


def _capture_workbook_graphs(source_paths: list[Path],
                             source_records: dict[Path, dict] | None = None) -> dict[str, Any]:
    """Capture lossless formula graphs for every Open XML workbook in an intake.

    V2 used to feed formula text to the LLM but then discarded the deterministic
    cell/dependency graph.  That made a formula impossible to inspect or bind
    after an upload.  The sidecar is deliberately separate from E3 claims:
    formulas remain source evidence, never LLM-invented claims.
    """
    from tools.source_graph import capture

    workbooks: list[dict[str, Any]] = []
    for source_path in source_paths:
        if source_path.suffix.lower() not in {".xlsx", ".xlsm"}:
            continue
        graph = capture(source_path).to_json()
        stats = graph.get("stats", {})
        workbooks.append({
            "source_id": (source_records or {}).get(source_path, _source_record(source_path))["source_id"],
            "source_path": str(source_path),
            "source_filename": source_path.name,
            "graph": graph,
            "summary": {
                "formula_count": int(stats.get("by_kind", {}).get("formula", 0)),
                "precedent_edge_count": int(stats.get("precedent_edges", 0)),
                "defined_name_count": int(stats.get("defined_names", 0)),
                "cached_formula_value_count": sum(
                    1 for cell in graph.get("cells", {}).values()
                    if cell.get("kind") == "formula" and cell.get("cached_value") is not None
                ),
                "evaluated_formula_count": sum(
                    1 for cell in graph.get("cells", {}).values()
                    if cell.get("kind") == "formula"
                    and cell.get("evaluation_status") == "CALCULATED_ACYCLIC"
                ),
                "cyclic_formula_count": sum(
                    1 for cell in graph.get("cells", {}).values()
                    if cell.get("kind") == "formula"
                    and cell.get("evaluation_status") == "CYCLIC_COMPONENT"
                ),
            },
        })
    return {"schema": "workbook-formula-graphs-1.0", "workbooks": workbooks}


def _compare_graphs(new_claims: list[dict], old_path: Path) -> dict:
    with open(old_path) as f:
        old_graph = json.load(f)
    old_nodes = {n["id"]: n for n in old_graph.get("nodes", [])
                 if n.get("type") == "claim"}
    new_by_metric_period = {}
    for c in new_claims:
        key = (c.get("metric", c.get("claim_id", "")), c.get("period", ""), c.get("perimeter", ""))
        new_by_metric_period[key] = c
    old_by_key = {}
    for n in old_nodes.values():
        key = (n.get("metric", ""), n.get("as_of", ""), n.get("perimeter", ""))
        old_by_key[key] = n
    only_new = [c for k, c in new_by_metric_period.items() if k not in old_by_key]
    only_old = [n for k, n in old_by_key.items() if k not in new_by_metric_period]
    changes = []
    for k in new_by_metric_period:
        if k in old_by_key:
            nv = str(new_by_metric_period[k].get("value", ""))
            ov = str(old_by_key[k].get("value", ""))
            if nv != ov:
                changes.append({"metric": k[0], "period": k[1],
                                 "old_value": ov, "new_value": nv})
    return {
        "new_total": len(new_claims),
        "old_total": len(old_nodes),
        "only_in_new": len(only_new),
        "only_in_old": len(only_old),
        "value_changes": len(changes),
        "samples_only_new": [
            {"metric": c.get("metric", c.get("claim_id", "")),
             "value": c.get("value"), "statement": c.get("statement", "")[:80]}
            for c in only_new[:5]
        ],
        "samples_only_old": [
            {"metric": n.get("metric", ""), "value": n.get("value"),
             "statement": n.get("statement", "")[:80]}
            for n in only_old[:5]
        ],
        "value_change_samples": changes[:5],
    }


def _sha256_file(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _w(path: Path, obj: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(obj, indent=2, default=str), encoding="utf-8")
    size = path.stat().st_size
    sha = _sha256_file(path)
    print(f"  {path.name:<45} {size//1024:4d}KB  sha256:{sha[:16]}...")


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main() -> int:
    ap = argparse.ArgumentParser(description="extract_v2 / E3 pipeline")
    src_group = ap.add_mutually_exclusive_group(required=True)
    src_group.add_argument("--manifest", choices=["K-PRE", "K-IC", "K-LIVE", "ALL"],
                           help="Run over all sources in manifest (no temporal leakage)")
    src_group.add_argument(
        "--source",
        help=(
            "Single source file (.pdf, .docx, .pptx, .xlsx/.xlsm, .csv, "
            ".md/.txt/.html, .srt/.vtt, .eml/.mbox)"
        ),
    )
    ap.add_argument(
        "--input-dir",
        type=Path,
        help=(
            "Directory containing manifest source files. Defaults to vault/inbox; "
            "use this for external or sensitive corpora without copying them into the repo."
        ),
    )
    ap.add_argument("--deal", required=True, help="Deal slug (e.g. keystone)")
    ap.add_argument(
        "--source-envelope", type=Path,
        help="Path to a panta.source-envelope/1.0 JSON record for a --source intake.",
    )
    ap.add_argument("--output", default="pipeline_out/e3",
                    help="Output directory (default: pipeline_out/e3)")
    ap.add_argument("--dry-run", action="store_true",
                    help="Show chunks without calling LLM")
    ap.add_argument("--compare", default=None,
                    help="Path to existing graph.json to compare against")
    ap.add_argument("--workers", type=int, default=3,
                    help="Parallel LLM workers (default: 3)")
    ap.add_argument("--chunk-words", type=int, default=CHUNK_WORDS,
                    help=f"Words per chunk (default: {CHUNK_WORDS})")
    ap.add_argument(
        "--resume-partial-cache",
        action="store_true",
        help=(
            "Migrate a legacy/incomplete raw cache by treating chunks with "
            "existing claims as complete and retrying the rest"
        ),
    )
    ap.add_argument(
        "--via-cli",
        action="store_true",
        help=(
            "Call the model through the Claude Code CLI instead of the API "
            "(subscription-backed). The schema is instructed, not enforced — do "
            "not compare a score from this path against one from the API path"
        ),
    )
    ap.add_argument("--cli-model", default="haiku",
                    help="Model alias for --via-cli (default: haiku)")
    args = ap.parse_args()

    # ── Collect source paths ──────────────────────────────────────────────
    if args.manifest:
        source_dir = args.input_dir or VAULT_INBOX
        if not source_dir.is_absolute():
            source_dir = ROOT / source_dir
        source_paths = load_manifest(args.manifest, args.deal, source_dir)
        manifest_label = args.manifest
        if not source_paths:
            print(f"ERROR: No sources found for manifest {args.manifest} in {source_dir}",
                  file=sys.stderr)
            print(f"  Expected stems: {MANIFEST_SOURCES[args.manifest]}")
            return 1
        print(f"\n[Manifest {args.manifest}] {len(source_paths)} sources:")
        for p in source_paths:
            rec = _source_record(p)
            print(f"  {rec['source_id']:<16} {p.name}  (known_at: {rec['known_at']})")
    else:
        source_path = Path(args.source)
        if not source_path.exists():
            source_path = ROOT / args.source
        if not source_path.exists():
            print(f"ERROR: source not found: {args.source}", file=sys.stderr)
            return 1
        source_paths = [source_path]
        manifest_label = "SINGLE"

    source_records: dict[Path, dict] = {}
    if args.source_envelope:
        if args.manifest:
            print("ERROR: --source-envelope can only be used with --source", file=sys.stderr)
            return 2
        try:
            envelope = json.loads(args.source_envelope.read_text(encoding="utf-8"))
            if envelope.get("schema") != "panta.source-envelope/1.0":
                raise ValueError("unsupported source envelope schema")
            if str(envelope.get("case_id")) != str(args.deal):
                raise ValueError("source envelope case_id does not match --deal")
            source_records[source_paths[0]] = extractor_source_record(envelope)
        except (OSError, ValueError, json.JSONDecodeError) as exc:
            print(f"ERROR: invalid --source-envelope: {exc}", file=sys.stderr)
            return 2

    out_dir = ROOT / args.output / manifest_label
    out_dir.mkdir(parents=True, exist_ok=True)

    # ── L1: Parse all sources ─────────────────────────────────────────────
    print(f"\n[L1] Parsing {len(source_paths)} source(s)...")
    all_chunks: list[Chunk] = []
    for sp in source_paths:
        try:
            chunks = parse_source(sp, args.chunk_words, source_records.get(sp))
        except UnsupportedSourceError as exc:
            print(f"ERROR: {json.dumps(exc.to_dict(), sort_keys=True)}", file=sys.stderr)
            return 2
        print(f"  {sp.name:<50} {len(chunks):3d} chunks")
        all_chunks.extend(chunks)
    print(f"  Total: {len(all_chunks)} chunks  "
          f"(avg {sum(c.word_count for c in all_chunks)//max(len(all_chunks),1)} w/chunk)")

    # Persist the lossless workbook structure before L2.  This is essential
    # even for a dry run or a later LLM failure: the formula text, precedents
    # and cached values are facts contained in the uploaded source.
    workbook_graphs = _capture_workbook_graphs(source_paths, source_records)
    workbook_graph_path = out_dir / "workbook_formula_graphs.json"
    _w(workbook_graph_path, workbook_graphs)
    for workbook in workbook_graphs["workbooks"]:
        summary = workbook["summary"]
        print(
            f"  {workbook['source_filename']:<50} "
            f"{summary['formula_count']:5d} formulas  "
            f"{summary['precedent_edge_count']:5d} precedent edges"
        )

    chunks_debug = [
        {"chunk_id": c.chunk_id, "locator": c.locator,
         "source_id": c.source_record["source_id"],
         "provenance": c.provenance,
         "period_context": c.period_context,
         "section_heading": c.section_heading,
         "page_or_slide_number": c.page_or_slide_number,
         "word_count": c.word_count, "preview": c.body[:100].replace("\n", " ") + "..."}
        for c in all_chunks
    ]
    _w(out_dir / "chunks_debug.json", chunks_debug)

    if args.dry_run:
        print(f"\n[DRY-RUN] Chunks written to {out_dir}/chunks_debug.json")
        _print_chunk_summary(all_chunks)
        return 0

    # ── L2: Annotate ──────────────────────────────────────────────────────
    # Raw claims cache: completed modern runs are reused. Interrupted/failed modern
    # runs retain their successful chunks and retry only the outstanding chunk IDs.
    # A cache without the companion status file is treated as a legacy complete cache.
    raw_cache_path = out_dir / "raw_claims_cache.json"
    chunk_status_path = out_dir / "l2_chunk_status.json"
    failed_chunks_path = out_dir / "failed_chunks.json"
    cached_raw: list[RawClaim] = []
    completed_chunk_ids: set[str] = set()
    completed_chunk_models: dict[str, str] = {}
    pending_chunks = list(all_chunks)
    modern_cache = raw_cache_path.exists() and chunk_status_path.exists()

    if raw_cache_path.exists() and not modern_cache and not args.resume_partial_cache:
        print(f"\n[L2] Cache found — loading raw claims from {raw_cache_path.name} (skipping API calls)")
        cached = json.loads(raw_cache_path.read_text())
        all_raw = [RawClaim(**c) for c in cached]
        print(f"  Loaded {len(all_raw)} raw claims from cache")
    else:
        if raw_cache_path.exists() and args.resume_partial_cache and not modern_cache:
            cached = json.loads(raw_cache_path.read_text())
            cached_raw = [RawClaim(**c) for c in cached]
            cached_locators = {claim.locator for claim in cached_raw}
            completed_chunk_ids = {
                chunk.chunk_id
                for chunk in all_chunks
                if any(
                    locator == chunk.locator
                    or locator.startswith(chunk.locator + ":")
                    for locator in cached_locators
                )
            }
            pending_chunks = [
                chunk for chunk in all_chunks
                if chunk.chunk_id not in completed_chunk_ids
            ]
            print(
                f"\n[L2] Migrating partial legacy cache — "
                f"{len(completed_chunk_ids)}/{len(all_chunks)} chunks have claims, "
                f"{len(pending_chunks)} to retry"
            )
        elif modern_cache:
            cached = json.loads(raw_cache_path.read_text())
            cached_raw = [RawClaim(**c) for c in cached]
            status = json.loads(chunk_status_path.read_text())
            completed_chunk_ids = set(status.get("completed_chunk_ids", []))
            completed_chunk_models = {
                str(chunk_id): str(model)
                for chunk_id, model in status.get("completed_chunk_models", {}).items()
                if chunk_id in completed_chunk_ids
            }
            if len(completed_chunk_models) < len(completed_chunk_ids):
                existing_model = os.environ.get(
                    "PEOS_EXISTING_CACHE_MODEL",
                    "legacy-cache",
                ).strip() or "legacy-cache"
                for chunk_id in completed_chunk_ids:
                    completed_chunk_models.setdefault(chunk_id, existing_model)
            pending_chunks = [
                chunk for chunk in all_chunks
                if chunk.chunk_id not in completed_chunk_ids
            ]
            if not pending_chunks and status.get("complete"):
                all_raw = cached_raw
                print(
                    f"\n[L2] Complete chunk cache found — loaded "
                    f"{len(all_raw)} raw claims (skipping API calls)"
                )
            else:
                print(
                    f"\n[L2] Resuming partial cache — "
                    f"{len(completed_chunk_ids)}/{len(all_chunks)} chunks complete, "
                    f"{len(pending_chunks)} to retry"
                )

        if modern_cache and not pending_chunks and status.get("complete"):
            pass
        else:
            if args.via_cli:
                # Subscription-backed path for re-runs. The schema is instructed
                # rather than enforced here (the CLI has no tool_choice), so a
                # number produced this way must not be compared against one
                # produced via the API — see tools/llm_cli_provider.py.
                from tools.llm_cli_provider import CliClient
                client = CliClient(model=args.cli_model)
                print(f"\n[L2] via Claude CLI (model={args.cli_model}) — "
                      f"schema instructed, not enforced")
            else:
                api_key = configured_api_key()
                if not api_key:
                    print(f"ERROR: {missing_key_message()}.", file=sys.stderr)
                    print("  Or use --dry-run to inspect chunks.")
                    return 1
                try:
                    import anthropic
                    client = anthropic.Anthropic(**anthropic_client_kwargs(api_key))
                except ImportError:
                    print("ERROR: pip install anthropic", file=sys.stderr)
                    return 1
            print(f"\n[L2] Annotating {len(pending_chunks)} chunk(s) "
                  f"(workers={args.workers}, model={MODEL})...")
            all_raw = list(cached_raw)
            processed = 0
            batch_errors: list[dict[str, str]] = []

            def _checkpoint(*, complete: bool = False) -> None:
                from dataclasses import asdict

                raw_cache_path.write_text(
                    json.dumps([asdict(r) for r in all_raw], indent=2, default=str),
                    encoding="utf-8",
                )
                chunk_status_path.write_text(
                    json.dumps(
                        {
                            "schema_version": "l2-chunk-status-1.0",
                            "total_chunks": len(all_chunks),
                            "completed_chunk_ids": sorted(completed_chunk_ids),
                            "completed_chunk_models": dict(
                                sorted(completed_chunk_models.items())
                            ),
                            "models_used": sorted(set(completed_chunk_models.values())),
                            "failed_chunks": batch_errors,
                            "complete": complete,
                        },
                        indent=2,
                    ),
                    encoding="utf-8",
                )
                failed_chunks_path.write_text(
                    json.dumps(batch_errors, indent=2),
                    encoding="utf-8",
                )

            def _process(chunk: Chunk) -> tuple[Chunk, list[RawClaim]]:
                for attempt in range(MAX_PROVIDER_RETRIES + 1):
                    try:
                        return chunk, annotate_chunk(
                            chunk,
                            client,
                            args.deal,
                            raise_errors=True,
                        )
                    except Exception as exc:
                        delay = _provider_retry_delay(exc, attempt)
                        if delay is None or attempt >= MAX_PROVIDER_RETRIES:
                            raise
                        print(
                            f"  [L2 RETRY] {chunk.chunk_id}: "
                            f"{type(exc).__name__}; retry "
                            f"{attempt + 1}/{MAX_PROVIDER_RETRIES} in {delay:g}s",
                            file=sys.stderr,
                        )
                        time.sleep(delay)
                raise AssertionError("unreachable provider retry state")

            with ThreadPoolExecutor(max_workers=args.workers) as pool:
                futures = {pool.submit(_process, c): c for c in pending_chunks}
                for fut in as_completed(futures):
                    chunk = futures[fut]
                    processed += 1
                    try:
                        _, raw_claims = fut.result()
                    except Exception as exc:
                        message = str(exc).splitlines()[0][:1000]
                        batch_errors.append(
                            {
                                "chunk_id": chunk.chunk_id,
                                "locator": chunk.locator,
                                "error_type": type(exc).__name__,
                                "message": message,
                            }
                        )
                        print(
                            f"  [L2 ERROR] {chunk.chunk_id}: "
                            f"{type(exc).__name__}: {message}",
                            file=sys.stderr,
                        )
                        if _is_fatal_provider_error(exc):
                            for queued in futures:
                                if queued is not fut:
                                    queued.cancel()
                            print(
                                "  [L2] Fatal provider/key error; remaining "
                                "queued chunks were cancelled.",
                                file=sys.stderr,
                            )
                            break
                    else:
                        all_raw.extend(raw_claims)
                        completed_chunk_ids.add(chunk.chunk_id)
                        completed_chunk_models[chunk.chunk_id] = MODEL
                        if raw_claims:
                            print(f"  [{processed:03d}/{len(pending_chunks):03d}] "
                                  f"{chunk.locator[:55]:<55} → {len(raw_claims)} claim(s)")
                    if processed % 10 == 0:
                        _checkpoint()

            is_complete = (
                not batch_errors
                and len(completed_chunk_ids) == len(all_chunks)
            )
            _checkpoint(complete=is_complete)
            print(f"  Raw claims checkpointed → {raw_cache_path.name}")
            if not is_complete:
                outstanding = len(all_chunks) - len(completed_chunk_ids)
                print(
                    f"ERROR: L2 incomplete — {outstanding} chunk(s) outstanding; "
                    "rerun the same command to retry only outstanding chunks.",
                    file=sys.stderr,
                )
                return 3

    print(f"  Total raw: {len(all_raw)}")

    # ── L3: Validate ──────────────────────────────────────────────────────
    print("\n[L3] Validating...")
    canonicals = [validate(r) for r in all_raw]
    invalid = [c for c in canonicals if c.validation_errors]
    if invalid:
        print(f"  Rejected: {len(invalid)}")
        for c in invalid[:5]:
            print(f"    x {c.metric}: {c.validation_errors}")

    # ── L4: Assemble ──────────────────────────────────────────────────────
    print("\n[L4] Assembling...")
    graph = assemble(canonicals)
    print(f"  Admitted: {graph.admitted_count}  "
          f"Rejected: {graph.rejected_count}  "
          f"Conflicts: {graph.conflict_count}")
    if graph.conflicts:
        for conf in graph.conflicts:
            print(f"    conflict: {conf['metric']} {conf['value_a']} vs {conf['value_b']}")

    # ── Output ────────────────────────────────────────────────────────────
    print(f"\n[Output] Writing to {out_dir}/...")
    sources_used = [source_records.get(p, _source_record(p)) for p in source_paths]
    workbook_graph_summary = [
        {
            "source_id": item["source_id"],
            "source_filename": item["source_filename"],
            **item["summary"],
            "artifact": "workbook_formula_graphs.json",
        }
        for item in workbook_graphs["workbooks"]
    ]
    e3 = _to_e3_manifest(
        graph,
        args.deal,
        manifest_label,
        sources_used,
        workbook_graph_summary,
    )
    if chunk_status_path.exists():
        l2_status = json.loads(chunk_status_path.read_text())
        e3["extraction_metadata"]["l2_complete"] = bool(
            l2_status.get("complete")
        )
        e3["extraction_metadata"]["llm_models"] = list(
            l2_status.get("models_used", [])
        )
    _w(out_dir / "e3_claims.json", e3)
    _w(out_dir / "rejected_claims.json", graph.rejected)
    _w(out_dir / "conflict_report.json", graph.conflicts)

    # Metric distribution
    metric_counts: dict[str, int] = {}
    for c in graph.claims:
        metric_counts[c.metric] = metric_counts.get(c.metric, 0) + 1

    report_lines = [
        f"E3 Extraction Report — {manifest_label}",
        "=" * 50,
        f"  Deal        : {args.deal}",
        f"  Manifest    : {manifest_label}",
        f"  Sources     : {len(source_paths)}",
        f"  Chunks      : {len(all_chunks)}",
        f"  Raw claims  : {len(all_raw)}",
        f"  Admitted    : {graph.admitted_count}",
        f"  Rejected    : {graph.rejected_count}",
        f"  Conflicts   : {graph.conflict_count}",
        "",
        "Admitted by metric:",
    ]
    for metric, count in sorted(metric_counts.items(), key=lambda x: -x[1]):
        report_lines.append(f"  {count:3d}  {metric}")
    if graph.conflicts:
        report_lines += ["", "Conflicts:"]
        for conf in graph.conflicts:
            report_lines.append(
                f"  {conf['metric']}: {conf['value_a']} ({conf['source_a']}) "
                f"vs {conf['value_b']} ({conf['source_b']})"
            )
    report_text = "\n".join(report_lines) + "\n"
    (out_dir / "extraction_report.txt").write_text(report_text, encoding="utf-8")

    # Hash manifest for version pinning
    e3_path = out_dir / "e3_claims.json"
    manifest_record = {
        "manifest_id": manifest_label,
        "deal": args.deal,
        "sources": [s["source_id"] for s in sources_used],
        "admitted_count": graph.admitted_count,
        "e3_claims_sha256": _sha256_file(e3_path),
    }
    _w(out_dir / "manifest_hash.json", manifest_record)

    print()
    print(report_text)

    # ── Compare ───────────────────────────────────────────────────────────
    if args.compare:
        compare_path = Path(args.compare)
        if not compare_path.exists():
            compare_path = ROOT / args.compare
        if compare_path.exists():
            print(f"[Compare] vs {compare_path.name}...")
            diff = _compare_graphs(e3["claims"], compare_path)
            _w(out_dir / "comparison.json", diff)
            print(f"  New: {diff['new_total']}  Old: {diff['old_total']}  "
                  f"Only-new: {diff['only_in_new']}  Only-old: {diff['only_in_old']}  "
                  f"Changed: {diff['value_changes']}")

    return 0


def _print_chunk_summary(chunks: list[Chunk]) -> None:
    print(f"\n  Chunk summary ({len(chunks)} total):")
    for c in chunks[:12]:
        preview = c.body[:55].replace("\n", " ")
        print(f"    {c.source_record['source_id']:<14} {c.locator:<50} "
              f"{c.word_count:3d}w  \"{preview}...\"")
    if len(chunks) > 12:
        print(f"    ... and {len(chunks) - 12} more")


if __name__ == "__main__":
    sys.exit(main())
