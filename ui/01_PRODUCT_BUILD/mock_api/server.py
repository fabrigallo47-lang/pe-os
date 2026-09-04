#!/usr/bin/env python3
"""PANTA V20 stateful mock API.

Synthetic only. This server demonstrates the V20 public contracts, including
bitemporal projection/replay, governed authority, blocked components and
partial settlement. It never contacts an external system and refuses to
masquerade as a production CONNECTED backend.
"""
from __future__ import annotations

import argparse
import base64
import copy
import hashlib
import hmac
import io
import json
import os
import re
import secrets
import threading
import zipfile
from datetime import date, datetime, time, timedelta, timezone
from http.server import ThreadingHTTPServer, SimpleHTTPRequestHandler
from pathlib import Path
from urllib.parse import parse_qs, unquote, urlparse

ROOT = Path(__file__).resolve().parents[1]
APP = ROOT / "app"
FIXTURES = ROOT / "fixtures"
SESSION_DIR = ROOT / "mock_api" / ".sessions"
SESSION_DIR.mkdir(parents=True, exist_ok=True)
LOCK = threading.RLock()
API = "/api/v20"
LEGACY_API = "/api/v19"
BINDINGS_ROUTE = API + "/cases/{case_id}/bindings"
AUTHORITY_LEDGER_ID = "PANTA-AUTHORITY-LEDGER-V1"
AUTHORITY_RECORD_VERSION = "authority-record/2.0"
MOCK_SIGNATURE_ALGORITHM = "SYNTHETIC-HMAC-SHA256"
NON_ATTESTABLE_HUMAN_STOP_REASONS = frozenset(
    {
        "AUTHORITY_POLICY_UNRESOLVED",
        "BATCH_VALUE_CONFLICT",
        "CIRCULAR_SUPPORT",
        "IMMUTABLE_HISTORICAL_FIELD",
        "MISSING_RULE_PROVENANCE",
        "NON_WAIVABLE_AXIOM",
        "OBJECT_TYPE_MISMATCH",
        "PRIOR_VALUE_MISMATCH",
        "UNKNOWN_OBJECT_ID",
        "UNKNOWN_TARGET_POSITION_ID",
        "UPSTREAM_INPUT_BLOCKED",
    }
)


class SessionAuthenticationError(ValueError):
    pass


def utcnow() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def today_utc() -> str:
    return datetime.now(timezone.utc).date().isoformat()


def sha(value) -> str:
    return "sha256:" + hashlib.sha256(
        json.dumps(value, sort_keys=True, separators=(",", ":"), ensure_ascii=False).encode("utf-8")
    ).hexdigest()


def normalized_human_stop(stop: dict) -> dict:
    result = clone(stop)
    reason_code = str(result.get("reason_code") or "")
    required_role = str(
        result.get("required_role")
        or result.get("required_authority_level")
        or "PROFESSIONAL_REVIEWER"
    )
    result["required_role"] = required_role
    result.setdefault("required_authority_level", required_role)
    attestable = bool(
        stop.get("attestable") is not False
        and required_role != "PREPARER"
        and reason_code not in NON_ATTESTABLE_HUMAN_STOP_REASONS
    )
    result["attestable"] = attestable
    if reason_code == "NON_WAIVABLE_AXIOM":
        result["resolution_kind"] = "NON_WAIVABLE_BLOCK"
    elif reason_code == "AUTHORITY_POLICY_UNRESOLVED":
        result["resolution_kind"] = "POLICY_CORRECTION"
    elif attestable:
        result["resolution_kind"] = "AUTHORITY_ATTESTATION"
    else:
        result["resolution_kind"] = "INPUT_OR_MODEL_CORRECTION"
    result.setdefault("downstream_scope", [])
    return result


def authority_record_payload_hash(record: dict) -> str:
    return sha(
        {
            key: value
            for key, value in record.items()
            if key not in {"record_hash", "record_signature"}
        }
    )


def mock_authority_signing_key() -> bytes:
    path = SESSION_DIR / ".authority_signing_hmac.key"
    try:
        return path.read_bytes()
    except FileNotFoundError:
        key = secrets.token_bytes(32)
        try:
            descriptor = os.open(
                path, os.O_WRONLY | os.O_CREAT | os.O_EXCL, 0o600
            )
        except FileExistsError:
            return path.read_bytes()
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(key)
            handle.flush()
            os.fsync(handle.fileno())
        return key


def mock_sign_authority_record(record: dict) -> dict:
    key = mock_authority_signing_key()
    signed = clone(record)
    signed["signature_algorithm"] = MOCK_SIGNATURE_ALGORITHM
    signed["signing_key_id"] = "sha256:" + hashlib.sha256(key).hexdigest()
    signed["record_hash"] = authority_record_payload_hash(signed)
    signature = hmac.new(
        key, signed["record_hash"].encode("ascii"), hashlib.sha256
    ).digest()
    signed["record_signature"] = base64.urlsafe_b64encode(signature).decode(
        "ascii"
    ).rstrip("=")
    return signed


def verify_mock_authority_signature(record: dict) -> None:
    key = mock_authority_signing_key()
    if record.get("signature_algorithm") != MOCK_SIGNATURE_ALGORITHM:
        raise ValueError("Authority record signature algorithm is unsupported")
    if record.get("signing_key_id") != "sha256:" + hashlib.sha256(key).hexdigest():
        raise ValueError("Authority record signing key is not trusted")
    if record.get("record_hash") != authority_record_payload_hash(record):
        raise ValueError("Authority record immutable payload hash mismatch")
    expected = hmac.new(
        key, record["record_hash"].encode("ascii"), hashlib.sha256
    ).digest()
    try:
        encoded = str(record.get("record_signature") or "")
        supplied = base64.urlsafe_b64decode(
            (encoded + "=" * (-len(encoded) % 4)).encode("ascii")
        )
    except (ValueError, UnicodeEncodeError) as exc:
        raise ValueError("Authority record signature is malformed") from exc
    if not hmac.compare_digest(expected, supplied):
        raise ValueError("Authority record signature verification failed")


def normalize_reference_ids(value, field: str) -> list[str]:
    if not isinstance(value, list):
        raise ValueError(f"{field} must be an array")
    normalized = []
    seen = set()
    for raw in value:
        if not isinstance(raw, str) or not raw.strip():
            raise ValueError(f"{field} must contain non-empty strings")
        reference_id = raw.strip()
        if reference_id in seen:
            raise ValueError(f"Duplicate {field} entry: {reference_id}")
        normalized.append(reference_id)
        seen.add(reference_id)
    return normalized


def read_json(path):
    return json.loads(Path(path).read_text(encoding="utf-8"))


def write_json(path, value):
    Path(path).write_text(json.dumps(value, indent=2, ensure_ascii=False), encoding="utf-8")


def clone(value):
    return json.loads(json.dumps(value))


def uid(prefix):
    return f"{prefix}-{secrets.token_hex(6).upper()}"


def parse_iso(value: str, *, end_of_day: bool = False) -> datetime:
    if not value:
        raise ValueError("A bitemporal date is required")
    text = str(value).strip()
    try:
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
            d = date.fromisoformat(text)
            return datetime.combine(d, time.max if end_of_day else time.min, tzinfo=timezone.utc)
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=timezone.utc)
        return parsed.astimezone(timezone.utc)
    except Exception as exc:
        raise ValueError(f"Invalid ISO-8601 date/time: {value}") from exc


def known_on_or_before(obj: dict, cutoff: datetime) -> bool:
    known = obj.get("known_at") or obj.get("timestamp") or obj.get("created_at") or obj.get("ingested_at")
    if not known:
        return False
    try:
        return parse_iso(known) <= cutoff
    except ValueError:
        return False


def list_cases():
    return sorted(x.name for x in FIXTURES.iterdir() if x.is_dir() and (x / "manifest.json").exists())


def default_case():
    cases = list_cases()
    if not cases:
        raise FileNotFoundError("No fixture packs available")
    return cases[0]


def load_pack(case_id):
    folder = FIXTURES / case_id
    if not folder.exists():
        raise FileNotFoundError(case_id)
    return {name: read_json(folder / f"{name}.json") for name in ("projection", "registry", "transitions", "manifest")}


def session_path(session_id):
    if not re.fullmatch(r"SESSION-[A-Za-z0-9_-]{32,128}", str(session_id)):
        raise SessionAuthenticationError("PANTA session token is malformed")
    return SESSION_DIR / f"{session_id}.json"


def save(session):
    with LOCK:
        write_json(session_path(session["session_id"]), session)


def actor_for(projection, key):
    target = "USR-PARTNER" if str(key).lower() == "partner" else "USR-ASSOCIATE"
    actor = next(
        (item for item in projection.get("actor_directory", []) if item.get("actor_id") == target),
        None,
    ) or {
        "actor_id": target,
        "name": str(key).title(),
        "role": str(key).title(),
        "authority_verbs": [],
    }
    return {**actor, "display_name": actor.get("display_name") or actor.get("name") or actor["actor_id"]}


def new_session(case_id=None, actor="partner", mode="MOCK_CONNECTED"):
    if mode == "CONNECTED":
        raise RuntimeError("CONNECTED_BACKEND_NOT_CONFIGURED")
    case_id = case_id or default_case()
    pack = load_pack(case_id)
    session_id = "SESSION-" + secrets.token_urlsafe(32)
    projection = clone(pack["projection"])
    projection["deal"]["projection_id"] = uid("PROJ")
    created_at = utcnow()
    expires_at = (
        parse_iso(created_at) + timedelta(hours=8)
    ).isoformat().replace("+00:00", "Z")
    session = {
        "session_id": session_id,
        "case_id": case_id,
        "mode": mode,
        "actor_key": actor,
        "projection": projection,
        "initial_as_of_state_id": projection.get("deal", {}).get("as_of_state_id"),
        "registry": clone(pack["registry"]),
        "runs": {},
        "authority_records": {},
        "packages": {},
        "jobs": {},
        "notes": [],
        "idempotency": {},
        "proposal_reviews": {},
        "mission_runs": {},
        "active_lens_id": projection.get("deal", {}).get("default_lens_id"),
        "created_at": created_at,
        "expires_at": expires_at,
    }
    save(session)
    return session


def load_session(session_id="", case_id=None, actor="partner", mode="MOCK_CONNECTED"):
    if session_id:
        path = session_path(session_id)
        if not path.exists():
            raise SessionAuthenticationError("PANTA session is unknown or expired")
        session = read_json(path)
        expires_at = parse_iso(session.get("expires_at", ""))
        if expires_at <= datetime.now(timezone.utc):
            raise SessionAuthenticationError("PANTA session is unknown or expired")
        if case_id and session.get("case_id") != case_id:
            raise SessionAuthenticationError("Session does not belong to the requested case")
        return session
    return new_session(case_id or default_case(), actor, mode)


def latest_known_date(session) -> str:
    values = []
    for event in session.get("registry", []):
        if event.get("known_at"):
            values.append(parse_iso(event["known_at"]))
    return max(values).date().isoformat() if values else today_utc()


def context(session, projection=None, **overrides):
    projection = projection or session["projection"]
    actor = actor_for(projection, session.get("actor_key", "partner"))
    level = "INVESTMENT_COMMITTEE" if "partner" in actor.get("role", "").lower() else "PROFESSIONAL_REVIEW"
    deal = projection["deal"]
    result = {
        "mode": session["mode"],
        "action_capability": "SIMULATED_SERVER",
        "case_id": session["case_id"],
        "projection_id": deal.get("projection_id"),
        "projection_hash": sha(projection),
        "as_of_state_id": deal.get("as_of_state_id"),
        "as_of_date": deal.get("as_of_date") or latest_known_date(session),
        "authenticated_actor": actor,
        "authentication": {
            "principal_id": actor["actor_id"],
            "case_id": session["case_id"],
            "authentication_method": "SYNTHETIC_SERVER_SESSION",
            "session_issued_at": session["created_at"],
            "session_expires_at": session["expires_at"],
        },
        "viewer_projection": "associate",
        "authority_assignments": [
            {
                "actor_id": actor["actor_id"],
                "authority_level": level,
                "verbs": actor.get("authority_verbs", []),
                "active": True,
            }
        ],
        "demo_session_id": session["session_id"],
        "synthetic": True,
        "no_external_effects": True,
        "contract_version": "20.0",
        "active_lens_id": deal.get("active_lens_id") or deal.get("default_lens_id"),
    }
    result.update(overrides)
    return result


def append_registry(
    session,
    kind,
    label,
    detail,
    actor="Mock Service",
    object_id=None,
    run_id=None,
    *,
    effective_date=None,
    known_at=None,
    epistemic_class="institutional_act",
):
    known_at = known_at or utcnow()
    effective_date = effective_date or parse_iso(known_at).date().isoformat()
    event = {
        "event_id": uid("REG"),
        "timestamp": known_at,
        "effective_date": effective_date,
        "known_at": known_at,
        "epistemic_class": epistemic_class,
        "kind": kind,
        "label": label,
        "detail": detail,
        "actor_id": actor,
        "actor_label": actor,
        "object_id": object_id,
        "case_id": session["case_id"],
        "run_id": run_id,
        "provenance": "mock_server",
        "synthetic": True,
    }
    session["registry"].append(event)
    return event


def idempotent(session, scope, key, payload, operation):
    if not key:
        raise ValueError("Idempotency-Key is required")
    token = f"{scope}:{key}"
    payload_hash = sha(payload)
    prior = session["idempotency"].get(token)
    if prior:
        if prior.get("payload_hash") != payload_hash:
            raise ValueError("IDEMPOTENCY_CONFLICT: the key was already used with a different payload")
        return clone(prior["result"])
    result = operation()
    session["idempotency"][token] = {"payload_hash": payload_hash, "result": clone(result)}
    return result


def event_key(pack, event_id):
    return next(
        (key for key, value in pack["projection"].get("events", {}).items() if value.get("event_id") == event_id),
        None,
    )


def find_object(projection, object_id):
    deal = projection.get("deal", {})
    lists = [
        deal.get("question_spine", []),
        deal.get("artifacts", []),
        deal.get("claims", []),
        deal.get("positions", []),
        deal.get("model_nodes", []),
        deal.get("cells", []),
        deal.get("rooms", {}).get("unknowns", {}).get("items", []),
        deal.get("rooms", {}).get("foundations", {}).get("sets", []),
        deal.get("rooms", {}).get("shadowIC", {}).get("theses", []),
        deal.get("source_center", {}).get("sources", []),
        deal.get("participants", []),
        deal.get("interactions", []),
        deal.get("utterances", []),
        deal.get("discrepancy_candidates", []),
        deal.get("derivations", []),
        deal.get("hypotheses", []),
        deal.get("agent_missions", []),
        deal.get("spine_change_proposals", []),
        deal.get("validation_envelopes", []),
        deal.get("lenses", []),
    ]
    for items in lists:
        for item in items:
            if object_id in (
                item.get("id"),
                item.get("claim_id"),
                item.get("position_id"),
                item.get("model_node_id"),
                item.get("cell_id"),
                item.get("source_id"),
                item.get("issue_id"),
                item.get("participant_id"),
                item.get("interaction_id"),
                item.get("utterance_id"),
                item.get("discrepancy_id"),
                item.get("derivation_id"),
                item.get("hypothesis_id"),
                item.get("mission_id"),
                item.get("proposal_id"),
                item.get("validation_envelope_id"),
                item.get("lens_id"),
            ):
                return item
    for artifact in deal.get("artifacts", []):
        for cell in artifact.get("cell_index", []):
            if object_id in (cell.get("id"), cell.get("cell_id"), cell.get("model_node_id")):
                return cell
    return None


def search_index(projection, query):
    query = str(query or "").lower().strip()
    output = []
    deal = projection.get("deal", {})
    groups = [
        (deal.get("question_spine", []), "QUESTION", "deal-command"),
        (deal.get("artifacts", []), "ARTIFACT", "artifacts"),
        (deal.get("claims", []), "CLAIM", "sources"),
        (deal.get("rooms", {}).get("unknowns", {}).get("items", []), "UNKNOWN", "unknowns"),
        (deal.get("model_nodes", []), "MODEL_NODE", "artifacts"),
        (deal.get("source_center", {}).get("sources", []), "SOURCE", "sources"),
        (deal.get("participants", []), "PERSON", "sources"),
        (deal.get("interactions", []), "INTERACTION", "sources"),
        (deal.get("utterances", []), "UTTERANCE", "sources"),
        (deal.get("discrepancy_candidates", []), "DISCREPANCY", "sources"),
        (deal.get("derivations", []), "DERIVATION", "sources"),
        (deal.get("hypotheses", []), "HYPOTHESIS", "sources"),
        (deal.get("agent_missions", []), "AGENT_MISSION", "work"),
        (deal.get("spine_change_proposals", []), "SPINE_CHANGE", "sources"),
        (deal.get("validation_envelopes", []), "VALIDATION_ENVELOPE", "foundations"),
        (deal.get("lenses", []), "LENS", "deal-command"),
    ]
    for values, label, route in groups:
        for item in values:
            object_id = (
                item.get("id") or item.get("claim_id") or item.get("model_node_id") or item.get("source_id") or item.get("interaction_id") or item.get("utterance_id") or item.get("discrepancy_id") or item.get("derivation_id") or item.get("hypothesis_id") or item.get("mission_id") or item.get("proposal_id") or item.get("validation_envelope_id") or item.get("lens_id")
            )
            text = " ".join(str(item.get(key, "")) for key in ("label", "title", "statement", "locator", "period", "perimeter", "name", "objective", "reason", "verbatim_text", "normalized_text", "description"))
            if not query or query in text.lower():
                output.append(
                    {
                        "id": object_id,
                        "type": label,
                        "label": item.get("label") or item.get("title") or item.get("name") or item.get("statement", "")[:90] or object_id,
                        "route": route,
                        "search_text": text,
                    }
                )
    return output[:50]


GENERIC_STOPWORDS = {
    "about", "after", "against", "before", "could", "current", "does", "from",
    "have", "into", "investment", "should", "their", "there", "these", "this",
    "under", "what", "when", "where", "which", "with", "would", "case", "deal",
}


def _semantic_tokens(value):
    return {
        token for token in re.findall(r"[a-z0-9]+", str(value or "").lower())
        if len(token) >= 4 and token not in GENERIC_STOPWORDS
    }


def question_bindings(projection, text):
    """Bind text only through question-provided semantics; never through case IDs in core code."""
    questions = projection.get("deal", {}).get("question_spine", [])
    lower = str(text or "").lower()
    input_tokens = _semantic_tokens(lower)
    ranked = []
    for question in questions:
        terms = set(str(item).lower() for item in question.get("semantic_keywords", []) if item)
        terms |= _semantic_tokens(question.get("label"))
        terms |= _semantic_tokens(question.get("description"))
        exact_hits = sum(1 for term in terms if term and term in lower)
        token_hits = len(input_tokens & _semantic_tokens(" ".join(terms)))
        score = exact_hits * 3 + token_hits
        if score:
            ranked.append((score, question.get("id")))
    ranked.sort(key=lambda item: (-item[0], str(item[1])))
    return [item[1] for item in ranked if item[1]]

def derive_replay_snapshots(registry, cutoff=None):
    events = []
    for event in registry:
        if not event.get("effective_date") or not event.get("known_at"):
            continue
        if cutoff is not None and parse_iso(event["known_at"]) > cutoff:
            continue
        events.append(clone(event))
    events.sort(key=lambda item: (parse_iso(item["known_at"]), item.get("event_id", "")))
    snapshots = []
    known, believed, approved, open_items = [], [], [], []
    for index, event in enumerate(events, start=1):
        kind = str(event.get("kind", "")).upper()
        summary = event.get("label") or event.get("detail") or event.get("event_id")
        if kind in {"SOURCE", "INGEST", "INGEST_REQUESTED", "COMPILER", "ORIGIN"}:
            known.append(summary)
        elif kind in {"AUTHORITY", "IC_RECORD", "DECISION"}:
            approved.append(summary)
        elif kind in {"COVERAGE", "BLOCK", "HUMAN_STOP"}:
            open_items.append(summary)
        else:
            believed.append(summary)
        state_id = event.get("result_state_id") or f"STATE-{event['event_id']}"
        snapshots.append(
            {
                "id": state_id,
                "event_id": event["event_id"],
                "date": event["effective_date"],
                "effective_date": event["effective_date"],
                "known_at": event["known_at"],
                "label": summary,
                "known": known[-4:],
                "believed": believed[-4:],
                "approved": approved[-4:],
                "open": open_items[-4:],
                "event_index": index,
                "derived_from_event_log": True,
                "stable_hash": sha(event),
            }
        )
    return snapshots



def _level(value):
    return {"VERY LOW": 1, "LOW": 2, "MEDIUM": 3, "HIGH": 4, "VERY HIGH": 5}.get(str(value or "").upper(), 3)


def _review_status(session, kind, object_id, default="PROPOSED"):
    review = session.get("proposal_reviews", {}).get(f"{kind}:{object_id}")
    return review.get("decision", default) if review else default


def generate_derivations(session, deal):
    claims = {item.get("claim_id") or item.get("id"): item for item in deal.get("claims", [])}
    output = []
    for spec in deal.get("derivation_specs", []):
        method = spec.get("method_type")
        params = clone(spec.get("parameters", {}))
        value = None
        unit = params.get("output_unit") or ""
        assumptions = []
        try:
            if method == "CIRCLE_AREA":
                claim = claims[params.get("radius_claim_id")]
                radius_m = float(claim.get("value"))
                value = 3.141592653589793 * (radius_m / 1000.0) ** 2 * 100.0
                unit = "hectares"
                assumptions = ["Circular unobstructed geometry", "Nominal range is a radius", "No overlap or environmental loss"]
            elif method == "IMPLIED_RADIUS_FROM_AREA_AND_NODE_COUNT":
                site_area_ha = float(params["site_area_ha"])
                node_count = float(params["node_count"])
                value = ((site_area_ha / node_count / 100.0) / 3.141592653589793) ** 0.5 * 1000.0
                unit = "m"
                assumptions = ["Equal non-overlapping circular coverage", "Node count reflects effective coverage rather than redundancy", "Site area is accurate"]
            elif method == "RUNWAY_MONTHS":
                value = float(params["cash_eur_m"]) / float(params["burn_eur_m"])
                unit = "months"
                assumptions = ["Burn remains constant", "No financing or grant receipts", "No working-capital discontinuity"]
            elif method == "POST_MONEY_OWNERSHIP":
                pre = float(params["pre_money_eur_m"])
                new = float(params["new_money_eur_m"])
                value = new / (pre + new) * 100.0
                unit = "%"
                assumptions = ["No option-pool top-up", "No convertibles or SAFE conversion", "Fully diluted basis"]
            else:
                continue
        except (KeyError, TypeError, ValueError, ZeroDivisionError):
            continue
        derivation_id = spec.get("derivation_id") or str(spec.get("derivation_spec_id", "DER-SPEC")).replace("-SPEC", "")
        value = round(value, 2)
        method_payload = {"method_type": method, "formula": spec.get("formula"), "parameters": params, "input_claim_ids": spec.get("input_claim_ids", [])}
        result = {
            "derivation_id": derivation_id,
            "id": derivation_id,
            "label": spec.get("label"),
            "statement": f"{spec.get('label')}: {value} {unit}",
            "method_type": method,
            "formula": spec.get("formula"),
            "input_claim_ids": spec.get("input_claim_ids", []),
            "assumptions": assumptions,
            "value": value,
            "unit": unit,
            "definition_id": spec.get("output_definition_id"),
            "period": spec.get("period"),
            "perimeter": spec.get("perimeter"),
            "question_ids": spec.get("question_ids", []),
            "bears_on": spec.get("question_ids", []),
            "epistemic_class": "derived",
            "review_status": _review_status(session, "derivation", derivation_id),
            "propagation_eligible": _review_status(session, "derivation", derivation_id) == "ADMITTED",
            "method_version": "panta-deterministic-derivation/1.0",
            "method_hash": sha(method_payload),
            "source_trace": spec.get("input_claim_ids", []),
            "effective_date": spec.get("effective_date"),
            "known_at": spec.get("known_at"),
            "synthetic": True,
        }
        output.append(result)
    return output


def _candidate_id(rule, object_ids, suffix=None):
    explicit = rule.get("candidate_id")
    if explicit:
        return explicit
    prefix = rule.get("candidate_id_prefix") or "DISC"
    payload = {"rule_id": rule.get("rule_id"), "objects": object_ids, "suffix": suffix}
    return f"{prefix}-{sha(payload).split(':')[-1][:12].upper()}"


def generate_discrepancies(session, deal, derivations):
    """Generate review candidates from declarative fixture/compiler rules.

    The core runtime contains no deal-specific definitions, IDs or conclusions. A candidate
    never changes institutional truth without professional review.
    """
    objects = list(deal.get("claims", [])) + list(derivations)
    candidates = []
    for rule in deal.get("discrepancy_rules", []):
        rule_type = rule.get("type")
        definition_id = rule.get("definition_id")
        if rule_type == "NUMERIC_INCOMPATIBILITY":
            matches = [item for item in objects if item.get("definition_id") == definition_id and isinstance(item.get("value"), (int, float))]
            if len(matches) < 2:
                continue
            values = [float(item["value"]) for item in matches]
            minimum = min(abs(value) for value in values)
            ratio = max(abs(value) for value in values) / max(minimum, 1e-9)
            threshold = float(rule.get("threshold", 1.0))
            if rule.get("comparison", "RATIO_GT") == "RATIO_GT" and ratio <= threshold:
                continue
            object_ids = [item.get("id") or item.get("claim_id") or item.get("derivation_id") for item in matches]
            did = _candidate_id(rule, object_ids)
            candidates.append({
                "discrepancy_id": did, "id": did, "type": rule_type,
                "label": rule.get("label") or "Materially incompatible values share one semantic definition",
                "definition_id": definition_id, "object_ids": object_ids,
                "values": [{"object_id": item.get("id") or item.get("claim_id") or item.get("derivation_id"), "value": item.get("value"), "unit": item.get("unit"), "perimeter": item.get("perimeter"), "epistemic_class": item.get("epistemic_class"), "source_id": item.get("source_id")} for item in matches],
                "ratio": round(ratio, 2), "aligned_dimensions": rule.get("aligned_dimensions", []),
                "non_aligned_dimensions": rule.get("non_aligned_dimensions", rule.get("dimensions", [])),
                "detector_version": "panta-discrepancy/1.0", "confidence": rule.get("confidence", "REVIEW_REQUIRED"),
                "review_status": _review_status(session, "discrepancy", did), "automatic_truth_change": False,
                "reason": rule.get("reason") or "A professional must determine whether the compared records share a valid perimeter.",
                "hypothesis_templates": rule.get("hypothesis_templates", []),
                "effective_date": max((item.get("effective_date", "") for item in matches), default=""),
                "known_at": max((item.get("known_at", "") for item in matches), default=""),
            })
        elif rule_type == "EXACT_CONTRADICTION":
            matches = [item for item in objects if item.get("definition_id") == definition_id and item.get("value") is not None]
            normalized = {json.dumps(item.get("value"), sort_keys=True) for item in matches}
            if len(matches) < 2 or len(normalized) < 2:
                continue
            object_ids = [item.get("id") or item.get("claim_id") or item.get("derivation_id") for item in matches]
            did = _candidate_id(rule, object_ids)
            candidates.append({
                "discrepancy_id": did, "id": did, "type": rule_type,
                "label": rule.get("label") or "The same semantic definition carries incompatible values",
                "definition_id": definition_id, "object_ids": object_ids,
                "values": [{"object_id": item.get("id") or item.get("claim_id") or item.get("derivation_id"), "value": item.get("value"), "unit": item.get("unit"), "perimeter": item.get("perimeter"), "source_id": item.get("source_id")} for item in matches],
                "detector_version": "panta-discrepancy/1.0", "confidence": rule.get("confidence", "REVIEW_REQUIRED"),
                "review_status": _review_status(session, "discrepancy", did), "automatic_truth_change": False,
                "reason": rule.get("reason") or "A professional must reconcile the incompatible representations.",
                "hypothesis_templates": rule.get("hypothesis_templates", []),
                "effective_date": max((item.get("effective_date", "") for item in matches), default=""),
                "known_at": max((item.get("known_at", "") for item in matches), default=""),
            })
        elif rule_type == "CONDITIONAL_INVALIDATION":
            relation = rule.get("edge_relation", "CONDITIONS")
            failed_statuses = set(rule.get("failed_statuses", ["NOT_SATISFIED", "FAILED"]))
            for edge in deal.get("condition_edges", []):
                if edge.get("relation") != relation or edge.get("condition_status") not in failed_statuses:
                    continue
                object_ids = [edge.get("from_condition_object_id"), edge.get("to_position_id")]
                did = _candidate_id(rule, object_ids, edge.get("edge_id"))
                candidates.append({
                    "discrepancy_id": did, "id": did, "type": rule_type,
                    "label": rule.get("label") or "A position is invalid under an explicit condition",
                    "definition_id": None, "object_ids": object_ids, "condition_edge_id": edge.get("edge_id"),
                    "detector_version": "panta-discrepancy/1.0", "confidence": rule.get("confidence", "DETERMINISTIC_EDGE"),
                    "review_status": _review_status(session, "discrepancy", did), "automatic_truth_change": False,
                    "reason": edge.get("condition_definition") or rule.get("reason"), "failure_effect": edge.get("failure_effect"),
                    "hypothesis_templates": rule.get("hypothesis_templates", []),
                    "effective_date": edge.get("effective_date"), "known_at": edge.get("known_at"),
                })
    return candidates


def generate_hypotheses(session, discrepancies):
    """Convert rule-supplied explanation prompts into reviewable hypotheses, never facts."""
    output = []
    for discrepancy in discrepancies:
        for index, template in enumerate(discrepancy.get("hypothesis_templates", []), start=1):
            text = template.get("label") if isinstance(template, dict) else str(template)
            hid = (template.get("hypothesis_id") if isinstance(template, dict) else None) or f"{discrepancy['id']}-H{index}"
            status = _review_status(session, "hypothesis", hid)
            output.append({
                "hypothesis_id": hid, "id": hid, "label": text,
                "discrepancy_id": discrepancy["id"], "input_object_ids": discrepancy.get("object_ids", []),
                "origin": "AI_REASONING_PROPOSAL", "review_status": status, "epistemic_class": "asserted",
                "propagation_eligible": status == "ADMITTED", "automatic_truth_change": False,
                "effective_date": discrepancy.get("effective_date"), "known_at": discrepancy.get("known_at"),
            })
    return output


def apply_lens(session, deal, requested_lens_id=None):
    lenses = deal.get("lenses", [])
    if not lenses:
        return
    lens_id = requested_lens_id or session.get("active_lens_id") or deal.get("default_lens_id") or lenses[0].get("lens_id")
    lens = next((item for item in lenses if (item.get("lens_id") or item.get("id")) == lens_id), None)
    if not lens and requested_lens_id:
        raise ValueError("Unknown lens_id")
    if not lens:
        fallback_id = deal.get("default_lens_id")
        lens = next((item for item in lenses if (item.get("lens_id") or item.get("id")) == fallback_id), None) or lenses[0]
        lens_id = lens.get("lens_id") or lens.get("id")
    session["active_lens_id"] = lens_id
    deal["active_lens_id"] = lens_id
    order = {qid: index for index, qid in enumerate(lens.get("question_order", []))}
    deal["question_spine"] = sorted(deal.get("question_spine", []), key=lambda item: (order.get(item.get("id"), 9999), item.get("id", "")))
    for question in deal.get("question_spine", []):
        question["lens_required"] = question.get("id") in set(lens.get("required_question_ids", []))
        question["active_lens_id"] = lens_id
    qorder = {item.get("id"): index for index, item in enumerate(deal.get("question_spine", []))}
    unknowns = deal.get("rooms", {}).get("unknowns", {}).get("items", [])
    unknowns.sort(key=lambda item: (qorder.get(item.get("question_id"), 9999), item.get("rank", 9999)))
    for index, item in enumerate(unknowns, start=1):
        item["lens_rank"] = index
    deal["work_items"] = sorted(deal.get("work_items", []), key=lambda item: (qorder.get(item.get("question_id"), 9999), item.get("id", "")))
    deal["lens_projection"] = {"lens_id": lens_id, "label": lens.get("label"), "description": lens.get("description"), "ranking_weights": lens.get("ranking_weights"), "required_question_ids": lens.get("required_question_ids", []), "facts_unchanged": True, "facts_hash": sha(deal.get("claims", []))}


def enrich_v20_projection(session, deal):
    derivations = generate_derivations(session, deal)
    discrepancies = generate_discrepancies(session, deal, derivations)
    hypotheses = generate_hypotheses(session, discrepancies)
    deal["derivations"] = derivations
    deal["discrepancy_candidates"] = discrepancies
    deal["hypotheses"] = hypotheses
    for collection, kind in ((deal.get("spine_change_proposals", []), "spine"), (deal.get("agent_missions", []), "mission")):
        for item in collection:
            oid = item.get("proposal_id") or item.get("mission_id") or item.get("id")
            if kind == "spine": item["status"] = _review_status(session, "spine", oid, item.get("status", "PROPOSED"))
            elif f"mission:{oid}" in session.get("proposal_reviews", {}): item["status"] = session["proposal_reviews"][f"mission:{oid}"]["decision"]

def projection_as_of(session, as_of_date=None, lens_id=None):
    if as_of_date:
        cutoff = parse_iso(as_of_date, end_of_day=True)
        display_date = cutoff.date().isoformat()
    else:
        display_date = latest_known_date(session)
        cutoff = parse_iso(display_date, end_of_day=True)
    projection = clone(session["projection"])
    deal = projection["deal"]
    deal["claims"] = [item for item in deal.get("claims", []) if known_on_or_before(item, cutoff)]
    for key in ("participants", "interactions", "utterances", "validation_envelopes", "agent_missions", "spine_change_proposals", "condition_edges", "derivation_specs", "lenses"):
        if isinstance(deal.get(key), list):
            deal[key] = [item for item in deal[key] if not item.get("known_at") or known_on_or_before(item, cutoff)]
    enrich_v20_projection(session, deal)
    apply_lens(session, deal, lens_id)
    canonical = projection.get("v16", {}).get("canonical", {})
    if isinstance(canonical.get("claims"), list):
        canonical["claims"] = [item for item in canonical["claims"] if known_on_or_before(item, cutoff)]
    source_center = deal.get("source_center", {})
    source_center["sources"] = [item for item in source_center.get("sources", []) if known_on_or_before(item, cutoff)]
    for source in source_center.get("sources", []):
        if isinstance(source.get("versions"), list):
            source["versions"] = [item for item in source["versions"] if known_on_or_before(item, cutoff)]
    projection["events"] = {
        key: value for key, value in projection.get("events", {}).items() if known_on_or_before(value, cutoff)
    }
    filtered_registry = [item for item in session.get("registry", []) if known_on_or_before(item, cutoff)]
    snapshots = derive_replay_snapshots(session.get("registry", []), cutoff)
    deal.setdefault("replay", {})["snapshots"] = snapshots
    deal["replay"]["source"] = "REGISTRY_EVENTS"
    deal["replay"]["hand_authored_snapshots"] = False
    deal["as_of_date"] = display_date
    # The date selector governs the knowledge projection. Institutional state identity
    # advances only through a recorded event that explicitly carries result_state_id.
    # A replay snapshot ID is never substituted for Current state.
    state_events = [
        item for item in filtered_registry
        if item.get("result_state_id") and item.get("known_at")
    ]
    state_events.sort(key=lambda item: (parse_iso(item["known_at"]), item.get("event_id", "")))
    deal["as_of_state_id"] = (
        state_events[-1]["result_state_id"]
        if state_events
        else session.get("initial_as_of_state_id")
        or session.get("projection", {}).get("deal", {}).get("as_of_state_id")
        or f"STATE-ORIGIN-{display_date}"
    )
    deal["projection_id"] = "PROJ-ASOF-" + sha({"case_id": session["case_id"], "date": display_date})[-16:]
    temporal = deal.setdefault("temporal", {})
    temporal.update(
        {
            "basis": "KNOWN_AT",
            "effective_axis": "effective_date",
            "knowledge_axis": "known_at",
            "replay_source": "REGISTRY_EVENTS",
            "available_dates": sorted(
                {
                    parse_iso(item["known_at"]).date().isoformat()
                    for item in session.get("registry", [])
                    if item.get("known_at")
                }
                | {
                    parse_iso(item["known_at"]).date().isoformat()
                    for item in session["projection"]["deal"].get("claims", [])
                    if item.get("known_at")
                }
                | {
                    parse_iso(item["known_at"]).date().isoformat()
                    for key in ("interactions", "utterances", "agent_missions", "spine_change_proposals")
                    for item in session["projection"]["deal"].get(key, [])
                    if item.get("known_at")
                }
            ),
        }
    )
    return projection, filtered_registry


def safe_name(v): return re.sub(r'[^A-Za-z0-9._-]+','_',str(v or 'source'))[:120]
def xml_text(data):
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as z:
            texts=[]
            for n in z.namelist():
                if n.endswith('.xml') and (n.startswith('word/') or n.startswith('ppt/slides/')):
                    raw=z.read(n).decode('utf-8','ignore'); texts += re.findall(r'<[^>]*?t[^>]*>(.*?)</[^>]+>',raw)
            return ' '.join(re.sub(r'<[^>]+>',' ',x) for x in texts)
    except Exception:return ''
def parse_native_source(job,s):
    raw=base64.b64decode(job.get('content_b64') or '') if job.get('content_b64') else b''
    name=safe_name(job.get('file_name') or job.get('source_name') or 'source'); ext=Path(name).suffix.lower(); result={'asset_url':None,'viewer':'METADATA','claims':[],'cells':[],'artifact':None,'stats':{}}
    if raw:
        folder=APP/'assets'/'uploads'/s['session_id']; folder.mkdir(parents=True,exist_ok=True); (folder/name).write_bytes(raw); result['asset_url']=f'/assets/uploads/{s["session_id"]}/{name}'
    p=s['projection']
    if ext in ('.xlsx','.xlsm') and raw:
        try:
            from openpyxl import load_workbook
            wb=load_workbook(io.BytesIO(raw),data_only=False,read_only=False); formula_count=0; nonempty=0; cells=[]; claims=[]
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is None: continue
                        nonempty+=1; is_formula=isinstance(cell.value,str) and cell.value.startswith('='); formula_count+=int(is_formula)
                        left=ws.cell(cell.row,max(1,cell.column-1)).value if cell.column>1 else None; above=ws.cell(max(1,cell.row-1),cell.column).value if cell.row>1 else None
                        label=str(left or above or '').strip(); locator=f'{ws.title}!{cell.coordinate}'
                        if len(cells)<250 and (is_formula or label):
                            cells.append({'cell_id':f'CELL:{locator}','id':f'CELL:{locator}','sheet':ws.title,'cell':cell.coordinate,'label':label or locator,'value':None if is_formula else cell.value,'unit':'','formula':cell.value if is_formula else None,'precedents':re.findall(r"(?:'([^']+)'|([A-Za-z0-9_ ]+))?!?\$?([A-Z]{1,3})\$?([0-9]+)",cell.value) if is_formula else [],'dependents':[],'period':'Workbook-defined','perimeter':'Workbook-defined','workbook_id':job.get('source_id'),'version_id':job.get('version_id'),'locator':locator})
                        text=f'{label}: {cell.value}' if label else ''
                        bindings=question_bindings(p,text) if text else []
                        if label and bindings and len(claims)<30:
                            claims.append({'statement':text,'locator':locator,'value':None if is_formula else cell.value,'unit':'','period':'Workbook-defined','perimeter':'Workbook-defined','topic':'Workbook extraction','direction':'context','bears_on':bindings})
            result.update({'viewer':'WORKBOOK','cells':cells,'claims':claims,'stats':{'sheets':len(wb.sheetnames),'nonempty_cells':nonempty,'formulas':formula_count}})
            aid=uid('ART'); result['artifact']={'id':aid,'type':'Excel model','title':job.get('source_name') or name,'current_version':'V1','status':'current','owner':'Compiler','trace_ids':[c['cell_id'] for c in cells[:20]],'versions':[{'version':'V1','date':utcnow(),'cause':'Native workbook ingested','status':'current'}],'preview_kind':'table','preview_rows':[{'label':c['label'],'value':str(c['value'] if c['value'] is not None else c['formula']),'source':c['locator'],'state':'Extracted'} for c in cells[:8]],'cell_index':cells[:100]}
        except Exception as e: result['stats']={'parse_error':str(e)}
    elif ext=='.pdf' and raw:
        result['viewer']='PDF'
        try:
            from pypdf import PdfReader
            rd=PdfReader(io.BytesIO(raw)); result['stats']={'pages':len(rd.pages)}
            for pi,page in enumerate(rd.pages[:5],1):
                text=(page.extract_text() or '').strip(); chunks=[x.strip() for x in re.split(r'\n{2,}|(?<=[.!?])\s+',text) if len(x.strip())>35]
                for ch in chunks[:3]: result['claims'].append({'statement':ch[:500],'locator':f'Page {pi}','value':None,'unit':'','period':'Source-defined','perimeter':'Source-defined','topic':'Document extraction','direction':'context','bears_on':question_bindings(p,ch)})
        except Exception as e: result['stats']={'parse_error':str(e)}
    elif ext in ('.docx','.pptx') and raw:
        result['viewer']='DOCUMENT'; text=xml_text(raw); result['stats']={'characters':len(text)}
        for i,ch in enumerate([x.strip() for x in re.split(r'(?<=[.!?])\s+',text) if len(x.strip())>40][:12],1): result['claims'].append({'statement':ch[:500],'locator':f'Object {i}','value':None,'unit':'','period':'Source-defined','perimeter':'Source-defined','topic':'Document extraction','direction':'context','bears_on':question_bindings(p,ch)})
    elif ext in ('.vtt','.srt','.transcript','.txt') and raw and re.search(r'(?m)^(?:\[?\d{1,2}:\d{2}|[A-Za-z][A-Za-z ._-]{1,40}:)', raw.decode('utf-8','ignore')):
        result['viewer']='TRANSCRIPT'; text=raw.decode('utf-8','ignore'); result['stats']={'characters':len(text),'native_interaction':True}
        iid=uid('INT'); participants={}; utterances=[]
        pattern=re.compile(r'(?m)^(?:\[?(?P<time>\d{1,2}:\d{2}(?::\d{2})?)\]?\s*)?(?P<speaker>[A-Za-z][A-Za-z ._-]{1,40}):\s*(?P<text>.+)$')
        for idx,match in enumerate(pattern.finditer(text),1):
            speaker=match.group('speaker').strip(); pid='PER-'+safe_name(speaker).upper(); participants[pid]={'participant_id':pid,'id':pid,'name':speaker,'role':'Source participant','affiliation':'Source-defined','party':'UNKNOWN','independence_group':pid,'relationship':'Pending professional review','effective_date':today_utc(),'known_at':utcnow()}
            locator=(match.group('time') or f'Utterance {idx}')
            utt={'utterance_id':uid('UTT'),'id':uid('UTTREF'),'interaction_id':iid,'speaker_id':pid,'speaker_role':'Source participant','party':'UNKNOWN','locator':locator,'verbatim_text':match.group('text')[:1000],'normalized_text':match.group('text')[:1000],'attribution_confidence':'REVIEW_REQUIRED','text_status':'VERBATIM_UNREVIEWED','effective_date':today_utc(),'known_at':utcnow()}; utt['id']=utt['utterance_id']; utterances.append(utt)
            result['claims'].append({'statement':match.group('text')[:500],'locator':f'{iid} · {utt["utterance_id"]} · {locator}','value':None,'unit':'','period':'Interaction date','perimeter':'Statement content; pending semantic review','topic':'Interaction extraction','direction':'context','bears_on':question_bindings(p,match.group('text')),'speaker_id':pid,'interaction_id':iid,'utterance_id':utt['utterance_id'],'asserting_actor_id':pid})
        result['interaction']={'interaction_id':iid,'id':iid,'interaction_type':'UPLOADED_TRANSCRIPT','title':job.get('source_name') or name,'start_at':utcnow(),'end_at':utcnow(),'participant_ids':list(participants),'organizer_id':None,'channel':'Uploaded transcript','source_id':job.get('source_id'),'source_version_id':job.get('version_id'),'transcript_status':'VERBATIM_UNREVIEWED','speaker_identification_confidence':'REVIEW_REQUIRED','consent_status':'NOT_RECORDED_IN_SOURCE','confidentiality_class':'REVIEW_REQUIRED','model_processing_allowed':True,'export_allowed':False,'effective_date':today_utc(),'known_at':utcnow()}
        result['participants']=list(participants.values()); result['utterances']=utterances
    elif raw:
        result['viewer']='TEXT'; text=raw.decode('utf-8','ignore'); result['stats']={'characters':len(text)}
        for i,ch in enumerate([x.strip() for x in text.splitlines() if len(x.strip())>25][:20],1): result['claims'].append({'statement':ch[:500],'locator':f'Line {i}','value':None,'unit':'','period':'Source-defined','perimeter':'Source-defined','topic':'Text extraction','direction':'context','bears_on':question_bindings(p,ch)})
    return result



def complete_job(session, job):
    projection = session["projection"]
    source_center = projection["deal"].setdefault("source_center", {})
    sources = source_center.setdefault("sources", [])
    name = job.get("source_name") or job.get("value") or "Uploaded source"
    source_id = uid("SRC")
    version_id = f"{source_id}-V1"
    known_at = utcnow()
    effective_date = job.get("effective_date") or parse_iso(known_at).date().isoformat()
    job["source_id"] = source_id
    job["version_id"] = version_id
    parsed = parse_native_source(job, session)
    raw = base64.b64decode(job.get("content_b64") or "") if job.get("content_b64") else json.dumps(job, sort_keys=True).encode()
    source = {
        "source_id": source_id,
        "name": name,
        "type": job.get("source_type") or (Path(job.get("file_name") or "").suffix.lstrip(".").upper() or "DOCUMENT"),
        "status": "INGESTED",
        "latest_version_id": version_id,
        "claim_count": len(parsed["claims"]) or 1,
        "ingested_at": known_at,
        "effective_date": effective_date,
        "known_at": known_at,
        "versions": [
            {
                "version_id": version_id,
                "created_at": known_at,
                "effective_date": effective_date,
                "known_at": known_at,
                "sha256": hashlib.sha256(raw).hexdigest(),
                "supersedes": None,
            }
        ],
        "file_name": job.get("file_name"),
        "viewer": parsed["viewer"],
        "asset_url": parsed.get("asset_url"),
        "synthetic": True,
        "native_stats": parsed.get("stats", {}),
    }
    if parsed.get("interaction"):
        parsed["interaction"]["source_id"] = source_id
        parsed["interaction"]["source_version_id"] = version_id
        source["interaction_id"] = parsed["interaction"]["interaction_id"]
        source["participant_ids"] = parsed["interaction"].get("participant_ids", [])
        source["speaker_level_provenance"] = True
        projection["deal"].setdefault("interactions", []).append(parsed["interaction"])
        projection["deal"].setdefault("participants", []).extend(parsed.get("participants", []))
        projection["deal"].setdefault("utterances", []).extend(parsed.get("utterances", []))
    sources.append(source)
    first_question = (projection["deal"].get("question_spine") or [{}])[0].get("id")
    new_claims = []
    raw_claims = parsed["claims"] or [
        {
            "statement": f"The newly ingested source '{name}' contains a decision-relevant statement pending professional review.",
            "locator": "Mock extraction · object 1",
            "value": None,
            "unit": None,
            "period": "As of ingest",
            "perimeter": "Source-defined; pending semantic review",
            "topic": "New source",
            "direction": "context",
            "bears_on": [first_question] if first_question else [],
        }
    ]
    for raw_claim in raw_claims:
        claim_id = uid("CL")
        claim = {
            "claim_id": claim_id,
            "id": claim_id,
            "source_id": source_id,
            "epistemic_class": "asserted",
            "definition_id": None,
            "notes": None,
            "ground_truth_flag": False,
            "validation_only": False,
            "position_ids": [],
            "author": source_id,
            "review_status": "REVIEW_REQUIRED",
            "effective_date": effective_date,
            "known_at": known_at,
            **raw_claim,
        }
        new_claims.append(claim)
    projection["deal"].setdefault("claims", []).extend(new_claims)
    if parsed.get("cells"):
        projection["deal"].setdefault("cells", []).extend(parsed["cells"])
    if parsed.get("artifact"):
        projection["deal"].setdefault("artifacts", []).append(parsed["artifact"])
    unbound = [claim["claim_id"] for claim in new_claims if not claim.get("bears_on")]
    if unbound:
        source_center.setdefault("pipeline_issues", []).append(
            {
                "issue_id": uid("PIPE"),
                "severity": "REVIEW",
                "kind": "QUESTION_BINDING",
                "label": f"{len(unbound)} new claim(s) require question binding",
                "object_ids": unbound,
                "recommended_action": "Review applicability and bind or mark as context.",
            }
        )
    job.update(
        {
            "status": "COMPLETE",
            "progress": 100,
            "stage": "PROJECTED",
            "stage_label": "Projection refreshed",
            "message": f"Source version created with {len(new_claims)} claim(s) and {len(parsed.get('cells', []))} inspectable cell objects.",
            "claim_ids": [claim["claim_id"] for claim in new_claims],
            "completed_at": known_at,
            "effective_date": effective_date,
            "known_at": known_at,
            "asset_url": parsed.get("asset_url"),
            "stats": parsed.get("stats", {}),
        }
    )
    projection["deal"]["projection_id"] = uid("PROJ")
    append_registry(
        session,
        "INGEST",
        "Source ingestion completed",
        name,
        "Compiler",
        source_id,
        effective_date=effective_date,
        known_at=known_at,
        epistemic_class="institutional_act",
    )


def build_package(session, course, authority_record):
    execution_room = session["projection"]["deal"]["executionRoom"]
    execution = course.get("execution") or {}
    known_at = utcnow()
    package = {
        "execution_package_id": uid("EXEC"),
        "run_id": authority_record["run_id"],
        "candidate_state_id": authority_record["candidate_state_id"],
        "course_id": course["id"],
        "authority_record_id": authority_record["authority_record_id"],
        "package_type": execution_room.get("type", "Execution package"),
        "destination": execution_room.get("recipient"),
        "sender": execution_room.get("sender"),
        "subject": execution_room.get("subject"),
        "message": execution_room.get("message"),
        "attachments": execution_room.get("attachments", []),
        "checks": execution_room.get("checks", []),
        "amount": execution.get("amount"),
        "currency": execution.get("currency"),
        "document_version": execution.get("document_version"),
        "status": "READY",
        "created_at": known_at,
        "effective_date": parse_iso(known_at).date().isoformat(),
        "known_at": known_at,
        "synthetic": True,
        "no_external_effects": True,
    }
    package["artifact_hash"] = sha(package)
    session["packages"][package["execution_package_id"]] = package
    return package


def valid_change_ids(transition):
    return {
        item.get("artifact_id")
        for item in transition.get("artifact_change_sets", [])
        if item.get("artifact_id")
    }


def authority_record_binding(transition: dict, stop: dict) -> dict:
    authority_resolution = transition.get("authority_resolution", {})
    if not isinstance(authority_resolution, dict):
        authority_resolution = {}
    normalized_stop = normalized_human_stop(stop)
    return {
        "human_stop_hash": sha(normalized_stop),
        "policy_refs_hash": sha(transition.get("policy_refs", {})),
        "authority_resolution_hash": sha(authority_resolution),
        "policy_rule_id": normalized_stop.get("policy_rule_id"),
        "authority_resolution_rule_id": authority_resolution.get("selected_rule_id"),
    }


def mock_authority_assignment_snapshot(
    actor: dict, required_role: str, captured_at: str
) -> dict:
    actor_id = str(actor.get("actor_id") or "")
    actor_role = str(actor.get("role") or "")
    granted_roles = sorted(
        {
            actor_role,
            required_role,
            *(str(item) for item in actor.get("authority_roles", []) if item),
        }
        - {""}
    )
    authority_verbs = sorted(
        {str(item) for item in actor.get("authority_verbs", []) if item}
    )
    declared_effective_from = actor.get("effective_from") or actor.get("effective_date")
    declared_known_at = actor.get("known_at")
    assignment_identity = {
        "actor_id": actor_id,
        "actor_role": actor_role,
        "granted_roles": granted_roles,
        "authority_verbs": authority_verbs,
        "declared_effective_from": declared_effective_from,
        "declared_effective_to": actor.get("effective_to") or actor.get("expires_at"),
        "declared_known_at": declared_known_at,
        "assignment_version": str(actor.get("authority_assignment_version") or actor.get("version") or "1"),
    }
    return {
        "assignment_id": str(
            actor.get("authority_assignment_id")
            or actor.get("assignment_id")
            or "ASSIGN-" + sha(assignment_identity).removeprefix("sha256:")[:20].upper()
        ),
        "assignment_version": assignment_identity["assignment_version"],
        "actor_id": actor_id,
        "actor_role": actor_role,
        "granted_roles": granted_roles,
        "authority_verbs": authority_verbs,
        "effective_from": declared_effective_from or captured_at,
        "effective_to": assignment_identity["declared_effective_to"],
        "known_at": declared_known_at or captured_at,
        "assignment_status": "ACTIVE",
        "revoked_effective_at": actor.get("revoked_effective_at"),
        "revocation_known_at": actor.get("revocation_known_at"),
        "temporal_basis": (
            "DECLARED_ASSIGNMENT"
            if declared_effective_from and declared_known_at
            else "ATTESTATION_TIME_FALLBACK"
        ),
        "captured_at": captured_at,
    }


def validate_mock_authority_ledger(session: dict) -> None:
    previous_hash = None
    records = sorted(
        session.get("authority_records", {}).values(),
        key=lambda item: item.get("ledger_sequence", 0),
    )
    for expected_sequence, record in enumerate(records, start=1):
        if (
            record.get("ledger_id") != AUTHORITY_LEDGER_ID
            or record.get("ledger_sequence") != expected_sequence
            or record.get("previous_record_hash") != previous_hash
        ):
            raise ValueError("Authority ledger hash chain is discontinuous")
        verify_mock_authority_signature(record)
        previous_hash = record["record_hash"]


def validate_authority_record(session: dict, run: dict, record: dict) -> None:
    transition = run["transition"]
    stop = next(
        (
            item
            for item in transition.get("human_stops", [])
            if item.get("stop_id") == record.get("human_stop_id")
        ),
        None,
    )
    if not stop or normalized_human_stop(stop)["attestable"] is not True:
        raise ValueError("Authority record references a non-attestable Human Stop")
    required_role = normalized_human_stop(stop)["required_role"]
    verb = stop.get("authority_verb") or session["projection"]["deal"]["decisionRoom"].get("verb")
    expected = {
        "authority_record_version": AUTHORITY_RECORD_VERSION,
        "case_id": session.get("case_id"),
        "run_id": run.get("run_id"),
        "candidate_state_id": run.get("candidate_state_id"),
        "human_stop_id": stop.get("stop_id"),
        "artifact_hash": transition.get("replay_hash"),
        "prepared_selection_hash": run.get("prepared_selection_hash"),
        "required_role": required_role,
        "authority_verb": verb,
        "status": "ATTESTED",
        **authority_record_binding(transition, stop),
    }
    if any(record.get(key) != value for key, value in expected.items()):
        raise ValueError("Authority record is not scoped to this prepared Candidate")
    if record.get("record_hash") != authority_record_payload_hash(record):
        raise ValueError("Authority record immutable payload hash mismatch")
    snapshot = record.get("authority_assignment")
    if not isinstance(snapshot, dict):
        raise ValueError("Authority assignment snapshot is missing")
    if record.get("authority_assignment_hash") != sha(snapshot):
        raise ValueError("Authority assignment snapshot hash mismatch")
    if (
        snapshot.get("actor_id") != record.get("actor_id")
        or snapshot.get("actor_role") != record.get("actor_role")
        or snapshot.get("captured_at") != record.get("timestamp")
        or required_role not in snapshot.get("granted_roles", [])
        or verb not in snapshot.get("authority_verbs", [])
    ):
        raise ValueError("Authority assignment snapshot does not authorize the act")
    authentication = record.get("authentication_context")
    if not isinstance(authentication, dict):
        raise ValueError("Authority authentication context is missing")
    if record.get("authentication_context_hash") != sha(authentication):
        raise ValueError("Authority authentication context hash mismatch")
    if (
        authentication.get("principal_id") != record.get("actor_id")
        or authentication.get("case_id") != session.get("case_id")
        or authentication.get("authentication_method") != "SYNTHETIC_SERVER_SESSION"
        or authentication.get("authenticated_at") != record.get("timestamp")
        or not re.fullmatch(
            r"sha256:[0-9a-f]{64}",
            str(authentication.get("session_id_hash") or ""),
        )
    ):
        raise ValueError("Authority authentication context does not authorize the act")
    issued_at = parse_iso(authentication.get("session_issued_at", ""))
    expires_at = parse_iso(authentication.get("session_expires_at", ""))
    authenticated_at = parse_iso(authentication.get("authenticated_at", ""))
    if not issued_at <= authenticated_at <= expires_at:
        raise ValueError("Authority act falls outside its authenticated session interval")
    verify_mock_authority_signature(record)
    course = next(
        (
            item
            for item in session["projection"]["deal"]["decisionRoom"].get("courses", [])
            if item.get("id") == record.get("course_id")
        ),
        None,
    )
    if not course or record.get("effect_type") != course.get("effect_type", "INTERNAL"):
        raise ValueError("Authority record course is no longer admissible")


def settle(session, body):
    for field in ("run_id", "candidate_state_id", "selected_change_ids", "as_of_state_id"):
        if body.get(field) in (None, ""):
            raise ValueError(f"Missing settlement field: {field}")
    selected = normalize_reference_ids(
        body.get("selected_change_ids", []), "selected_change_ids"
    )
    if not selected:
        raise ValueError("At least one selected change is required")
    run = session["runs"].get(body["run_id"])
    if not run:
        raise ValueError("Run not found")
    if run.get("status") != "PREPARED":
        raise ValueError("Run must be PREPARED before settlement")
    if run.get("candidate_state_id") != body["candidate_state_id"]:
        raise ValueError("Candidate is stale or does not belong to this run")
    if body.get("as_of_state_id") != session["projection"]["deal"].get("as_of_state_id"):
        raise ValueError("As-of state is stale")
    transition = run["transition"]
    validate_mock_authority_ledger(session)
    allowed = valid_change_ids(transition)
    if not set(selected).issubset(allowed):
        raise ValueError("Settlement references a change outside the calculated affected set")
    if set(selected) != set(run.get("selected_change_ids", [])):
        raise ValueError("Settlement selection differs from the prepared change set")

    normalize_reference_ids(body.get("human_stop_ids", []), "human_stop_ids")
    authority_ids = normalize_reference_ids(
        body.get("authority_record_ids", []), "authority_record_ids"
    )
    if body.get("authority_record_id"):
        if body["authority_record_id"] in authority_ids:
            raise ValueError("Duplicate authority_record_ids entry")
        authority_ids.append(body["authority_record_id"])
    records = [session["authority_records"].get(item) for item in authority_ids]
    for record in records:
        if record and record.get("run_id") == run["run_id"]:
            validate_authority_record(session, run, record)
    stops = transition.get("human_stops", [])
    for stop in stops:
        matching = [
            record
            for record in records
            if record
            and record.get("run_id") == run["run_id"]
            and record.get("candidate_state_id") == run["candidate_state_id"]
            and record.get("human_stop_id") == stop.get("stop_id")
            and record.get("status") == "ATTESTED"
        ]
        if not matching:
            raise ValueError(f"Human Stop {stop.get('stop_id')} requires a scoped authority record")

    package_ids = normalize_reference_ids(
        body.get("execution_package_ids", []), "execution_package_ids"
    )
    if body.get("execution_package_id"):
        if body["execution_package_id"] in package_ids:
            raise ValueError("Duplicate execution_package_ids entry")
        package_ids.append(body["execution_package_id"])
    for record in records:
        if record and record.get("effect_type") == "EXTERNAL_PACKAGE":
            packages = [session["packages"].get(item) for item in package_ids]
            if not any(
                package
                and package.get("authority_record_id") == record["authority_record_id"]
                and package.get("run_id") == run["run_id"]
                and package.get("candidate_state_id") == run["candidate_state_id"]
                and package.get("status") == "ACCEPTED"
                for package in packages
            ):
                raise ValueError("External execution package lacks a scoped server acknowledgment")

    blocked = transition.get("blocked_components", [])
    if blocked and not body.get("allow_partial_settlement"):
        raise ValueError("Blocked components require an explicit bounded partial settlement")

    projection = session["projection"]
    prior_state_id = projection["deal"].get("as_of_state_id")
    known_at = utcnow()
    effective_date = body.get("effective_date") or parse_iso(known_at).date().isoformat()
    current_state_id = uid("STATE")
    projection["deal"]["as_of_state_id"] = current_state_id
    projection["deal"]["as_of_date"] = parse_iso(known_at).date().isoformat()
    projection["deal"]["projection_id"] = uid("PROJ")
    projection["deal"].setdefault("branches", {})["current"] = f"Settled {parse_iso(known_at).date().isoformat()}"
    run["status"] = "SETTLED"
    registry_event = append_registry(
        session,
        "SETTLEMENT",
        "Selected changes settled",
        f"{len(selected)} explicit change set(s); {len(blocked)} blocked component(s) preserved.",
        body.get("actor_id", "Settlement Service"),
        current_state_id,
        body["run_id"],
        effective_date=effective_date,
        known_at=known_at,
        epistemic_class="institutional_act",
    )
    registry_event["source_state_id"] = prior_state_id
    registry_event["result_state_id"] = current_state_id
    registry_event["candidate_state_id"] = body["candidate_state_id"]
    registry_event["selected_change_ids"] = selected
    registry_event["partial"] = bool(blocked)
    result_projection, filtered_registry = projection_as_of(session, projection["deal"]["as_of_date"])
    return {
        "settlement_id": uid("SETTLE"),
        "case_id": session["case_id"],
        "run_id": body["run_id"],
        "candidate_state_id": body["candidate_state_id"],
        "prior_state_id": prior_state_id,
        "current_state_id": current_state_id,
        "approved_state_id": projection["deal"].get("branches", {}).get("approved"),
        "selected_change_ids": selected,
        "partial": bool(blocked),
        "blocked_components": clone(blocked),
        "summary": "Current state updated; Approved history preserved; blocked scope remains explicit.",
        "replay_hash": sha(registry_event),
        "timestamp": known_at,
        "effective_date": effective_date,
        "known_at": known_at,
        "as_of_state_id": current_state_id,
        "as_of_date": projection["deal"]["as_of_date"],
        "projection": result_projection,
        "context": context(session, result_projection),
        "registry": filtered_registry,
    }


class Handler(SimpleHTTPRequestHandler):
    server_version = "PANTA-V20-Mock/1.0"

    def translate_path(self, path):
        relative = urlparse(path).path
        if relative.startswith("/api/"):
            return str(APP / "__not_file__")
        return str(APP / (relative.lstrip("/") or "index.html"))

    def end_headers(self):
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Headers", "Content-Type,Idempotency-Key,X-Panta-Session")
        self.send_header("Access-Control-Allow-Methods", "GET,POST,OPTIONS")
        self.send_header("Cache-Control", "no-store")
        super().end_headers()

    def do_OPTIONS(self):
        self.send_response(204)
        self.end_headers()

    def json_response(self, status, payload):
        encoded = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(encoded)))
        self.end_headers()
        self.wfile.write(encoded)

    def body(self):
        length = int(self.headers.get("Content-Length", "0"))
        raw = self.rfile.read(length) if length else b"{}"
        try:
            return json.loads(raw or b"{}")
        except Exception as exc:
            raise ValueError("Invalid JSON body") from exc

    def session(self, query, case_id=None, actor="partner", mode="MOCK_CONNECTED"):
        query_session = (query.get("session_id") or [""])[0].strip()
        header_session = self.headers.get("X-Panta-Session", "").strip()
        if query_session and header_session and not secrets.compare_digest(query_session, header_session):
            raise SessionAuthenticationError("Session header and query token do not match")
        session_id = header_session or query_session
        return load_session(session_id, case_id, actor, mode)

    def do_GET(self):
        parsed = urlparse(self.path)
        if not parsed.path.startswith("/api/"):
            return super().do_GET()
        try:
            path = API + parsed.path[len(LEGACY_API):] if parsed.path.startswith(LEGACY_API) else parsed.path
            self.get_api(path, parse_qs(parsed.query))
        except FileNotFoundError as exc:
            self.json_response(404, {"error": {"code": "NOT_FOUND", "message": str(exc)}})
        except SessionAuthenticationError as exc:
            self.json_response(401, {"error": {"code": "AUTHENTICATION_REQUIRED", "message": str(exc)}})
        except ValueError as exc:
            self.json_response(400, {"error": {"code": "INVALID_TEMPORAL_QUERY", "message": str(exc)}})
        except Exception as exc:
            self.json_response(500, {"error": {"code": "MOCK_SERVER_ERROR", "message": str(exc)}})

    def do_POST(self):
        parsed = urlparse(self.path)
        if not parsed.path.startswith("/api/"):
            return self.json_response(405, {"error": {"code": "METHOD_NOT_ALLOWED", "message": "POST is only available under the API"}})
        try:
            path = API + parsed.path[len(LEGACY_API):] if parsed.path.startswith(LEGACY_API) else parsed.path
            self.post_api(path, parse_qs(parsed.query), self.body())
        except SessionAuthenticationError as exc:
            self.json_response(401, {"error": {"code": "AUTHENTICATION_REQUIRED", "message": str(exc)}})
        except ValueError as exc:
            self.json_response(400, {"error": {"code": "INVALID_REQUEST", "message": str(exc)}})
        except Exception as exc:
            self.json_response(500, {"error": {"code": "MOCK_SERVER_ERROR", "message": str(exc)}})

    def get_api(self, path, query):
        if path == f"{API}/bootstrap":
            case_id = query.get("case_id", [default_case()])[0]
            actor = query.get("actor", ["partner"])[0]
            mode = query.get("mode", ["MOCK_CONNECTED"])[0]
            if mode == "CONNECTED":
                return self.json_response(
                    503,
                    {
                        "error": {
                            "code": "CONNECTED_BACKEND_NOT_CONFIGURED",
                            "message": "The bundled V20 server is synthetic Mock Connected only. It will not return fixture data for CONNECTED mode.",
                        }
                    },
                )
            try:
                session = self.session(query, case_id, actor, mode)
            except SessionAuthenticationError:
                session = new_session(case_id, actor, mode)
            save(session)
            projection, _ = projection_as_of(session, latest_known_date(session), query.get("lens_id", [""])[0] or None)
            return self.json_response(
                200,
                {
                    "session_id": session["session_id"],
                    "available_cases": list_cases(),
                    "context": context(session, projection),
                },
            )

        match = re.fullmatch(API + r"/cases/([^/]+)/projection", path)
        if match:
            session = self.session(query, unquote(match.group(1)))
            as_of_date = query.get("as_of_date", [""])[0] or latest_known_date(session)
            projection, registry = projection_as_of(session, as_of_date, query.get("lens_id", [""])[0] or None)
            save(session)
            return self.json_response(200, {"projection": projection, "context": context(session, projection), "registry": registry})

        match = re.fullmatch(API + r"/cases/([^/]+)/search", path)
        if match:
            session = self.session(query, unquote(match.group(1)))
            projection, _ = projection_as_of(session, query.get("as_of_date", [""])[0] or latest_known_date(session))
            return self.json_response(200, {"results": search_index(projection, query.get("q", [""])[0])})

        match = re.fullmatch(API + r"/cases/([^/]+)/(claims|questions|cells)/(.+)", path)
        if match:
            case_id, kind, object_id = unquote(match.group(1)), match.group(2), unquote(match.group(3))
            session = self.session(query, case_id)
            projection, _ = projection_as_of(session, query.get("as_of_date", [""])[0] or latest_known_date(session))
            obj = find_object(projection, object_id)
            if not obj:
                return self.json_response(404, {"error": {"code": "OBJECT_NOT_FOUND", "message": f"{kind[:-1].title()} not found"}})
            return self.json_response(200, {kind[:-1]: obj})

        match = re.fullmatch(API + r"/cases/([^/]+)/(coverage|bindings|compiler-report)", path)
        if match:
            case_id, kind = unquote(match.group(1)), match.group(2)
            session = self.session(query, case_id)
            projection, _ = projection_as_of(session, query.get("as_of_date", [""])[0] or latest_known_date(session))
            deal = projection["deal"]
            if kind == "coverage":
                output = {"coverage": deal.get("coverage", {}), "pipeline_issues": deal.get("source_center", {}).get("pipeline_issues", [])}
            elif kind == "bindings":
                output = {
                    "claim_question_bindings": [
                        {"claim_id": claim.get("claim_id"), "question_ids": claim.get("bears_on", [])}
                        for claim in deal.get("claims", [])
                    ],
                    "position_model_bindings": deal.get("position_model_bindings", []),
                }
            else:
                output = {
                    "report": {
                        "claims": len(deal.get("claims", [])),
                        "questions": len(deal.get("question_spine", [])),
                        "bound_claims": sum(bool(claim.get("bears_on")) for claim in deal.get("claims", [])),
                        "sources": len(deal.get("source_center", {}).get("sources", [])),
                        "pipeline_issues": deal.get("source_center", {}).get("pipeline_issues", []),
                        "projection_hash": context(session, projection)["projection_hash"],
                        "as_of_date": deal.get("as_of_date"),
                        "interactions": len(deal.get("interactions", [])),
                        "utterances": len(deal.get("utterances", [])),
                        "generated_discrepancies": len(deal.get("discrepancy_candidates", [])),
                        "deterministic_derivations": len(deal.get("derivations", [])),
                        "hypotheses_pending_review": sum(item.get("review_status") == "PROPOSED" for item in deal.get("hypotheses", [])),
                        "spine_change_proposals": len(deal.get("spine_change_proposals", [])),
                    }
                }
            return self.json_response(200, output)

        match = re.fullmatch(API + r"/cases/([^/]+)/compiler-proposals", path)
        if match:
            session = self.session(query, unquote(match.group(1)))
            projection, _ = projection_as_of(session, query.get("as_of_date", [""])[0] or latest_known_date(session), query.get("lens_id", [""])[0] or None)
            deal = projection["deal"]
            return self.json_response(200, {"discrepancies": deal.get("discrepancy_candidates", []), "derivations": deal.get("derivations", []), "hypotheses": deal.get("hypotheses", []), "spine_changes": deal.get("spine_change_proposals", [])})

        match = re.fullmatch(API + r"/cases/([^/]+)/objects/(.+)", path)
        if match:
            session = self.session(query, unquote(match.group(1)))
            projection, _ = projection_as_of(session, query.get("as_of_date", [""])[0] or latest_known_date(session))
            obj = find_object(projection, unquote(match.group(2)))
            return self.json_response(200, {"object": obj}) if obj else self.json_response(404, {"error": {"code": "OBJECT_NOT_FOUND", "message": "Object not found"}})

        match = re.fullmatch(API + r"/cases/([^/]+)/(claims|questions|cells)", path)
        if match:
            case_id, kind = unquote(match.group(1)), match.group(2)
            session = self.session(query, case_id)
            projection, _ = projection_as_of(session, query.get("as_of_date", [""])[0] or latest_known_date(session))
            deal = projection["deal"]
            rows = deal.get(kind if kind != "questions" else "question_spine", [])
            return self.json_response(200, {kind: rows, "total": len(rows)})

        match = re.fullmatch(API + r"/cases/([^/]+)/sources", path)
        if match:
            session = self.session(query, unquote(match.group(1)))
            projection, _ = projection_as_of(session, query.get("as_of_date", [""])[0] or latest_known_date(session))
            return self.json_response(200, {"sources": projection["deal"].get("source_center", {}).get("sources", [])})

        match = re.fullmatch(API + r"/cases/([^/]+)/inbox", path)
        if match:
            session = self.session(query, unquote(match.group(1)))
            return self.json_response(200, {"items": session["projection"]["deal"].get("source_center", {}).get("inbox", [])})

        match = re.fullmatch(API + r"/jobs/([^/]+)", path)
        if match:
            job_id = unquote(match.group(1))
            session = self.session(query)
            job = session["jobs"].get(job_id)
            if not job:
                return self.json_response(404, {"error": {"code": "JOB_NOT_FOUND", "message": "Ingest job not found"}})
            if job["status"] not in ("COMPLETE", "FAILED", "CANCELLED"):
                job["poll_count"] = job.get("poll_count", 0) + 1
                stages = [
                    ("QUEUED", 8, "Waiting for compiler"),
                    ("PARSING", 28, "Parsing native source structure"),
                    ("EXTRACTING", 52, "Extracting values, formulas and statements"),
                    ("BINDING", 76, "Binding evidence to questions and objects"),
                    ("VALIDATING", 91, "Running grounding and coverage checks"),
                ]
                if job["poll_count"] <= len(stages):
                    job["status"], job["progress"], job["stage_label"] = stages[job["poll_count"] - 1]
                else:
                    complete_job(session, job)
                save(session)
            return self.json_response(200, {"job": job})

        match = re.fullmatch(API + r"/cases/([^/]+)/replay", path)
        if match:
            session = self.session(query, unquote(match.group(1)))
            event_id = query.get("event_id", [""])[0]
            requested_date = query.get("as_of_date", [""])[0]
            event = None
            if event_id:
                event = next((item for item in session["registry"] if item.get("event_id") == event_id), None)
                if not event:
                    return self.json_response(404, {"error": {"code": "REPLAY_EVENT_NOT_FOUND", "message": "Replay event not found"}})
                requested_date = parse_iso(event["known_at"]).date().isoformat()
            if not requested_date:
                requested_date = latest_known_date(session)
            projection, registry = projection_as_of(session, requested_date)
            snapshots = projection["deal"].get("replay", {}).get("snapshots", [])
            if event:
                snapshot = next((item for item in snapshots if item.get("event_id") == event_id), None)
            else:
                snapshot = snapshots[-1] if snapshots else None
                event = next((item for item in registry if snapshot and item.get("event_id") == snapshot.get("event_id")), None)
            if not snapshot or not event:
                return self.json_response(404, {"error": {"code": "REPLAY_EMPTY", "message": "No event was known by the requested date"}})
            return self.json_response(
                200,
                {
                    "snapshot": snapshot,
                    "event": event,
                    "source_state_id": event.get("source_state_id") or snapshot.get("id"),
                    "result_state_id": event.get("result_state_id") or snapshot.get("id"),
                    "stable_hash": snapshot["stable_hash"],
                    "effective_date": event["effective_date"],
                    "known_at": event["known_at"],
                    "as_of_date": requested_date,
                    "read_only": True,
                    "derived_from_event_log": True,
                },
            )

        return self.json_response(404, {"error": {"code": "NOT_FOUND", "message": "Unknown V20 API route"}})

    def post_api(self, path, query, body):
        if path == f"{API}/sessions":
            mode = body.get("mode", "MOCK_CONNECTED")
            if mode == "CONNECTED":
                return self.json_response(503, {"error": {"code": "CONNECTED_BACKEND_NOT_CONFIGURED", "message": "The bundled server cannot create a production Connected session."}})
            session = new_session(body.get("case_id") or default_case(), body.get("actor", "partner"), mode)
            return self.json_response(201, {"session_id": session["session_id"]})

        if path == f"{API}/open-deal":
            session = self.session({"session_id": [body.get("session_id", "")]})
            projection = session["projection"]
            projection["deal"]["name"] = body.get("case_name") or projection["deal"].get("name")
            projection["deal"]["company"] = body.get("company") or projection["deal"].get("company")
            projection["deal"]["objective"] = {
                "verb": "decide",
                "target": body.get("case_name") or "the investment",
                "statement": body.get("objective") or "Determine whether to invest.",
                "deadline": "Open",
                "status": "active",
            }
            lines = [line.strip() for line in str(body.get("thesis", "")).splitlines() if line.strip()]
            known_at = utcnow()
            projection["deal"]["structured_opening"] = {
                "thesis_decomposition": lines,
                "recorded_at": known_at,
                "effective_date": body.get("effective_date") or parse_iso(known_at).date().isoformat(),
                "known_at": known_at,
                "actor_id": body.get("actor_id", "Professional"),
            }
            projection["deal"]["projection_id"] = uid("PROJ")
            append_registry(session, "DEAL_OPENING", "Structured deal opening recorded", body.get("objective", ""), body.get("actor_id", "Professional"))
            save(session)
            return self.json_response(201, {"status": "ACKNOWLEDGED", "message": "Deal decomposition recorded in the mock case store.", "projection": projection, "registry": session["registry"]})

        match = re.fullmatch(API + r"/cases/([^/]+)/ingest", path)
        if match:
            session = self.session(query, unquote(match.group(1)))
            job_id = uid("JOB")
            known_at = utcnow()
            job = {
                "job_id": job_id,
                "status": "QUEUED",
                "progress": 0,
                "stage": "QUEUED",
                "stage_label": "Queued",
                "message": "The source will enter the projection only after completion.",
                "source_name": body.get("file_name") or body.get("value") or "Source",
                "file_name": body.get("file_name"),
                "value": body.get("value"),
                "method": body.get("method"),
                "purpose": body.get("purpose"),
                "content_b64": body.get("content_b64"),
                "source_type": body.get("source_type"),
                "effective_date": body.get("effective_date") or parse_iso(known_at).date().isoformat(),
                "known_at": known_at,
                "created_at": known_at,
                "poll_count": 0,
            }
            session["jobs"][job_id] = job
            append_registry(session, "INGEST_REQUESTED", "Source ingest requested", job["source_name"], body.get("actor_id", "Professional"), job_id, effective_date=job["effective_date"], known_at=known_at)
            save(session)
            return self.json_response(202, {"job": job})

        match = re.fullmatch(API + r"/cases/([^/]+)/sources/([^/]+)/remove", path)
        if match:
            session = self.session(query, unquote(match.group(1)))
            source_id = unquote(match.group(2))
            source = next((item for item in session["projection"]["deal"].get("source_center", {}).get("sources", []) if item.get("source_id") == source_id), None)
            if not source:
                return self.json_response(404, {"error": {"code": "SOURCE_NOT_FOUND", "message": "Source not found"}})
            known_at = utcnow()
            source["status"] = "RETIRED"
            source["retired_at"] = known_at
            append_registry(session, "SOURCE_RETIRED", "Source retired from Current projection", "Historical version remains available.", "Source Registry", source_id, effective_date=parse_iso(known_at).date().isoformat(), known_at=known_at)
            save(session)
            return self.json_response(200, {"status": "RETIRED", "source": source})

        match = re.fullmatch(API + r"/cases/([^/]+)/notes", path)
        if match:
            session = self.session(query, unquote(match.group(1)))
            known_at = utcnow()
            body["server_ack_at"] = known_at
            body["effective_date"] = body.get("effective_date") or parse_iso(known_at).date().isoformat()
            body["known_at"] = known_at
            session["notes"].append(body)
            append_registry(session, body.get("kind", "ANNOTATION"), "Professional input acknowledged", body.get("decision") or body.get("text", ""), body.get("actor_id", "Professional"), body.get("object_id"), effective_date=body["effective_date"], known_at=known_at)
            save(session)
            return self.json_response(201, {"status": "ACKNOWLEDGED", "note": body, "registry": session["registry"]})

        match = re.fullmatch(API + r"/cases/([^/]+)/ic-record", path)
        if match:
            session = self.session(query, unquote(match.group(1)))
            known_at = utcnow()
            append_registry(session, "IC_RECORD", "IC decision ritual recorded", f"{body.get('decision')} · {body.get('conditions', '')}", body.get("actor_id", "IC"), session["case_id"], effective_date=body.get("effective_date") or parse_iso(known_at).date().isoformat(), known_at=known_at, epistemic_class="institutional_act")
            save(session)
            return self.json_response(201, {"status": "ACKNOWLEDGED", "message": "IC decision, conditions and dissent recorded.", "registry": session["registry"]})

        match = re.fullmatch(API + r"/cases/([^/]+)/events/([^/]+)/admit", path)
        if match:
            case_id, event_id = map(unquote, match.groups())
            session = self.session(query, case_id)
            key = self.headers.get("Idempotency-Key") or body.get("idempotency_key")
            pack = load_pack(case_id)
            event_name = event_key(pack, event_id)
            if not event_name:
                return self.json_response(404, {"error": {"code": "EVENT_NOT_FOUND", "message": "Event not found"}})
            missing = [field for field in ("treatment_id", "treatment_hash", "source_version_id", "event_id", "actor_id", "as_of_state_id") if not body.get(field)]
            if missing:
                return self.json_response(400, {"error": {"code": "ADMISSION_CONTEXT_MISSING", "message": "Admission payload incomplete", "details": {"missing": missing}}})
            if body.get("event_id") != event_id:
                return self.json_response(409, {"error": {"code": "EVENT_CONTEXT_MISMATCH", "message": "Payload event_id does not match route event_id"}})
            if body.get("as_of_state_id") != session["projection"]["deal"].get("as_of_state_id"):
                return self.json_response(409, {"error": {"code": "STALE_AS_OF_STATE", "message": "Admission was prepared against a stale state"}})

            def operation():
                transition = clone(pack["transitions"][event_name])
                transition["run_id"] = uid("RUN")
                transition["candidate_state_id"] = uid("CAND")
                transition["prior_state_id"] = session["projection"]["deal"]["as_of_state_id"]
                transition["as_of_state_id"] = transition["prior_state_id"]
                transition["as_of_date"] = session["projection"]["deal"].get("as_of_date")
                transition["projection_id"] = session["projection"]["deal"].get("projection_id")
                run = {
                    "run_id": transition["run_id"],
                    "transition": transition,
                    "candidate_state_id": transition["candidate_state_id"],
                    "status": "CANDIDATE",
                    "event_id": event_id,
                    "selected_change_ids": [],
                    "effective_date": transition.get("effective_date"),
                    "known_at": utcnow(),
                }
                session["runs"][run["run_id"]] = run
                append_registry(session, "ADMISSION", "Professional treatment admitted", body["treatment_id"], body["actor_id"], event_id, run["run_id"], effective_date=transition.get("effective_date"), known_at=run["known_at"], epistemic_class="institutional_act")
                append_registry(session, "TRANSITION", "Candidate calculated", f"{len(transition.get('affected_set', []))} mapped consequences", "Transition Engine", transition["candidate_state_id"], run["run_id"], effective_date=transition.get("effective_date"), known_at=run["known_at"], epistemic_class="derived")
                return {"run": run, "transition": transition, "context": context(session, run_id=run["run_id"], candidate_state_id=run["candidate_state_id"], human_stop_id=(transition.get("human_stops") or [{}])[0].get("stop_id")), "registry": session["registry"]}

            try:
                output = idempotent(session, f"admit:{case_id}:{event_id}", key, body, operation)
            except ValueError as exc:
                return self.json_response(409, {"error": {"code": "IDEMPOTENCY_CONFLICT", "message": str(exc)}})
            save(session)
            return self.json_response(200, output)

        match = re.fullmatch(API + r"/cases/([^/]+)/compiler-proposals/(discrepancy|derivation|hypothesis|spine)/([^/]+)/review", path)
        if match:
            case_id, kind, object_id = unquote(match.group(1)), match.group(2), unquote(match.group(3))
            session = self.session(query, case_id)
            decision = str(body.get("decision", "")).upper()
            if decision not in {"ADMITTED", "REJECTED", "CORRECTED", "ACCEPTED"}:
                return self.json_response(400, {"error": {"code": "INVALID_REVIEW_DECISION", "message": "Use ADMITTED, ACCEPTED, CORRECTED or REJECTED."}})
            projection, _ = projection_as_of(session, latest_known_date(session))
            collection = {"discrepancy": "discrepancy_candidates", "derivation": "derivations", "hypothesis": "hypotheses", "spine": "spine_change_proposals"}[kind]
            obj = next((item for item in projection["deal"].get(collection, []) if object_id in (item.get("id"), item.get("discrepancy_id"), item.get("derivation_id"), item.get("hypothesis_id"), item.get("proposal_id"))), None)
            if not obj:
                return self.json_response(404, {"error": {"code": "COMPILER_PROPOSAL_NOT_FOUND", "message": "Compiler proposal not found"}})
            actor = actor_for(session["projection"], session.get("actor_key", "partner"))
            if kind == "spine" and decision in {"ADMITTED", "ACCEPTED"} and "approve_spine_change" not in actor.get("authority_verbs", []):
                return self.json_response(403, {"error": {"code": "SPINE_AUTHORITY_REQUIRED", "message": "The authenticated actor lacks approve_spine_change."}})
            known_at = utcnow()
            review = {"review_id": uid("REVIEW"), "kind": kind, "object_id": object_id, "decision": decision, "rationale": body.get("rationale"), "actor_id": actor["actor_id"], "effective_date": body.get("effective_date") or today_utc(), "known_at": known_at}
            session.setdefault("proposal_reviews", {})[f"{kind}:{object_id}"] = review
            if kind == "spine" and decision in {"ADMITTED", "ACCEPTED"}:
                base = next((item for item in session["projection"]["deal"].get("spine_change_proposals", []) if item.get("proposal_id") == object_id), None)
                if base and base.get("proposed_question"):
                    proposed = clone(base["proposed_question"])
                    if not any(q.get("id") == proposed.get("id") for q in session["projection"]["deal"].get("question_spine", [])):
                        proposed.update({"owner": actor["actor_id"], "versions": {today_utc(): {"view": "Newly admitted spine question; institutional view not yet formed.", "strength": "UNKNOWN", "economic_weight": "Not yet quantified", "cause": "Governed spine change", "effective_date": today_utc(), "known_at": known_at}}, "claim_ids": base.get("binding_migration", {}).get("claim_ids", []), "model_node_ids": [], "work_plan": [], "decision_axes": {}})
                        session["projection"]["deal"].setdefault("question_spine", []).append(proposed)
            append_registry(session, "COMPILER_REVIEW", f"{kind.title()} proposal {decision.lower()}", obj.get("label") or object_id, actor["actor_id"], object_id, effective_date=review["effective_date"], known_at=known_at)
            save(session)
            projection, registry = projection_as_of(session, latest_known_date(session))
            return self.json_response(200, {"review": review, "projection": projection, "context": context(session, projection), "registry": registry})

        match = re.fullmatch(API + r"/cases/([^/]+)/missions/([^/]+)/(prepare|run)", path)
        if match:
            case_id, mission_id, action = unquote(match.group(1)), unquote(match.group(2)), match.group(3)
            session = self.session(query, case_id)
            mission = next((item for item in session["projection"]["deal"].get("agent_missions", []) if item.get("mission_id") == mission_id), None)
            if not mission:
                return self.json_response(404, {"error": {"code": "MISSION_NOT_FOUND", "message": "Mission not found"}})
            known_at = utcnow(); actor = actor_for(session["projection"], session.get("actor_key", "partner"))
            if action == "prepare":
                draft = {"mission_run_id": uid("MISSION-DRAFT"), "mission_id": mission_id, "status": "PREPARED", "objective": mission.get("objective"), "allowed_sources": mission.get("allowed_sources", []), "prohibited_sources": mission.get("prohibited_sources", []), "confidential_context_policy": mission.get("confidential_context_policy"), "data_egress_policy": mission.get("data_egress_policy"), "synthetic": True, "no_external_effects": True, "effective_date": today_utc(), "known_at": known_at}
                session.setdefault("mission_runs", {})[draft["mission_run_id"]] = draft
                append_registry(session, "MISSION_PREPARED", "Governed mission prepared", mission.get("label"), actor["actor_id"], mission_id, effective_date=today_utc(), known_at=known_at)
                save(session)
                return self.json_response(201, {"mission_run": draft, "message": "Mission draft prepared. No research or human contact occurred."})
            if mission.get("external_human_contact") or not mission.get("auto_executable_in_mock"):
                return self.json_response(409, {"error": {"code": "MISSION_AUTHORITY_REQUIRED", "message": "This mission involves human contact, a physical test or another external effect. V20 will prepare it but will not auto-run it."}})
            run = {"mission_run_id": uid("MISSION-RUN"), "mission_id": mission_id, "status": "COMPLETE", "execution_mode": "MOCK_SYNTHETIC_RESEARCH", "queries_redacted": True, "source_quality": "SYNTHETIC_REFERENCE_ONLY", "synthetic": True, "no_external_effects": True, "effective_date": today_utc(), "known_at": known_at}
            source_id = uid("SRC-MISSION")
            source = {"source_id": source_id, "id": source_id, "name": f"Synthetic mission result · {mission.get('label')}", "type": "RESEARCH_RESULT", "status": "INGESTED", "latest_version_id": source_id + "-V1", "claim_count": 1, "effective_date": today_utc(), "known_at": known_at, "ingested_at": known_at, "versions": [{"version_id": source_id + "-V1", "created_at": known_at, "effective_date": today_utc(), "known_at": known_at, "sha256": sha(run), "supersedes": None}], "viewer": "METADATA", "mission_id": mission_id, "retrieval_trace": {"allowed_sources": mission.get("allowed_sources", []), "prohibited_sources": mission.get("prohibited_sources", []), "data_egress_policy": mission.get("data_egress_policy"), "mock": True}, "synthetic": True}
            proposed_claim = {"claim_id": uid("CL-MISSION"), "id": uid("CL-MISSION-REF"), "statement": f"Synthetic research result for professional review: {mission.get('expected_output')}", "source_id": source_id, "locator": "Mission result · object 1", "epistemic_class": "asserted", "value": None, "unit": None, "definition_id": None, "period": today_utc(), "perimeter": "Synthetic mock research result", "topic": "Governed research", "direction": "context", "bears_on": mission.get("question_ids", []), "position_ids": [], "review_status": "REVIEW_REQUIRED", "ground_truth_flag": False, "validation_only": False, "mission_id": mission_id, "effective_date": today_utc(), "known_at": known_at}
            proposed_claim["id"] = proposed_claim["claim_id"]
            session["projection"]["deal"].setdefault("source_center", {}).setdefault("sources", []).append(source)
            session["projection"]["deal"].setdefault("claims", []).append(proposed_claim)
            run.update({"source_id": source_id, "proposed_claim_ids": [proposed_claim["claim_id"]]})
            session.setdefault("mission_runs", {})[run["mission_run_id"]] = run
            append_registry(session, "MISSION_COMPLETED", "Synthetic read-only mission completed", mission.get("label"), actor["actor_id"], mission_id, effective_date=today_utc(), known_at=known_at)
            save(session)
            projection, registry = projection_as_of(session, latest_known_date(session))
            return self.json_response(200, {"mission_run": run, "source": source, "proposed_claims": [proposed_claim], "projection": projection, "context": context(session, projection), "registry": registry})

        match = re.fullmatch(API + r"/runs/([^/]+)/prepare", path)
        if match:
            run_id = unquote(match.group(1))
            session = self.session(query)
            run = session["runs"].get(run_id)
            if not run:
                return self.json_response(404, {"error": {"code": "RUN_NOT_FOUND", "message": "Run not found"}})
            if run.get("status") not in {"CANDIDATE", "PREPARED"}:
                return self.json_response(409, {"error": {"code": "RUN_NOT_PREPARABLE", "message": "Only a Candidate run may be prepared"}})
            raw_selected = body.get("selected_change_ids", [])
            if not isinstance(raw_selected, list) or any(
                not isinstance(item, str) or not item.strip()
                for item in raw_selected
            ):
                return self.json_response(400, {"error": {"code": "INVALID_CHANGE_IDS", "message": "selected_change_ids must contain non-empty strings"}})
            selected = [item.strip() for item in raw_selected]
            if not selected:
                return self.json_response(409, {"error": {"code": "NO_CHANGES_SELECTED", "message": "Select at least one change explicitly."}})
            if len(selected) != len(set(selected)):
                return self.json_response(409, {"error": {"code": "DUPLICATE_CHANGE_ID", "message": "Duplicate selected change IDs are not allowed"}})
            allowed = valid_change_ids(run["transition"])
            unknown = sorted(set(selected) - allowed)
            if unknown:
                return self.json_response(409, {"error": {"code": "UNKNOWN_CHANGE_ID", "message": "Prepared change is outside the transition output", "details": {"unknown_change_ids": unknown}}})
            selection_hash = sha(
                {
                    "run_id": run_id,
                    "candidate_state_id": run.get("candidate_state_id"),
                    "selected_change_ids": sorted(selected),
                }
            )
            if run.get("status") == "PREPARED":
                if run.get("prepared_selection_hash") != selection_hash:
                    return self.json_response(409, {"error": {"code": "PREPARED_SELECTION_CONFLICT", "message": "Run is already PREPARED with another selection"}})
                return self.json_response(200, {"run": run})
            run["selected_change_ids"] = selected
            run["prepared_selection_hash"] = selection_hash
            run["status"] = "PREPARED"
            run["prepared_at"] = utcnow()
            append_registry(session, "PREPARATION", "Explicit change set prepared", ", ".join(selected), actor_for(session["projection"], session.get("actor_key", "partner"))["actor_id"], run["candidate_state_id"], run_id, effective_date=run["transition"].get("effective_date"), known_at=run["prepared_at"])
            save(session)
            return self.json_response(200, {"run": run})

        match = re.fullmatch(API + r"/runs/([^/]+)/authority/attest", path)
        if match:
            run_id = unquote(match.group(1))
            session = self.session(query)
            key = self.headers.get("Idempotency-Key") or body.get("idempotency_key")
            run = session["runs"].get(run_id)
            if not run:
                return self.json_response(404, {"error": {"code": "RUN_NOT_FOUND", "message": "Run not found"}})
            if run.get("status") != "PREPARED":
                return self.json_response(409, {"error": {"code": "RUN_NOT_PREPARED", "message": "Authority requires an explicitly prepared change set"}})
            if not run.get("prepared_selection_hash"):
                return self.json_response(409, {"error": {"code": "PREPARED_SELECTION_MISSING", "message": "Authority requires a durably prepared selection hash"}})
            if body.get("run_id") and body.get("run_id") != run_id:
                return self.json_response(409, {"error": {"code": "RUN_CONTEXT_MISMATCH", "message": "Authority request run_id does not match the route"}})
            if body.get("candidate_state_id") != run.get("candidate_state_id"):
                return self.json_response(409, {"error": {"code": "CANDIDATE_CONTEXT_MISMATCH", "message": "Authority request Candidate does not match the run"}})
            transition = run["transition"]
            stop = next((item for item in transition.get("human_stops", []) if item.get("stop_id") == body.get("human_stop_id")), None)
            if not stop or stop.get("status") not in {"OPEN", None}:
                return self.json_response(409, {"error": {"code": "HUMAN_STOP_NOT_OPEN", "message": "Human Stop is not open for this run"}})
            normalized_stop = normalized_human_stop(stop)
            if normalized_stop["attestable"] is not True:
                return self.json_response(409, {"error": {"code": "HUMAN_STOP_NOT_ATTESTABLE", "message": "Human Stop requires correction or replay and cannot be authority-attested"}})
            actor = actor_for(session["projection"], session.get("actor_key", "partner"))
            if not isinstance(body.get("actor_id"), str) or not body.get("actor_id", "").strip():
                return self.json_response(400, {"error": {"code": "ACTOR_ID_REQUIRED", "message": "actor_id is required for authority attestation"}})
            if body["actor_id"].strip() != actor.get("actor_id"):
                return self.json_response(403, {"error": {"code": "ACTOR_CONTEXT_MISMATCH", "message": "actor_id does not match the authenticated session actor"}})
            verb = stop.get("authority_verb") or session["projection"]["deal"]["decisionRoom"].get("verb")
            if verb not in actor.get("authority_verbs", []):
                return self.json_response(403, {"error": {"code": "INSUFFICIENT_AUTHORITY", "message": f"{actor['actor_id']} lacks {verb}"}})
            course = next((item for item in session["projection"]["deal"]["decisionRoom"].get("courses", []) if item.get("id") == body.get("course_id")), None)
            if not course:
                return self.json_response(400, {"error": {"code": "COURSE_NOT_ADMISSIBLE", "message": "Course not found"}})
            existing = [record for record in session["authority_records"].values() if record.get("run_id") == run_id and record.get("human_stop_id") == stop["stop_id"]]
            if existing and any(record.get("course_id") != course["id"] for record in existing):
                return self.json_response(409, {"error": {"code": "CONFLICTING_ATTESTATION", "message": "An incompatible course is already attested for this Human Stop"}})
            if not body.get("artifact_hash"):
                return self.json_response(400, {"error": {"code": "ARTIFACT_HASH_REQUIRED", "message": "Authority record requires the prepared artifact hash"}})
            if body.get("artifact_hash") != transition.get("replay_hash"):
                return self.json_response(409, {"error": {"code": "ARTIFACT_HASH_MISMATCH", "message": "Authority artifact hash does not match this Candidate replay"}})

            def operation():
                known_at = utcnow()
                assignment = mock_authority_assignment_snapshot(
                    actor, normalized_stop["required_role"], known_at
                )
                authentication_context = {
                    "principal_id": actor["actor_id"],
                    "case_id": session["case_id"],
                    "authentication_method": "SYNTHETIC_SERVER_SESSION",
                    "session_id_hash": "sha256:" + hashlib.sha256(
                        session["session_id"].encode("utf-8")
                    ).hexdigest(),
                    "session_issued_at": session["created_at"],
                    "session_expires_at": session["expires_at"],
                    "authenticated_at": known_at,
                }
                existing_records = sorted(
                    session["authority_records"].values(),
                    key=lambda item: item.get("ledger_sequence", 0),
                )
                record = {
                    "authority_record_version": AUTHORITY_RECORD_VERSION,
                    "authority_record_id": uid("AUTH"),
                    "case_id": session["case_id"],
                    "run_id": run_id,
                    "candidate_state_id": run["candidate_state_id"],
                    "human_stop_id": stop["stop_id"],
                    "course_id": course["id"],
                    "actor_id": actor["actor_id"],
                    "actor_role": actor["role"],
                    "timestamp": known_at,
                    "effective_date": transition.get("effective_date") or parse_iso(known_at).date().isoformat(),
                    "known_at": known_at,
                    "artifact_hash": body["artifact_hash"],
                    "authority_verb": verb,
                    "required_role": normalized_stop["required_role"],
                    "prepared_selection_hash": run["prepared_selection_hash"],
                    "authority_assignment": assignment,
                    "authority_assignment_hash": sha(assignment),
                    "authentication_context": authentication_context,
                    "authentication_context_hash": sha(authentication_context),
                    "ledger_id": AUTHORITY_LEDGER_ID,
                    "ledger_sequence": len(existing_records) + 1,
                    "previous_record_hash": (
                        existing_records[-1]["record_hash"]
                        if existing_records
                        else None
                    ),
                    "effect_type": course.get("effect_type", "INTERNAL"),
                    "status": "ATTESTED",
                    "synthetic": True,
                    **authority_record_binding(transition, stop),
                }
                record = mock_sign_authority_record(record)
                session["authority_records"][record["authority_record_id"]] = record
                append_registry(session, "AUTHORITY", "Authority decision attested", course["label"], actor["actor_id"], record["authority_record_id"], run_id, effective_date=record["effective_date"], known_at=known_at, epistemic_class="institutional_act")
                package = build_package(session, course, record) if course.get("effect_type") == "EXTERNAL_PACKAGE" else None
                return {"authority_record": record, "execution_package": package, "registry": session["registry"]}

            try:
                output = idempotent(session, f"authority:{run_id}:{stop['stop_id']}", key, body, operation)
            except ValueError as exc:
                return self.json_response(409, {"error": {"code": "IDEMPOTENCY_CONFLICT", "message": str(exc)}})
            save(session)
            return self.json_response(200, output)

        match = re.fullmatch(API + r"/runs/([^/]+)/execution-packages", path)
        if match:
            return self.json_response(409, {"error": {"code": "PACKAGE_SERVER_DERIVED", "message": "Packages are created only from an attested course."}})

        match = re.fullmatch(API + r"/execution-packages/([^/]+)/send", path)
        if match:
            package_id = unquote(match.group(1))
            session = self.session(query)
            package = session["packages"].get(package_id)
            if not package:
                return self.json_response(404, {"error": {"code": "PACKAGE_NOT_FOUND", "message": "Package not found"}})
            known_at = utcnow()
            if body.get("simulate_failure"):
                package["status"] = "FAILED"
                package["failed_at"] = known_at
                save(session)
                return self.json_response(503, {"error": {"code": "DELIVERY_FAILED", "message": "Simulated delivery failed; nothing was sent."}})
            package["status"] = "ACCEPTED"
            package["ack_id"] = uid("ACK")
            package["acknowledged_at"] = known_at
            append_registry(session, "EXECUTION", "Simulated package accepted", "No external system was contacted.", "Execution Service", package_id, package["run_id"], effective_date=parse_iso(known_at).date().isoformat(), known_at=known_at, epistemic_class="institutional_act")
            save(session)
            return self.json_response(200, {"execution_package": package, "registry": session["registry"]})

        match = re.fullmatch(API + r"/runs/([^/]+)/settle", path)
        if match:
            session = self.session(query)
            key = self.headers.get("Idempotency-Key") or body.get("idempotency_key")
            body = {**body, "run_id": unquote(match.group(1))}
            try:
                output = idempotent(session, f"settle:{body['run_id']}", key, body, lambda: settle(session, body))
            except ValueError as exc:
                return self.json_response(409, {"error": {"code": "SETTLEMENT_INVARIANT_FAILED", "message": str(exc)}})
            save(session)
            return self.json_response(200, output)

        match = re.fullmatch(API + r"/cases/([^/]+)/work-items/([^/]+)/prepare", path)
        if match:
            session = self.session(query, unquote(match.group(1)))
            known_at = utcnow()
            draft = {
                "draft_id": uid("DRAFT"),
                "work_item_id": unquote(match.group(2)),
                "status": "DRAFT",
                "created_at": known_at,
                "effective_date": body.get("effective_date") or parse_iso(known_at).date().isoformat(),
                "known_at": known_at,
                **body,
                "synthetic": True,
            }
            return self.json_response(201, {"draft": draft, "message": "Draft prepared in PANTA. It was not dispatched externally."})

        return self.json_response(404, {"error": {"code": "NOT_FOUND", "message": "Unknown V20 API route"}})


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=4191)
    args = parser.parse_args()
    os.chdir(APP)
    server = ThreadingHTTPServer((args.host, args.port), Handler)
    print(
        f"PANTA V20 mock server: http://{args.host}:{args.port}/?mode=mock&case={default_case()}&actor=partner",
        flush=True,
    )
    server.serve_forever()


if __name__ == "__main__":
    main()
