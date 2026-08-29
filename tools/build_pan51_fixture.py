#!/usr/bin/env python3
"""Build the non-sensitive PAN-51 formula coverage workbook fixture."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName


ROOT = Path(__file__).resolve().parent.parent
OUTPUT = ROOT / "tools" / "fixtures" / "pan51_formula_model.xlsx"


def build(path: Path = OUTPUT) -> Path:
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs.append(["Input", "Value"])
    inputs.append(["Revenue base", 100])
    inputs.append(["Growth", 0.2])
    inputs.append(["Costs", 50])
    inputs.append(["Lookup key", 2])

    lookup = workbook.create_sheet("Lookup")
    lookup.append(["Key", "Multiple"])
    lookup.append([1, 8.0])
    lookup.append([2, 9.5])
    lookup.append([3, 11.0])

    outputs = workbook.create_sheet("Outputs")
    outputs.append(["Metric", "Formula output"])
    outputs.append(["Revenue", "=Inputs!B2*(1+Inputs!B3)"])
    outputs.append(["Input aggregate", "=SUM(Inputs!B2:B4)"])
    outputs.append(["Conditional costs", "=IF(Inputs!B2>90,Inputs!B4,0)"])
    outputs.append(["Selected multiple", "=VLOOKUP(Inputs!B5,Lookup!A2:B4,2,FALSE)"])
    outputs.append(["Named revenue", "=RevenueBase*(1+Inputs!B3)"])
    outputs.append(["External dependency", "='[External.xlsx]Data'!A1+Inputs!B2"])
    outputs.append(["Unsupported function", '=CUBEVALUE("Connection","[Measures].[Revenue]")'])
    outputs.append(["Key underwriting output", "=Outputs!B2-Inputs!B4"])

    workbook.defined_names.add(
        DefinedName("RevenueBase", attr_text="'Inputs'!$B$2")
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


if __name__ == "__main__":
    built = build()
    print(f"[pan51-fixture] -> {built}")
