#!/usr/bin/env python3
"""
sheet_classifier — the one judgment a model makes, made as stable as possible.

Everything else in L2 is deterministic and generalises: label columns, header
rows, unit columns, block boundaries. Two things do not, and both are questions
about meaning rather than layout:

  * what kind of sheet this is — a model, or a table of records
  * whether a column header names a period — FY2025A, P1, T1, "Anno 1"

Period vocabulary is unbounded, so a parser recognises only what it has met. A
French model heading its columns P1/P2 dropped to zero confidence for exactly
that reason. This asks a model those two questions and nothing else.

How stability is obtained
-------------------------
Not by hoping the model is consistent. By making a second run unable to differ:

  1. Fingerprint. A sheet is reduced to its structure — dimensions, header
     candidates, label samples, column types — with values rounded away. The
     fingerprint is the sha256 of that canonical form.
  2. Cache. A fingerprint that has been answered is never asked again. Identical
     structure therefore yields an identical answer by construction, across
     runs, machines and model versions.
  3. Bounded output. tool_use with an enum; the model selects, it does not
     compose. Note that temperature is not available on this SDK for this model
     family — OutputConfigParam exposes effort and format only — so sampling
     cannot be pinned. That is precisely why the cache, not the decoding
     parameters, carries the stability guarantee.
  4. Structural veto. An answer that contradicts something the deterministic
     layer knows for certain is rejected, not accepted. A model_sheet needs a
     header row; a record_table needs more rows than a handful. The model can
     resolve ambiguity, it cannot overrule evidence.
  5. Durable record. Judgments live in the vault with their fingerprint and the
     reason given, so they are auditable and diffable rather than re-derived.

The cache is the load-bearing part. Without it, "stable" means "the model
behaved last time"; with it, the second run is not a model call at all.

    python3 tools/sheet_classifier.py --workbook model.xlsx [--apply]
    python3 tools/sheet_classifier.py --selftest
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import sys
from dataclasses import dataclass, field, asdict
from datetime import datetime, date, timezone
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parent.parent
CACHE = ROOT / "vault" / "policy" / "sheet_classifications.json"

SHEET_KINDS = ["model_sheet", "record_table", "scalar_block", "unknown"]


# ── fingerprint ──────────────────────────────────────────────────────────────

def _cell_shape(v: Any) -> str:
    """The shape of a value, never the value: two models with different numbers
    but the same layout must fingerprint identically."""
    if v is None:
        return "-"
    if isinstance(v, bool):
        return "b"
    if isinstance(v, (datetime, date)):
        return "d"
    if isinstance(v, (int, float)):
        return "n"
    s = str(v)
    if s.startswith("="):
        return "f"
    return "t"


def fingerprint_sheet(ws, probe_rows: int = 25, probe_cols: int = 12) -> tuple[str, dict]:
    """(digest, the canonical structure it was computed from)."""
    max_r = min(ws.max_row, probe_rows)
    max_c = min(ws.max_column, probe_cols)

    grid = ["".join(_cell_shape(ws.cell(row=r, column=c).value)
                    for c in range(1, max_c + 1))
            for r in range(1, max_r + 1)]

    # Text is what the judgment turns on, so a bounded, ordered sample of it is
    # part of the fingerprint — but truncated, so an edited label does not
    # invalidate the whole sheet.
    texts: list[str] = []
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip() and not v.startswith("="):
                texts.append(f"{r},{c}:{v.strip()[:40]}")
            elif isinstance(v, (datetime, date)):
                texts.append(f"{r},{c}:<date>")
    texts = texts[:60]

    canonical = {
        "name": ws.title,
        "dims": [ws.max_row, ws.max_column],
        "grid": grid,
        "texts": texts,
    }
    blob = json.dumps(canonical, sort_keys=True, separators=(",", ":"),
                      ensure_ascii=False)
    return hashlib.sha256(blob.encode()).hexdigest()[:16], canonical


# ── judgments ────────────────────────────────────────────────────────────────

@dataclass
class Judgment:
    sheet: str
    fingerprint: str
    kind: str
    period_headers: dict[str, str] = field(default_factory=dict)  # header text -> kind
    reason: str = ""
    source: str = "model"          # model | cache | veto | deterministic
    vetoed: str = ""               # what the veto overruled, if anything
    decided_at: str = ""


def load_cache(path: Path = CACHE) -> dict[str, dict]:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8")).get("judgments", {})
    except Exception:
        return {}


def save_cache(judgments: dict[str, dict], path: Path = CACHE) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps({
        "schema": "sheet-classifications-1",
        "note": "Judgments keyed by structural fingerprint. A fingerprint that "
                "appears here is never sent to a model again.",
        "judgments": judgments,
    }, indent=2, ensure_ascii=False, sort_keys=True), encoding="utf-8")


# ── structural veto ──────────────────────────────────────────────────────────

def veto(kind: str, canonical: dict) -> tuple[str, str]:
    """
    Reject an answer that contradicts what is certain from structure.

    Returns (kind_to_use, reason_if_overruled). The model resolves ambiguity;
    it does not get to overrule evidence.
    """
    rows, cols = canonical["dims"]
    grid = canonical["grid"]

    if kind == "record_table" and rows < 5:
        return "unknown", f"record_table rifiutato: solo {rows} righe"
    if kind == "model_sheet":
        # a model needs at least one row that is mostly text or dates followed
        # by a row that is mostly numbers or formulas
        has_header = any(r.count("t") + r.count("d") >= 2 for r in grid)
        has_values = any(r.count("n") + r.count("f") >= 2 for r in grid)
        if not (has_header and has_values):
            return "unknown", "model_sheet rifiutato: nessuna riga header sopra valori"
    if kind not in SHEET_KINDS:
        return "unknown", f"tipo non ammesso: {kind!r}"
    return kind, ""


# ── the model call ───────────────────────────────────────────────────────────

CLASSIFY_TOOL = {
    "name": "classify_sheet",
    "description": "Classify one spreadsheet sheet and identify its period headers.",
    "input_schema": {
        "type": "object",
        "required": ["kind", "reason", "period_headers"],
        "additionalProperties": False,
        "properties": {
            "kind": {
                "type": "string",
                "enum": SHEET_KINDS,
                "description": (
                    "model_sheet: line items down the left, periods across the top. "
                    "record_table: one row per record, columns are fields. "
                    "scalar_block: labels and one or two value columns, no period axis. "
                    "unknown: genuinely cannot tell."
                ),
            },
            "period_headers": {
                "type": "object",
                "additionalProperties": {"type": "string"},
                "description": (
                    "Header texts that denote a time period, mapped to one of "
                    "quarter, fiscal_year, month, point, anchor, sequence. "
                    "Include P1/T1/'Anno 1' style sequence labels. Empty if none."
                ),
            },
            "reason": {"type": "string", "description": "One sentence, in Italian."},
        },
    },
}

SYSTEM = (
    "Classifichi la struttura di un foglio di calcolo finanziario. "
    "Ricevi solo la forma del foglio: tipi di cella e un campione di testo, mai i numeri. "
    "Rispondi esclusivamente attraverso lo strumento. "
    "Se il foglio è ambiguo scegli 'unknown': una risposta sbagliata con sicurezza "
    "costa più di un'astensione."
)


def build_prompt(canonical: dict) -> str:
    rows, cols = canonical["dims"]
    lines = [
        f"FOGLIO: {canonical['name']}",
        f"DIMENSIONI: {rows} righe x {cols} colonne",
        "",
        "FORMA (t=testo d=data n=numero f=formula b=bool -=vuota), una riga per riga:",
    ]
    for i, g in enumerate(canonical["grid"], start=1):
        lines.append(f"  r{i:<3} {g}")
    lines += ["", "TESTO (riga,colonna:contenuto):"]
    lines += [f"  {t}" for t in canonical["texts"]]
    return "\n".join(lines)


def classify_with_model(canonical: dict, api_key: str) -> tuple[str, dict, str]:
    import anthropic
    client = anthropic.Anthropic(api_key=api_key)
    resp = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=700,
        system=SYSTEM,
        tools=[CLASSIFY_TOOL],
        tool_choice={"type": "tool", "name": "classify_sheet"},
        messages=[{"role": "user", "content": build_prompt(canonical)}],
    )
    for block in resp.content:
        if block.type == "tool_use" and block.name == "classify_sheet":
            d = block.input
            return (str(d.get("kind", "unknown")),
                    dict(d.get("period_headers") or {}),
                    str(d.get("reason", "")))
    return "unknown", {}, "nessuna risposta dallo strumento"


# ── orchestration ────────────────────────────────────────────────────────────

def classify_workbook(path: Path, api_key: str | None,
                      cache_path: Path = CACHE) -> list[Judgment]:
    import openpyxl
    wb = openpyxl.load_workbook(str(path), data_only=False)
    cache = load_cache(cache_path)
    out: list[Judgment] = []

    for ws in wb:
        fp, canonical = fingerprint_sheet(ws)
        hit = cache.get(fp)
        if hit:
            out.append(Judgment(sheet=ws.title.upper(), fingerprint=fp,
                                kind=hit["kind"],
                                period_headers=hit.get("period_headers", {}),
                                reason=hit.get("reason", ""), source="cache",
                                decided_at=hit.get("decided_at", "")))
            continue

        if not api_key:
            out.append(Judgment(ws.title.upper(), fp, "unknown", {},
                                "nessuna API key: nessun giudizio espresso",
                                source="deterministic"))
            continue

        kind, periods, reason = classify_with_model(canonical, api_key)
        final, overruled = veto(kind, canonical)
        j = Judgment(sheet=ws.title.upper(), fingerprint=fp, kind=final,
                     period_headers=periods, reason=reason,
                     source="veto" if overruled else "model",
                     vetoed=overruled,
                     decided_at=datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"))
        out.append(j)
        cache[fp] = {"sheet": j.sheet, "kind": j.kind,
                     "period_headers": j.period_headers, "reason": j.reason,
                     "vetoed": j.vetoed, "decided_at": j.decided_at}

    save_cache(cache, cache_path)
    return out


# ── self-test ────────────────────────────────────────────────────────────────

def _selftest() -> int:
    import openpyxl
    from openpyxl import Workbook
    ok = True

    def check(name, cond, detail=""):
        nonlocal ok
        ok &= bool(cond)
        print(f"  {'✓' if cond else '✗'} {name}" + (f"  [{detail}]" if detail else ""))

    print("=" * 62)
    print("sheet_classifier — stabilità")
    print("=" * 62)

    wb = Workbook(); ws = wb.active; ws.title = "M"
    ws["A1"], ws["B1"], ws["C1"] = "Line", "FY2025A", "FY2026E"
    ws["A2"], ws["B2"], ws["C2"] = "Revenue", 74.0, "=B2*1.06"
    ws["A3"], ws["B3"], ws["C3"] = "EBITDA", 11.4, "=C2*0.15"

    fp1, canon1 = fingerprint_sheet(ws)
    fp2, _ = fingerprint_sheet(ws)
    check("l'impronta è stabile fra due letture", fp1 == fp2, fp1)

    # same layout, different numbers -> same fingerprint, so the cache holds
    wb2 = Workbook(); w2 = wb2.active; w2.title = "M"
    w2["A1"], w2["B1"], w2["C1"] = "Line", "FY2025A", "FY2026E"
    w2["A2"], w2["B2"], w2["C2"] = "Revenue", 999.0, "=B2*1.06"
    w2["A3"], w2["B3"], w2["C3"] = "EBITDA", 222.0, "=C2*0.15"
    fp3, _ = fingerprint_sheet(w2)
    check("numeri diversi, stessa struttura => stessa impronta", fp1 == fp3)

    # changed label -> different fingerprint, so the judgment is re-asked
    w2["A2"] = "Turnover"
    fp4, _ = fingerprint_sheet(w2)
    check("etichetta cambiata => impronta diversa", fp1 != fp4)

    # veto
    k, why = veto("model_sheet", {"dims": [3, 3], "grid": ["ttt", "tnf", "tnf"]})
    check("model_sheet con header e valori passa", k == "model_sheet")
    k, why = veto("model_sheet", {"dims": [2, 2], "grid": ["nn", "nn"]})
    check("model_sheet senza header viene respinto", k == "unknown", why)
    k, why = veto("record_table", {"dims": [3, 4], "grid": ["tttt"]})
    check("record_table con 3 righe viene respinto", k == "unknown", why)
    k, why = veto("inventato", {"dims": [9, 4], "grid": ["tttt"]})
    check("un tipo fuori enum viene respinto", k == "unknown", why)

    # cache round-trip
    import tempfile
    with tempfile.TemporaryDirectory() as td:
        cp = Path(td) / "c.json"
        save_cache({fp1: {"sheet": "M", "kind": "model_sheet",
                          "period_headers": {"FY2025A": "fiscal_year"},
                          "reason": "x", "decided_at": "t"}}, cp)
        js = classify_workbook_from(wb, None, cp)
        check("un'impronta in cache non chiede al modello",
              js and js[0].source == "cache" and js[0].kind == "model_sheet",
              js[0].source if js else "-")
        check("senza API key e senza cache non si inventa un giudizio",
              classify_workbook_from(wb2, None, Path(td) / "vuota.json")[0].kind == "unknown")

    print("\n" + "=" * 62)
    print("PASS" if ok else "FAIL")
    print("=" * 62)
    return 0 if ok else 1


def classify_workbook_from(wb, api_key, cache_path):
    """classify_workbook against an already-open workbook (used by the self-test)."""
    cache = load_cache(cache_path)
    out = []
    for ws in wb:
        fp, canonical = fingerprint_sheet(ws)
        hit = cache.get(fp)
        if hit:
            out.append(Judgment(ws.title.upper(), fp, hit["kind"],
                                hit.get("period_headers", {}), hit.get("reason", ""),
                                source="cache"))
        elif not api_key:
            out.append(Judgment(ws.title.upper(), fp, "unknown", {},
                                "nessuna API key", source="deterministic"))
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description="Classify sheets, stably")
    ap.add_argument("--workbook", type=Path)
    ap.add_argument("--cache", type=Path, default=CACHE)
    ap.add_argument("--selftest", action="store_true")
    a = ap.parse_args()

    if a.selftest:
        return _selftest()
    if not a.workbook:
        ap.error("serve --workbook, oppure --selftest")

    key = os.environ.get("ANTHROPIC_API_KEY", "")
    js = classify_workbook(a.workbook, key or None, a.cache)

    print(f"[sheet_classifier] {a.workbook.name}   cache: {a.cache}")
    counts: dict[str, int] = {}
    for j in js:
        counts[j.source] = counts.get(j.source, 0) + 1
        flag = f"  ⟵ veto: {j.vetoed}" if j.vetoed else ""
        periods = f"  periodi={len(j.period_headers)}" if j.period_headers else ""
        print(f"  {j.sheet:22} {j.kind:13} [{j.source}] {j.fingerprint}{periods}{flag}")
    print(f"\n  origine dei giudizi: {counts}")
    if not key:
        print("  (ANTHROPIC_API_KEY non impostata: nessuna chiamata effettuata)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
