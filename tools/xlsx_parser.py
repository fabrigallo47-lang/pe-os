#!/usr/bin/env python3
"""Direct xlsx parser for LBO model structure — zero LLM.

Discovers structural model nodes from an Excel workbook using sheet conventions:
  - Inputs sheet: capital-structure scalar inputs + assumption series
  - QoE_Bridge: EBITDA-definition totals + firm-view adjustments
  - Ownership_Returns: MOIC / XIRR outputs per scenario case
  - S&U_Opening: sources-and-uses check cells
  - Scenario sheets (SB_Base, SB_Down, Acq_Base, Combined_Risk): balance-check series
  - Scenario_Drivers: integration-spend series

Output: list of ModelNode dicts (fields matching model_nodes.csv target).
"""
from __future__ import annotations

import csv
import hashlib
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

ROOT = Path(__file__).resolve().parent.parent

# ── column letters used in the assumption series block ──────────────────────
SERIES_COLS = ["C", "D", "E", "F", "G", "H"]  # FY2026E … FY2031E
MOIC_COL, IRR_COL = "K", "L"
SCENARIO_COL_RANGE = "C:V"  # quarterly scenario columns (20 quarters)

# ── driver-keyword filters for the assumption-series block ───────────────────
# We include rows whose label contains any of these; gross margin excluded by
# design (it's an intermediate metric, not a model driver for returns calc).
DRIVER_KEYWORDS = [
    "platform growth", "firm ebitda margin", "dso",
    "wip / revenue", "capex / revenue", "exit multiple",
    "integration spend",          # Scenario_Drivers sheet
]

# ── keywords for capital-structure scalar inputs ──────────────────────────────
SCALAR_KEYWORDS = [
    "enterprise value",
    "debt",                        # first-lien / term loan debt
    "sponsor initial cash equity",
    "seller rollover",
    "opening cash",
    "concentration",               # largest ultimate-parent concentration
    "customer concentration",
    "net working capital",
    "nwc",
]

# ── QoE perimeter column → name mapping ──────────────────────────────────────
QOE_PERIMETERS = {
    "C": ("seller", "Alderstone consolidated seller-adjusted EBITDA"),
    "D": ("qoe",    "Alderstone consolidated QoE-normalized EBITDA"),
    "E": ("firm",   "Alderstone consolidated Firm-underwritten EBITDA"),
    "F": ("cov",    "Alderstone consolidated Covenant EBITDA"),
}

# ── scenario rows in Ownership_Returns ───────────────────────────────────────
SCENARIO_ROWS = {
    4: ("base",           "Standalone Base"),
    5: ("down",           "Standalone Downside"),
    7: ("acq",            "Acquisition Base"),
    8: ("combined-risk",  "Combined Risk"),
}


@dataclass
class ModelNode:
    node_id:    str
    name:       str
    kind:       str          # input | assumption_series | output | adjustment_input | model_control | model_control_series
    sheet:      str
    cell:       str          # "B3" or "C40:H40"
    value:      object       # scalar or list
    unit:       str = ""
    definition: str = ""
    period:     str = ""
    perimeter:  str = ""

    def to_dict(self) -> dict:
        v = self.value
        if isinstance(v, list):
            v_str = str([round(x, 6) if isinstance(x, float) else x for x in v])
        elif isinstance(v, float):
            v_str = str(round(v, 9))
        else:
            v_str = str(v) if v is not None else ""
        return {
            "model_node_id": self.node_id,
            "name":          self.name,
            "kind":          self.kind,
            "sheet":         self.sheet,
            "cell":          self.cell,
            "value":         v_str,
            "unit":          self.unit,
            "definition":    self.definition,
            "period":        self.period,
            "perimeter":     self.perimeter,
        }


def _slug(s: str, n: int = 30) -> str:
    return re.sub(r"[^a-z0-9]+", "-", s.lower()).strip("-")[:n]


def _node_id(prefix: str, sheet: str, label: str = "") -> str:
    key = f"{prefix}|{sheet}|{label}"
    h = hashlib.md5(key.encode()).hexdigest()[:6]
    return f"MN-{_slug(label or sheet, 28).upper().replace('-', '-')}-{h}" if not label.startswith("MN-") else label


# ── openpyxl helpers ─────────────────────────────────────────────────────────

def _load(path: Path):
    """Load workbook twice: formula text + computed values."""
    try:
        import openpyxl
    except ModuleNotFoundError:
        sys.exit("openpyxl is required: run `make setup` or `.venv/bin/pip install openpyxl`")
    wb_v = openpyxl.load_workbook(str(path), data_only=True)
    return wb_v


def _cell_val(ws, coord: str):
    return ws[coord].value


def _row_vals(ws, row: int, cols: list[str]) -> list:
    return [ws[f"{c}{row}"].value for c in cols]


def _col_letter(n: int) -> str:
    """Convert 1-based column index to letter(s)."""
    result = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


# ── parsers per structural section ───────────────────────────────────────────

def _parse_inputs_scalars(wb) -> list[ModelNode]:
    """Scalar LOCKED inputs from the Inputs sheet (capital structure + key metrics)."""
    ws = wb["Inputs"]
    nodes: list[ModelNode] = []
    for row in ws.iter_rows(min_row=3, max_row=36, min_col=1, max_col=6):
        label_cell, val_cell, unit_cell, status_cell = row[0], row[1], row[2], row[3]
        label = label_cell.value
        status = status_cell.value if status_cell.value else ""
        value = val_cell.value
        unit = unit_cell.value or ""
        if not isinstance(label, str) or label.strip() == "":
            continue
        if status not in ("LOCKED",):
            continue
        if not isinstance(value, (int, float)):
            continue
        label_l = label.lower()
        if not any(kw in label_l for kw in SCALAR_KEYWORDS):
            continue
        # Build a stable ID from the label
        node_id = f"MN-{_slug(label, 32).upper().replace('-', '-')}"
        node_id = re.sub(r"-+", "-", node_id).strip("-")
        coord = f"B{label_cell.row}"
        period = "Opening / 2026-03-31"
        if "concentration" in label_l:
            period = "FY2025A"
        if "working capital" in label_l or "nwc" in label_l:
            period = "Closing reference"
        nodes.append(ModelNode(
            node_id=node_id, name=label.strip(), kind="input",
            sheet="Inputs", cell=coord, value=float(value), unit=unit,
            period=period, perimeter="Project Keystone enterprise value" if "enterprise" in label_l else "",
        ))
    return nodes


def _parse_qoe_ebitda(wb) -> list[ModelNode]:
    """Adjusted EBITDA totals (row 19) for Seller, QoE, Firm, Covenant perimeters."""
    ws = wb["QoE_Bridge"]
    nodes: list[ModelNode] = []
    # Locate the "Adjusted EBITDA" row
    ebitda_row = None
    for row in ws.iter_rows(min_row=15, max_row=22, min_col=1, max_col=1):
        label = row[0].value
        if isinstance(label, str) and "adjusted ebitda" in label.lower():
            ebitda_row = row[0].row
            break
    if ebitda_row is None:
        return nodes

    for col_letter, (slug, perimeter_name) in QOE_PERIMETERS.items():
        val = _cell_val(ws, f"{col_letter}{ebitda_row}")
        if val is None:
            continue
        node_id = f"MN-{slug.upper()}-EBITDA"
        definitions = {
            "seller": "DEF-EBITDA-SELLER",
            "qoe":    "DEF-EBITDA-QOE",
            "firm":   "DEF-EBITDA-FIRM",
            "cov":    "DEF-EBITDA-COV",
        }
        nodes.append(ModelNode(
            node_id=node_id,
            name=f"{slug.replace('-', ' ').title()}-adjusted EBITDA",
            kind="input", sheet="QoE_Bridge",
            cell=f"{col_letter}{ebitda_row}",
            value=float(val), unit="$mm",
            definition=definitions.get(slug, ""),
            period="FY2025A" + (" / seller adjustment bridge" if slug == "seller" else ""),
            perimeter=perimeter_name,
        ))
    return nodes


def _parse_qoe_adjustments(wb) -> list[ModelNode]:
    """Firm-view adjustments in QoE_Bridge that are firm-specific (non-zero in Firm View only)."""
    ws = wb["QoE_Bridge"]
    nodes: list[ModelNode] = []
    # Key adjustment rows: E7 (integration) and E14 (related-party rent)
    ADJUSTMENT_TARGETS = {
        "E7":  ("MN-INTEGRATION-COST-ADJ",   "Integration and systems costs (Firm View adjustment)", "DEF-ADJ-INTEGRATION"),
        "E14": ("MN-RELATED-PARTY-RENT-NORM", "Related-party rent normalization (Firm View)",         "DEF-ADJ-RENT"),
    }
    for coord, (node_id, name, defn) in ADJUSTMENT_TARGETS.items():
        val = _cell_val(ws, coord)
        if val is None:
            continue
        row_num = int(re.search(r"\d+", coord).group())
        row_label_cell = ws[f"A{row_num}"]
        label = row_label_cell.value or name
        nodes.append(ModelNode(
            node_id=node_id, name=str(label), kind="adjustment_input",
            sheet="QoE_Bridge", cell=coord, value=float(val),
            unit="$mm", definition=defn,
            period="FY2025A", perimeter="Alderstone consolidated Firm-underwritten EBITDA",
        ))
    return nodes


def _parse_assumption_series(wb) -> list[ModelNode]:
    """Multi-year assumption rows from the Inputs sheet (rows 38+)."""
    ws = wb["Inputs"]
    nodes: list[ModelNode] = []

    # Find the header row (year numbers)
    header_row = None
    for row in ws.iter_rows(min_row=36, max_row=42, min_col=1, max_col=8):
        vals = [c.value for c in row]
        if any(isinstance(v, int) and 2020 < v < 2040 for v in vals):
            header_row = row[0].row
            break
    if header_row is None:
        return nodes

    year_row = header_row
    year_cols = []
    for col_idx in range(3, 9):  # C to H
        cell = ws.cell(row=year_row, column=col_idx)
        if isinstance(cell.value, int) and 2020 < cell.value < 2040:
            year_cols.append(_col_letter(col_idx))

    if not year_cols:
        return nodes

    years = [ws[f"{c}{year_row}"].value for c in year_cols]
    period_str = f"FY{years[0]}E-FY{years[-1]}E"

    # Scan assumption rows
    for row in ws.iter_rows(min_row=year_row + 1, max_row=year_row + 50, min_col=1, max_col=9):
        label_cell = row[0]
        unit_cell = row[1]
        label = label_cell.value
        if not isinstance(label, str) or label.strip() == "":
            continue
        label_l = label.lower()
        if not any(kw in label_l for kw in DRIVER_KEYWORDS):
            continue
        values = [ws.cell(row=label_cell.row, column=ws[f"{c}1"].column).value
                  for c in year_cols]
        values = [v for v in values]  # keep None as-is for range reporting
        unit = unit_cell.value or ""

        # Derive scenario prefix and driver name
        parts = [p.strip() for p in label.split(" - ", 1)]
        scenario = parts[0] if len(parts) == 2 else "unknown"
        driver = parts[1] if len(parts) == 2 else label

        scenario_slug = _slug(scenario, 20).upper().replace("-", "-")
        driver_slug = _slug(driver, 20).upper().replace("-", "-")
        node_id = f"MN-{scenario_slug}-{driver_slug}"
        node_id = re.sub(r"-+", "-", node_id).strip("-")

        first_col = year_cols[0]
        last_col = year_cols[-1]
        cell_range = f"{first_col}{label_cell.row}:{last_col}{label_cell.row}"

        perimeter = f"{scenario} scenario"
        nodes.append(ModelNode(
            node_id=node_id, name=label.strip(), kind="assumption_series",
            sheet="Inputs", cell=cell_range,
            value=[v for v in values],
            unit=unit, period=period_str, perimeter=perimeter,
        ))
    return nodes


def _parse_scenario_drivers(wb) -> list[ModelNode]:
    """Integration-spend series from Scenario_Drivers sheet."""
    ws = wb["Scenario_Drivers"]
    nodes: list[ModelNode] = []
    # Scan for integration spend row
    for row in ws.iter_rows(min_row=155, max_row=175, min_col=1, max_col=1):
        label = row[0].value
        if isinstance(label, str) and "integration" in label.lower() and "spend" in label.lower():
            r = row[0].row
            # Read the quarterly range C:V
            vals = [ws.cell(row=r, column=c).value for c in range(3, 23)]  # C to V = cols 3-22
            # Get column letters
            first_col = _col_letter(3)
            last_col = _col_letter(22)
            nodes.append(ModelNode(
                node_id="MN-COMBINED-RISK-INTEGRATION-SPEND",
                name=str(label).strip(),
                kind="assumption_series",
                sheet="Scenario_Drivers",
                cell=f"{first_col}{r}:{last_col}{r}",
                value=vals,
                unit="$mm",
                period="Quarterly 2026-06-30 to 2031-03-31",
                perimeter="Combined Risk scenario",
            ))
    return nodes


def _parse_ownership_returns(wb) -> list[ModelNode]:
    """MOIC and IRR outputs per scenario from Ownership_Returns sheet."""
    ws = wb["Ownership_Returns"]
    nodes: list[ModelNode] = []
    exit_period = "Closing 2026-03-31 to exit 2031-03-31"
    for row_num, (slug, case_name) in SCENARIO_ROWS.items():
        for col_letter, metric, unit in [(MOIC_COL, "MOIC", "x"), (IRR_COL, "IRR", "%")]:
            val = _cell_val(ws, f"{col_letter}{row_num}")
            if val is None:
                continue
            metric_slug = _slug(f"{slug} {metric}", 28).upper().replace("-", "-")
            node_id = f"MN-{metric_slug}"
            node_id = re.sub(r"-+", "-", node_id).strip("-")
            # Normalize: MN-COMBINED-RISK-MOIC etc.
            node_id = node_id.replace("COMBINED-RISK-COMBINED-RISK", "COMBINED-RISK")
            nodes.append(ModelNode(
                node_id=node_id,
                name=f"{case_name} — Gross {metric}",
                kind="output",
                sheet="Ownership_Returns",
                cell=f"{col_letter}{row_num}",
                value=float(val),
                unit=unit,
                period=exit_period,
                perimeter=f"{case_name} scenario",
            ))
    return nodes


def _parse_su_check_cells(wb) -> list[ModelNode]:
    """Sources-and-uses check cells (should be 0)."""
    ws = wb["S&U_Opening"]
    nodes: list[ModelNode] = []
    checks = [
        ("F11", "MN-CHECK-SOURCES-USES", "Sources equals uses check"),
        ("B27", "MN-CHECK-OPENING-BS",   "Opening balance sheet check"),
    ]
    for coord, node_id, name in checks:
        val = _cell_val(ws, coord)
        nodes.append(ModelNode(
            node_id=node_id, name=name, kind="model_control",
            sheet="S&U_Opening", cell=coord,
            value=float(val) if val is not None else 0.0,
            unit="", period="Opening / 2026-03-31", perimeter="",
        ))
    return nodes


def _parse_scenario_balance_checks(wb) -> list[ModelNode]:
    """Balance-check rows from scenario model sheets (row 87, C:V)."""
    nodes: list[ModelNode] = []
    SCENARIO_SHEETS = [
        ("SB_Base",       "MN-CHECK-SB-BASE-BS",  "Standalone Base balance sheet check"),
        ("SB_Down",       "MN-CHECK-SB-DOWN-BS",  "Standalone Downside balance sheet check"),
        ("Acq_Base",      "MN-CHECK-ACQ-BASE-BS",  "Acquisition Base balance sheet check"),
        ("Combined_Risk", "MN-CHECK-COMBINED-BS",  "Combined Risk balance sheet check"),
    ]
    first_col = _col_letter(3)  # C
    last_col  = _col_letter(22)  # V
    for sheet_name, node_id, name in SCENARIO_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        # Scan for the balance-check row near row 87
        check_row = None
        for r in range(80, 95):
            label = ws[f"A{r}"].value
            if isinstance(label, str) and ("balance" in label.lower() or "check" in label.lower() or "bs check" in label.lower()):
                check_row = r
                break
        if check_row is None:
            check_row = 87  # default fallback
        vals = [ws.cell(row=check_row, column=c).value for c in range(3, 23)]
        nodes.append(ModelNode(
            node_id=node_id, name=name, kind="model_control_series",
            sheet=sheet_name,
            cell=f"{first_col}{check_row}:{last_col}{check_row}",
            value=vals,
            unit="", period="Quarterly 2026-06-30 to 2031-03-31", perimeter="",
        ))
    return nodes


# ── main parse function ───────────────────────────────────────────────────────

def parse_workbook(path: Path) -> list[ModelNode]:
    """Parse an LBO model xlsx and return all discovered model nodes."""
    wb = _load(path)
    nodes: list[ModelNode] = []
    nodes += _parse_inputs_scalars(wb)
    nodes += _parse_qoe_ebitda(wb)
    nodes += _parse_qoe_adjustments(wb)
    nodes += _parse_assumption_series(wb)
    nodes += _parse_scenario_drivers(wb)
    nodes += _parse_ownership_returns(wb)
    nodes += _parse_su_check_cells(wb)
    nodes += _parse_scenario_balance_checks(wb)
    # De-duplicate by cell coordinate
    seen: set[str] = set()
    deduped: list[ModelNode] = []
    for n in nodes:
        key = f"{n.sheet}!{n.cell}"
        if key not in seen:
            seen.add(key)
            deduped.append(n)
    return deduped


def to_csv_rows(nodes: list[ModelNode]) -> list[dict]:
    return [n.to_dict() for n in nodes]


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    import argparse
    ap = argparse.ArgumentParser(description="Parse LBO model xlsx → model nodes CSV")
    ap.add_argument("xlsx", help="Path to the xlsx file")
    ap.add_argument("--out", default="-", help="Output CSV path (default: stdout)")
    ap.add_argument("--score", metavar="TARGET_CSV",
                    help="Score against a target model_nodes.csv and print precision/recall")
    args = ap.parse_args()

    path = Path(args.xlsx)
    if not path.exists():
        sys.exit(f"File not found: {path}")

    nodes = parse_workbook(path)
    rows = to_csv_rows(nodes)

    FIELDS = ["model_node_id", "name", "kind", "sheet", "cell", "value",
              "unit", "definition", "period", "perimeter"]

    if args.out == "-":
        import io
        buf = io.StringIO()
        w = csv.DictWriter(buf, fieldnames=FIELDS, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
        print(buf.getvalue(), end="")
    else:
        out = Path(args.out)
        out.parent.mkdir(parents=True, exist_ok=True)
        with open(out, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=FIELDS, extrasaction="ignore")
            w.writeheader()
            w.writerows(rows)
        print(f"Wrote {len(nodes)} nodes → {out}")

    if args.score:
        _score(nodes, Path(args.score))


def _normalize_cell(c: str) -> str:
    """Normalize a cell reference for comparison (remove spaces, uppercase)."""
    return c.strip().upper().replace(" ", "")


def _score(nodes: list[ModelNode], target_csv: Path, silent: bool = False) -> dict:
    """Compare parsed nodes against the canonical target CSV. Returns metrics dict."""
    with open(target_csv, encoding="utf-8-sig") as f:
        targets = list(csv.DictReader(f))

    # Index targets by cell coordinate (sheet!cell)
    def tkey(t: dict) -> str:
        return f"{t['sheet']}!{_normalize_cell(t['cell'])}"

    target_by_coord = {tkey(t): t for t in targets}
    parsed_by_coord = {f"{n.sheet}!{_normalize_cell(n.cell)}": n for n in nodes}

    tp_coords = set(target_by_coord) & set(parsed_by_coord)
    fp_coords = set(parsed_by_coord) - set(target_by_coord)
    fn_coords = set(target_by_coord) - set(parsed_by_coord)

    precision = len(tp_coords) / len(parsed_by_coord) if parsed_by_coord else 0
    recall    = len(tp_coords) / len(target_by_coord) if target_by_coord else 0
    f1        = 2 * precision * recall / (precision + recall) if (precision + recall) else 0

    # Per-kind breakdown
    kind_targets: dict[str, list] = {}
    for t in targets:
        kind_targets.setdefault(t["kind"], []).append(t)
    per_kind = {
        kind: {"tp": sum(1 for t in kt if tkey(t) in tp_coords), "total": len(kt)}
        for kind, kt in sorted(kind_targets.items())
    }

    if not silent:
        print(f"\n── Model Parser Benchmark ──────────────────────────────")
        print(f"  Target nodes:   {len(targets):3d}")
        print(f"  Parsed nodes:   {len(nodes):3d}")
        print(f"  True positive:  {len(tp_coords):3d}  (found & correct)")
        print(f"  False positive: {len(fp_coords):3d}  (found, not in target)")
        print(f"  False negative:  {len(fn_coords):3d}  (in target, not found)")
        print(f"  Precision:    {precision:.1%}")
        print(f"  Recall:       {recall:.1%}")
        print(f"  F1:           {f1:.1%}")

        if fp_coords:
            print(f"\n  False positives (extra nodes we generated):")
            for coord in sorted(fp_coords):
                n = parsed_by_coord[coord]
                print(f"    {n.node_id:<35} {coord}")

        if fn_coords:
            print(f"\n  False negatives (target nodes we missed):")
            for coord in sorted(fn_coords):
                t = target_by_coord[coord]
                print(f"    {t['model_node_id']:<35} {coord}  ({t['name'][:40]})")

        print(f"\n  Per-kind recall:")
        for kind, v in per_kind.items():
            print(f"    {kind:<25} {v['tp']}/{v['total']}")

    return {
        "total": len(targets),
        "parsed": len(nodes),
        "tp": len(tp_coords),
        "fp": len(fp_coords),
        "fn": len(fn_coords),
        "precision": round(precision, 3),
        "recall": round(recall, 3),
        "f1": round(f1, 3),
        "per_kind": per_kind,
        "fp_list": sorted(fp_coords),
        "fn_list": sorted(fn_coords),
    }


if __name__ == "__main__":
    main()
