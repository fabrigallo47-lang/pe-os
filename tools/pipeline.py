#!/usr/bin/env python3
"""
PANTA Extraction Pipeline — standalone entry point for Anto's tests.

Given any source file (PDF, xlsx, markdown, txt), this script:
  1. Converts the file to text
  2. Runs LLM claim extraction
  3. Writes a canonical claims JSON
  4. Builds a semantic graph and saves it as a graph JSON + SQLite-ready structure

Usage
-----
    .venv/bin/python3 tools/pipeline.py <file> [options]

    .venv/bin/python3 tools/pipeline.py docs/cim.pdf
    .venv/bin/python3 tools/pipeline.py model.xlsx --deal keystone
    .venv/bin/python3 tools/pipeline.py notes.md   --deal keystone --out /tmp/out/

    Outputs (all in --out dir, default: ./pipeline_out/):
        claims.json       — extracted claims list (one object per claim)
        graph.json        — semantic graph (node-link format, loads into DealGraph)
        graph.db          — SQLite with nodes + edges tables (for Anto's algo)

Environment
-----------
    ANTHROPIC_API_KEY   required for extraction step
    PEOS_MODEL          model name (default: claude-sonnet-5)
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import shutil
import sqlite3
import subprocess
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from tools.extract import SYSTEM_PROMPT, llm_extract, parse_json
from vercel.api._claim_graph import claims_to_graph
from tools.graph_store import DealGraph


# ── File → text ──────────────────────────────────────────────────────────────

def _pdf_to_text(path: Path) -> str:
    data = path.read_bytes()
    pdftotext = (
        shutil.which("pdftotext") or
        next((p for p in [
            "/opt/homebrew/bin/pdftotext",
            "/usr/local/bin/pdftotext",
            "/usr/bin/pdftotext",
        ] if os.path.isfile(p)), None)
    )
    if pdftotext:
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
            f.write(data)
            tmp = f.name
        try:
            r = subprocess.run(
                [pdftotext, "-layout", tmp, "-"],
                capture_output=True, text=True, timeout=60,
            )
            if r.returncode == 0 and r.stdout.strip():
                return r.stdout
        finally:
            os.unlink(tmp)
    try:
        import pypdf, io
        reader = pypdf.PdfReader(io.BytesIO(data))
        return "\n\n".join(p.extract_text() or "" for p in reader.pages)
    except ImportError:
        pass
    raise RuntimeError(
        "Cannot convert PDF. Install poppler: brew install poppler  "
        "or: .venv/bin/pip install pypdf"
    )


def _xlsx_to_text(path: Path) -> str:
    """Convert a workbook to evidence text without discarding formulas or cell locators."""
    try:
        import openpyxl
        # Formula text identifies a model relationship; cached values identify
        # the last calculated value. Both are evidence and neither may replace
        # the other silently.
        wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
        cached = openpyxl.load_workbook(path, data_only=True, read_only=True)
        parts: list[str] = []
        for name in wb.sheetnames:
            ws = wb[name]
            cached_ws = cached[name]
            lines: list[str] = []
            for row, cached_row in zip(ws.iter_rows(), cached_ws.iter_rows()):
                cells: list[str] = []
                for cell, cached_cell in zip(row, cached_row):
                    value = cell.value
                    if value is None:
                        continue
                    if isinstance(value, str) and value.startswith("="):
                        cached_value = cached_cell.value
                        cells.append(f"{cell.coordinate}=FORMULA({value}); cached={cached_value!r}")
                    else:
                        cells.append(f"{cell.coordinate}={value}")
                if cells:
                    lines.append(" | ".join(cells))
            if not lines:
                continue
            parts.append(f"## Sheet: {name}\n" + "\n".join(lines))
        return "\n\n".join(parts)
    except ImportError as exc:
        raise RuntimeError("Cannot convert workbook. Install openpyxl: .venv/bin/pip install openpyxl") from exc
    # Fallback: xlsx_parser for LBO-style workbooks
    try:
        from tools.xlsx_parser import parse_workbook
        nodes = parse_workbook(path)
        lines = [f"{n.node_id}: {n.label} = {n.value} {n.unit or ''}"
                 for n in nodes]
        return "\n".join(lines)
    except Exception:
        pass
    raise RuntimeError("Cannot convert workbook; openpyxl could not read it")


def file_to_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return _pdf_to_text(path)
    if suffix in (".xlsx", ".xlsm"):
        return _xlsx_to_text(path)
    if suffix == ".xls":
        raise RuntimeError("Legacy .xls is not supported; save it as .xlsx before ingesting.")
    # Plaintext formats: md, txt, csv, html, json
    return path.read_text(encoding="utf-8", errors="replace")


# ── Extraction ────────────────────────────────────────────────────────────────

def extract_claims(text: str, deal: str, source_name: str) -> list[dict]:
    """Run LLM extraction on text. Returns list of raw claim dicts."""
    if len(text) > 100_000:
        print(f"  [truncating {len(text):,} → 100,000 chars]")
        text = text[:100_000]

    user = (
        f"DEAL: {deal}\n"
        f"DEAL QUESTIONS:\n(no pre-loaded questions)\n\n"
        f"EXISTING SUBJECTS (reuse when same quantity/basis): []\n\n"
        f"ARTIFACT ({source_name}):\n{text}"
    )
    print(f"  Calling extraction API for '{source_name}' ({len(text):,} chars)…")
    raw = llm_extract(SYSTEM_PROMPT, user)
    claims = parse_json(raw)
    print(f"  Extracted {len(claims)} claims")
    return claims


# ── Graph DB (SQLite) ─────────────────────────────────────────────────────────

def build_sqlite(graph_json: dict, db_path: Path) -> None:
    """Write nodes + edges tables into a SQLite file from a node-link graph dict."""
    db_path.parent.mkdir(parents=True, exist_ok=True)
    con = sqlite3.connect(db_path)
    con.execute("DROP TABLE IF EXISTS nodes")
    con.execute("DROP TABLE IF EXISTS edges")
    con.execute("""
        CREATE TABLE nodes (
            id      TEXT PRIMARY KEY,
            type    TEXT,
            label   TEXT,
            data    TEXT   -- full JSON blob of all node attrs
        )
    """)
    con.execute("""
        CREATE TABLE edges (
            src     TEXT,
            tgt     TEXT,
            rel     TEXT,
            data    TEXT,  -- full JSON blob of all edge attrs
            PRIMARY KEY (src, tgt, rel)
        )
    """)

    # node-link format from networkx: {"nodes": [...], "edges": [...]}
    for n in graph_json.get("nodes", []):
        nid = str(n.get("id", ""))
        con.execute(
            "INSERT OR REPLACE INTO nodes (id, type, label, data) VALUES (?,?,?,?)",
            (nid, n.get("type", ""), n.get("label", ""), json.dumps(n)),
        )
    for e in graph_json.get("edges", []):
        src = str(e.get("source", ""))
        tgt = str(e.get("target", ""))
        rel = str(e.get("rel", e.get("label", "")))
        con.execute(
            "INSERT OR REPLACE INTO edges (src, tgt, rel, data) VALUES (?,?,?,?)",
            (src, tgt, rel, json.dumps(e)),
        )

    con.commit()
    con.close()


# ── Stats ─────────────────────────────────────────────────────────────────────

def print_stats(dg: DealGraph, claims: list[dict]) -> None:
    stats = dg.stats()
    print(f"\n  Graph stats:")
    print(f"    nodes        : {stats['nodes']}")
    print(f"    edges        : {stats['edges']}")
    print(f"    claims       : {len(claims)}")
    contras = dg.contradictions()
    if contras:
        print(f"    contradictions: {len(contras)}")
        for c in contras[:3]:
            a = c.get("source", {}).get("label", c.get("source", {}).get("id", "?"))
            b = c.get("target", {}).get("label", c.get("target", {}).get("id", "?"))
            print(f"      {a} ↔ {b}")
    top = dg.top_by_pagerank(n=5)
    if top:
        print(f"    top claims (PageRank):")
        for n in top:
            print(f"      {n['label']}")


# ── Main ──────────────────────────────────────────────────────────────────────

def e3_to_claims(e3: dict) -> list[dict]:
    """Convert an extract_v2 manifest into the claim shape this pipeline emits.

    CAP-003 keeps the claim record to frozen fields and parks compiler metadata
    (metric, direction, topic, author, derivation) alongside it, so the two are
    joined back on claim_id here.
    """
    sidecar = {
        m.get("claim_id"): m
        for m in e3.get("extraction_metadata", {}).get("compiler_fields_per_claim", [])
        if m.get("claim_id")
    }
    out = []
    for c in e3.get("claims", []):
        meta = sidecar.get(c.get("claim_id"), {})
        perimeter = c.get("perimeter") or ""
        out.append({
            # subject is the economic scope the claim is about; the extractor
            # records that as the perimeter.
            "subject":    perimeter.split(",")[0].strip() or "unknown",
            "metric":     meta.get("metric", ""),
            "value":      c.get("value"),
            "unit":       c.get("unit", ""),
            "as_of":      c.get("period") or "",
            "period":     c.get("period") or "",
            "perimeter":  perimeter,
            "topic":      meta.get("topic", ""),
            "source_doc": c.get("source_id", ""),
            "epistemic":  c.get("epistemic_class", "asserted"),
            "direction":  meta.get("direction", "context"),
            "bears_on":   [],
            "locator":    c.get("locator", ""),
            "author":     meta.get("author"),
            "statement":  c.get("statement", ""),
            "derivation": meta.get("derivation"),
        })
    return out


def _run_from_e3(e3_path: Path, out_dir: Path, deal: str) -> None:
    if not e3_path.exists():
        print(f"ERROR: file not found: {e3_path}", file=sys.stderr)
        sys.exit(1)
    out_dir.mkdir(parents=True, exist_ok=True)
    e3 = json.loads(e3_path.read_text(encoding="utf-8"))
    claims = e3_to_claims(e3)

    print("\n=== PANTA Extraction Pipeline (from E3) ===")
    print(f"  source : {e3_path}")
    print(f"  deal   : {deal}")
    print(f"  output : {out_dir}\n")

    claims_path = out_dir / "claims.json"
    claims_path.write_text(json.dumps(claims, indent=2, ensure_ascii=False))
    print(f"[1/4] {len(claims)} claim da {e3.get('manifest_id', '?')} → {claims_path.name}")

    print("[2/4] Building semantic graph…")
    graph_dict = claims_to_graph(claims, source_name=e3.get("manifest_id", "E3"))
    dg = DealGraph(deal=deal)
    dg.load_from_claims_graph(claims, graph_dict)
    graph_json = dg.to_json()
    (out_dir / "graph.json").write_text(
        json.dumps(graph_json, indent=2, ensure_ascii=False))
    print_stats(dg, claims)

    # storage_export writes graph.db, nodes.csv and edges.csv from this one
    # snapshot, so the three reconcile row by row rather than merely in counts.
    print("\n[3/4] Writing SQLite + CSVs…")
    if "links" in graph_json and "edges" not in graph_json:
        graph_json["edges"] = graph_json.pop("links")
    from tools import storage_export
    errs = storage_export.validate_graph(graph_json)
    if errs:
        print("  graph non esportabile:", *errs[:5], sep="\n   - ")
        sys.exit(1)
    stats = storage_export.export(graph_json, out_dir)
    print(f"  {stats['nodes']} nodes, {stats['edges']} edges → graph.db / nodes.csv / edges.csv")

    print("\n[4/4] Reconciling…")
    problems = storage_export.verify(out_dir, graph_json)
    if problems:
        print("  RICONCILIAZIONE FALLITA:", *problems, sep="\n   - ")
        sys.exit(1)
    print("  graph.json ↔ CSV ↔ SQLite coerenti")
    print(f"\nDone → {out_dir}")


def main() -> None:
    ap = argparse.ArgumentParser(
        description="PANTA extraction pipeline: file → claims JSON + graph DB",
    )
    ap.add_argument("file", nargs="?",
                    help="Input file (PDF, xlsx, md, txt, csv, …)")
    ap.add_argument("--from-e3", metavar="E3_CLAIMS",
                    help="Build the graph from an existing extract_v2 e3_claims.json "
                         "instead of extracting. No API key needed.")
    ap.add_argument("--deal", default="pipeline", help="Deal slug (default: pipeline)")
    ap.add_argument("--out",  default="pipeline_out", help="Output directory")
    ap.add_argument("--no-api", action="store_true",
                    help="Skip LLM extraction (use if claims.json already exists)")
    args = ap.parse_args()

    if args.from_e3:
        return _run_from_e3(Path(args.from_e3).expanduser().resolve(),
                            Path(args.out).expanduser().resolve(), args.deal)

    if not args.file:
        ap.error("serve un file di input, oppure --from-e3")
    src = Path(args.file).expanduser().resolve()
    if not src.exists():
        print(f"ERROR: file not found: {src}", file=sys.stderr)
        sys.exit(1)

    out_dir = Path(args.out).expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    claims_path = out_dir / "claims.json"
    graph_path  = out_dir / "graph.json"
    db_path     = out_dir / "graph.db"

    print(f"\n=== PANTA Extraction Pipeline ===")
    print(f"  source : {src}")
    print(f"  deal   : {args.deal}")
    print(f"  output : {out_dir}\n")

    # ── 1. File → text ────────────────────────────────────────────────────────
    if not args.no_api:
        print("[1/4] Converting file to text…")
        try:
            text = file_to_text(src)
            print(f"  {len(text):,} characters extracted from {src.name}")
        except RuntimeError as e:
            print(f"ERROR: {e}", file=sys.stderr)
            sys.exit(1)

        # ── 2. Extract claims ─────────────────────────────────────────────────
        print("[2/4] Running LLM extraction…")
        if not os.environ.get("ANTHROPIC_API_KEY"):
            print("ERROR: ANTHROPIC_API_KEY not set", file=sys.stderr)
            sys.exit(1)

        claims = extract_claims(text, args.deal, src.name)
        claims_path.write_text(json.dumps(claims, indent=2, ensure_ascii=False))
        print(f"  Saved {len(claims)} claims → {claims_path}")
    else:
        if not claims_path.exists():
            print(f"ERROR: --no-api set but {claims_path} does not exist", file=sys.stderr)
            sys.exit(1)
        print("[1-2/4] Skipping extraction (--no-api); loading existing claims…")
        claims = json.loads(claims_path.read_text())
        print(f"  Loaded {len(claims)} claims from {claims_path}")

    # ── 3. Claims → semantic graph ────────────────────────────────────────────
    print("[3/4] Building semantic graph…")
    graph_dict = claims_to_graph(claims, source_name=src.name)
    dg = DealGraph(deal=args.deal)
    dg.load_from_claims_graph(claims, graph_dict)

    graph_json = dg.to_json()
    graph_path.write_text(json.dumps(graph_json, indent=2, ensure_ascii=False))
    print(f"  Graph saved → {graph_path}")
    print_stats(dg, claims)

    # ── 4. Graph → SQLite ─────────────────────────────────────────────────────
    print("\n[4/4] Writing SQLite database…")
    build_sqlite(graph_json, db_path)
    con = sqlite3.connect(db_path)
    con.row_factory = sqlite3.Row
    n_nodes = con.execute("SELECT COUNT(*) FROM nodes").fetchone()[0]
    n_edges = con.execute("SELECT COUNT(*) FROM edges").fetchone()[0]
    print(f"  {n_nodes} nodes, {n_edges} edges → {db_path}")

    # ── 5. SQLite → CSV ───────────────────────────────────────────────────────
    print("\n[5/5] Writing CSVs…")
    nodes_csv = out_dir / "nodes.csv"
    edges_csv = out_dir / "edges.csv"

    rows = con.execute("SELECT * FROM nodes").fetchall()
    with open(nodes_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["id", "type", "label", "data"])
        w.writeheader()
        w.writerows([dict(r) for r in rows])

    rows = con.execute("SELECT * FROM edges").fetchall()
    with open(edges_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["src", "tgt", "rel", "data"])
        w.writeheader()
        w.writerows([dict(r) for r in rows])

    con.close()
    print(f"  {n_nodes} rows → {nodes_csv}")
    print(f"  {n_edges} rows → {edges_csv}")

    print(f"\nDone. Hand Anto:")
    print(f"  claims     : {claims_path}")
    print(f"  graph      : {graph_path}")
    print(f"  sqlite     : {db_path}")
    print(f"  nodes.csv  : {nodes_csv}")
    print(f"  edges.csv  : {edges_csv}")
    print()


if __name__ == "__main__":
    main()
