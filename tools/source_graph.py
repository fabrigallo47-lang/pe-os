#!/usr/bin/env python3
"""
source_graph — L1: a lossless record of what a workbook contains.

Architecture this belongs to
----------------------------
  L1  source graph   what the files literally contain           (this module)
  L2  proposals      an LLM proposes semantic identities
  L3  resolver       a global solver decides which bindings are admissible

The split exists because neither extreme works. A fully agentic extractor can
read the same workbook twice and interpret it differently. A purely
deterministic parser reads cells and text perfectly and understands no
economics. So L1 commits to representation only, L2 to proposal only, and L3
decides — over the whole deal at once, as a constraint system, never cell by
cell.

L1 therefore contains no interpretation whatsoever. No concept names, no
confidence, no units inferred from wording. If a downstream layer disagrees
about meaning, it must be able to come back here and find the original
untouched.

Why the earlier attempts were not this
--------------------------------------
tools/xlsx_parser.py reads cached values with data_only=True and recognises
sheets by hard-coded name, yielding 59 named concepts on Keystone and no ability
to recompute. tools/extract_v3.py builds a real formula dependency graph — 29476
cells, 22971 edges — but stores only {sheet, ref, value} per cell, discarding
the formula text, the number format and the 71 merged ranges. Neither is
lossless, so neither can serve as L1.

What is recorded
----------------
  cells        raw value, formula text, number format, type, merge role
  precedents   references each formula depends on, parsed from its text
  ranges       merged ranges, native tables, defined names
  sheets       dimensions, visibility
  provenance   file digest and size, so a re-import can be diffed against this

Locators are `SHEET!REF` throughout — the same address the workbook uses, so a
claim can point back into the source without translation.

    python3 tools/source_graph.py --workbook model.xlsx --out DIR
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from dataclasses import dataclass, field, asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

# ── reference parsing ────────────────────────────────────────────────────────

# 'Sheet Name'!$A$1:$B$2 | Sheet!A1 | $A$1:$B$2 | A1
_REF_RE = re.compile(
    r"(?:(?P<q>'[^']+'|[A-Za-z_][A-Za-z0-9_.]*)!)?"
    r"(?P<c1>\$?[A-Z]{1,3}\$?\d+)"
    r"(?::(?P<c2>\$?[A-Z]{1,3}\$?\d+))?"
)
# Text inside quotes must not be mined for references, and neither must the
# names of functions that merely look like column letters.
_STRING_RE = re.compile(r'"[^"]*"')
_SUPPORTED_FUNCTIONS = {
    "ABS", "AND", "AVERAGE", "CHOOSE", "COUNT", "COUNTIF", "COUNTIFS",
    "FALSE", "HLOOKUP", "IF", "IFERROR", "INDEX", "IRR", "MATCH", "MAX",
    "MIN", "NA", "NOT", "NPV", "OR", "PMT", "ROUND", "SUM", "SUMIF",
    "SUMIFS", "TRUE", "VLOOKUP", "XIRR", "XLOOKUP", "XNPV",
}
_FUNCTION_RE = re.compile(r"(?<![A-Za-z0-9_.])([A-Za-z_][A-Za-z0-9_.]*)\s*\(")
_EXTERNAL_REF_RE = re.compile(
    r"(?:'(?P<quoted>\[[^\]]+\][^']+)'|(?P<plain>\[[^\]]+\][A-Za-z_][A-Za-z0-9_. ]*))!"
    r"(?P<c1>\$?[A-Z]{1,3}\$?\d+)(?::(?P<c2>\$?[A-Z]{1,3}\$?\d+))?",
    re.IGNORECASE,
)


def _strip(ref: str) -> str:
    return ref.replace("$", "").upper()


def external_references(formula: str) -> list[str]:
    """Return workbook-qualified references without pretending they are local."""
    if not formula or not formula.startswith("="):
        return []
    references = []
    for match in _EXTERNAL_REF_RE.finditer(_STRING_RE.sub('\"\"', formula)):
        book_sheet = match.group("quoted") or match.group("plain") or ""
        ref = _strip(match.group("c1"))
        if match.group("c2"):
            ref += ":" + _strip(match.group("c2"))
        locator = f"{book_sheet.upper()}!{ref}"
        if locator not in references:
            references.append(locator)
    return references


def formula_functions(formula: str) -> list[str]:
    """List function names exactly enough to disclose unsupported evaluation."""
    if not formula or not formula.startswith("="):
        return []
    body = _STRING_RE.sub('\"\"', formula)
    return sorted({match.group(1).upper() for match in _FUNCTION_RE.finditer(body)})


def _named_reference_destinations(
    formula: str,
    defined_names: dict[str, str] | None,
) -> dict[str, str]:
    if not formula or not defined_names:
        return {}
    body = _STRING_RE.sub('\"\"', formula)
    found = {}
    for name, destination in defined_names.items():
        pattern = rf"(?<![A-Za-z0-9_.]){re.escape(name)}(?![A-Za-z0-9_.(])"
        if re.search(pattern, body, re.IGNORECASE):
            found[str(name)] = str(destination)
    return found


def parse_precedents(
    formula: str,
    own_sheet: str,
    defined_names: dict[str, str] | None = None,
) -> list[str]:
    """
    Locators a formula reads. A range is kept whole (SHEET!A1:B2) rather than
    expanded: the workbook expressed a range, and L1 records what it expressed.
    """
    if not formula or not formula.startswith("="):
        return []
    body = _STRING_RE.sub('""', formula)
    # An external reference is not a dependency on a similarly named local
    # cell. Keep it in ``external_references`` and remove it from this pass.
    body = _EXTERNAL_REF_RE.sub("", body)
    out: list[str] = []
    for m in _REF_RE.finditer(body):
        sheet = m.group("q")
        if sheet:
            sheet = sheet.strip("'")
        else:
            # A bare token immediately followed by "(" is a function call.
            end = m.end()
            if end < len(body) and body[end] == "(":
                continue
            if _strip(m.group("c1")) in _SUPPORTED_FUNCTIONS:
                continue
            sheet = own_sheet
        ref = _strip(m.group("c1"))
        if m.group("c2"):
            ref = f"{ref}:{_strip(m.group('c2'))}"
        loc = f"{sheet.upper()}!{ref}"
        if loc not in out:
            out.append(loc)
    for destination in _named_reference_destinations(formula, defined_names).values():
        # Defined names may be constants or dynamic formulas. Only direct
        # cell/range destinations become edges; all names remain preserved on
        # the CellRecord for review.
        destination_body = destination.removeprefix("=")
        for match in _REF_RE.finditer(destination_body):
            sheet = (match.group("q") or own_sheet).strip("'")
            ref = _strip(match.group("c1"))
            if match.group("c2"):
                ref += ":" + _strip(match.group("c2"))
            locator = f"{sheet.upper()}!{ref}"
            if locator not in out:
                out.append(locator)
    return out


# ── records ──────────────────────────────────────────────────────────────────

@dataclass
class CellRecord:
    locator: str                       # SHEET!B3
    sheet: str
    ref: str
    row: int
    col: int
    kind: str                          # formula | number | text | bool | date | blank
    value: Any = None                  # raw stored value (formula text if a formula)
    number_format: str = ""
    precedents: list[str] = field(default_factory=list)
    merged_into: str | None = None     # anchor locator when part of a merge
    # Excel's last computed value. Often None here: a workbook written by a
    # script and never opened in Excel caches nothing, which is precisely why
    # reading with data_only=True loses the model.
    cached_value: Any = None
    function_names: list[str] = field(default_factory=list)
    unsupported_functions: list[str] = field(default_factory=list)
    external_references: list[str] = field(default_factory=list)
    named_references: dict[str, str] = field(default_factory=dict)
    formula_status: str | None = None
    evaluation_status: str | None = None
    human_stop_reason: str | None = None
    # Deterministically recomputed only when the formula lies in an acyclic
    # region that the evaluator supports. Never substitutes source evidence.
    evaluated_value: Any = None


@dataclass
class SheetRecord:
    name: str
    max_row: int
    max_col: int
    state: str = "visible"
    merged_ranges: list[str] = field(default_factory=list)
    tables: list[dict] = field(default_factory=list)


@dataclass
class SourceGraph:
    workbook: str
    digest: str
    size_bytes: int
    captured_at: str
    sheets: list[SheetRecord] = field(default_factory=list)
    cells: dict[str, CellRecord] = field(default_factory=dict)
    defined_names: dict[str, str] = field(default_factory=dict)

    def stats(self) -> dict:
        kinds: dict[str, int] = {}
        formula_statuses: dict[str, int] = {}
        for c in self.cells.values():
            kinds[c.kind] = kinds.get(c.kind, 0) + 1
            if c.formula_status:
                formula_statuses[c.formula_status] = formula_statuses.get(c.formula_status, 0) + 1
        return {
            "sheets": len(self.sheets),
            "cells": len(self.cells),
            "by_kind": kinds,
            "precedent_edges": sum(len(c.precedents) for c in self.cells.values()),
            "merged_ranges": sum(len(s.merged_ranges) for s in self.sheets),
            "defined_names": len(self.defined_names),
            "formula_statuses": formula_statuses,
            "human_stops": sum(c.evaluation_status == "HUMAN_STOP" for c in self.cells.values()),
        }

    def to_json(self) -> dict:
        return {
            "schema": "source-graph-1",
            "workbook": self.workbook,
            "digest": self.digest,
            "size_bytes": self.size_bytes,
            "captured_at": self.captured_at,
            "sheets": [asdict(s) for s in self.sheets],
            "defined_names": self.defined_names,
            "cells": {k: asdict(v) for k, v in self.cells.items()},
            "stats": self.stats(),
        }


# ── capture ──────────────────────────────────────────────────────────────────

def _kind_of(value: Any, is_formula: bool) -> str:
    if is_formula:
        return "formula"
    if value is None:
        return "blank"
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, (int, float)):
        return "number"
    if isinstance(value, datetime):
        return "date"
    return "text"


def capture(path: Path) -> SourceGraph:
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.exit("serve openpyxl")

    raw = path.read_bytes()
    graph = SourceGraph(
        workbook=path.name,
        digest="sha256:" + hashlib.sha256(raw).hexdigest(),
        size_bytes=len(raw),
        captured_at=datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
    )

    # data_only=False keeps the formulas; the cached values are a separate read
    # because openpyxl cannot expose both from one load.
    wb = openpyxl.load_workbook(str(path), data_only=False)
    try:
        wb_vals = openpyxl.load_workbook(str(path), data_only=True)
    except Exception:
        wb_vals = None

    for name, dest in getattr(wb, "defined_names", {}).items():
        try:
            graph.defined_names[str(name)] = str(dest.value)
        except Exception:
            graph.defined_names[str(name)] = ""

    for ws in wb:
        merged = [str(r) for r in ws.merged_cells.ranges]
        graph.sheets.append(SheetRecord(
            name=ws.title, max_row=ws.max_row, max_col=ws.max_column,
            state=getattr(ws, "sheet_state", "visible"),
            merged_ranges=merged,
            tables=[{"name": n, "ref": str(t.ref)}
                    for n, t in getattr(ws, "tables", {}).items()],
        ))

        # anchor lookup so a merged cell knows which cell carries its content
        anchor_of: dict[str, str] = {}
        for rng in ws.merged_cells.ranges:
            anchor = f"{ws.title.upper()}!{get_column_letter(rng.min_col)}{rng.min_row}"
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    loc = f"{ws.title.upper()}!{get_column_letter(c)}{r}"
                    if loc != anchor:
                        anchor_of[loc] = anchor

        vs = wb_vals[ws.title] if wb_vals is not None else None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None and cell.coordinate not in anchor_of:
                    continue
                is_formula = isinstance(cell.value, str) and cell.value.startswith("=")
                loc = f"{ws.title.upper()}!{cell.coordinate}"
                rec = CellRecord(
                    locator=loc, sheet=ws.title.upper(), ref=cell.coordinate,
                    row=cell.row, col=cell.column,
                    kind=_kind_of(cell.value, is_formula),
                    value=cell.value,
                    number_format=cell.number_format or "",
                    precedents=(
                        parse_precedents(cell.value, ws.title, graph.defined_names)
                        if is_formula else []
                    ),
                    merged_into=anchor_of.get(loc),
                )
                # A formula's last computed value is part of what the file
                # contains, so it is recorded — clearly separated from the
                # formula itself rather than replacing it.
                if is_formula and vs is not None:
                    try:
                        rec.cached_value = vs[cell.coordinate].value  # type: ignore[attr-defined]
                    except Exception:
                        pass
                if is_formula:
                    rec.function_names = formula_functions(str(cell.value))
                    rec.unsupported_functions = sorted(
                        set(rec.function_names) - _SUPPORTED_FUNCTIONS
                    )
                    rec.external_references = external_references(str(cell.value))
                    rec.named_references = _named_reference_destinations(
                        str(cell.value), graph.defined_names
                    )
                    if rec.external_references:
                        rec.formula_status = "EXTERNAL_LINK"
                        rec.evaluation_status = "HUMAN_STOP"
                        rec.human_stop_reason = (
                            "External workbook dependency is unavailable; cached_value is display-only."
                        )
                    elif rec.unsupported_functions:
                        rec.formula_status = "UNSUPPORTED_FUNCTION"
                        rec.evaluation_status = "HUMAN_STOP"
                        rec.human_stop_reason = (
                            "Unsupported Excel function(s): "
                            + ", ".join(rec.unsupported_functions)
                            + "; no value was evaluated."
                        )
                    else:
                        rec.formula_status = "PRESERVED"
                        if rec.cached_value is None:
                            rec.evaluation_status = "HUMAN_STOP"
                            rec.human_stop_reason = (
                                "The workbook has no cached/displayed value and capture does not calculate formulas."
                            )
                        else:
                            rec.evaluation_status = "CACHED_VALUE_AVAILABLE"
                graph.cells[loc] = rec
    _evaluate_acyclic_formulas(path, graph)
    return graph


def _evaluate_acyclic_formulas(path: Path, graph: SourceGraph) -> None:
    """Evaluate supported acyclic formulas without obscuring Excel evidence.

    Workbooks created by code often have no cached formula result. Cyclic and
    unsupported formulas remain explicit unknowns, never coerced to a value.
    """
    formula_cells = [cell for cell in graph.cells.values() if cell.kind == "formula"]
    if not formula_cells:
        return
    try:
        from tools.extract_v3 import compile_workbook
        compiled = compile_workbook(path)
    except Exception as exc:
        for cell in formula_cells:
            cell.evaluation_status = f"UNAVAILABLE:{type(exc).__name__}"
        return

    cyclic = compiled.cyclic_cells
    for cell in formula_cells:
        if cell.locator in cyclic:
            cell.evaluation_status = "CYCLIC_COMPONENT"
            continue
        result = compiled.cells.get(cell.locator)
        if result is None:
            cell.evaluation_status = "UNAVAILABLE:NOT_IN_EVALUATOR_GRAPH"
            continue
        value = result.get("value")
        if isinstance(value, str) and value.startswith("#"):
            cell.evaluation_status = "UNAVAILABLE:EVALUATOR_ERROR"
            continue
        cell.evaluated_value = value
        cell.evaluation_status = "CALCULATED_ACYCLIC"


def main() -> int:
    ap = argparse.ArgumentParser(description="Capture a workbook losslessly (L1)")
    ap.add_argument("--workbook", type=Path, required=True)
    ap.add_argument("--out", type=Path)
    a = ap.parse_args()

    g = capture(a.workbook)
    s = g.stats()
    print(f"[source_graph] {g.workbook}  {g.digest[:19]}…")
    print(f"  fogli            : {s['sheets']}")
    print(f"  celle            : {s['cells']}")
    for k, v in sorted(s["by_kind"].items(), key=lambda kv: -kv[1]):
        print(f"      {k:9}: {v}")
    print(f"  archi precedent  : {s['precedent_edges']}")
    print(f"  range uniti      : {s['merged_ranges']}")
    print(f"  defined names    : {s['defined_names']}")

    if a.out:
        a.out.mkdir(parents=True, exist_ok=True)
        p = a.out / "source_graph.json"
        p.write_text(json.dumps(g.to_json(), indent=2, ensure_ascii=False, default=str),
                     encoding="utf-8")
        print(f"  → {p} ({p.stat().st_size // 1024}KB)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
