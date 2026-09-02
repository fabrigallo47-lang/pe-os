#!/usr/bin/env python3
"""Acceptance tests for PAN-102 XLSX semantic chunk context."""

from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from tools.extract_v2 import parse_xlsx


class PAN102XlsxSemanticContextTests(unittest.TestCase):
    def _workbook(self, directory: str) -> Path:
        path = Path(directory) / "pan102_semantic_model.xlsx"
        workbook = Workbook()
        model = workbook.active
        model.title = "Model"
        inputs = workbook.create_sheet("Inputs")
        inputs["A1"] = "Metric"
        inputs["C1"] = "FY2025E"
        inputs["A2"] = "EBITDA"
        inputs["C2"] = 15

        model.merge_cells("A1:D1")
        model["A1"] = "Operating Model"
        model["A2"] = "Line Item"
        model["B2"] = "Unit"
        model.merge_cells("C2:D2")
        model["C2"] = "Forecast"
        model["C3"] = "FY2025E"
        model["D3"] = "FY2026E"

        blue = Font(color="0000FF")
        black = Font(color="000000")
        green = Font(color="008000")
        red = Font(color="FF0000")

        model["A4"] = "Revenue"
        model["B4"] = "$m"
        model["C4"] = 90
        model["D4"] = 100
        model["C4"].font = blue
        model["D4"].font = blue

        model["A5"] = "Revenue Growth"
        model["B5"] = "%"
        model["C5"] = "=C4/80-1"
        model["D5"] = "=D4/C4-1"
        model["C5"].font = black
        model["D5"].font = black

        model["A6"] = "Linked EBITDA"
        model["B6"] = "$m"
        model["C6"] = "='Inputs'!C2"
        model["C6"].font = green

        model["A7"] = "External Price"
        model["B7"] = "$m"
        model["C7"] = "='[external.xlsx]Sheet1'!C2"
        model["C7"].font = red

        model.merge_cells("A8:D8")
        model["A8"] = "Working Capital"
        model["A9"] = "DSO"
        model["B9"] = "days"
        model["C9"] = 64
        model["C9"].font = blue
        model["D9"] = 62
        model["D9"].fill = PatternFill(fill_type="solid", fgColor="0000FF")

        # Formula-bearing sheets can contain large numeric grids that do not
        # carry labels or declared input formatting.  They remain out of the
        # L2 text path and in the separate cell graph.
        model["C10"] = 111
        model["D10"] = 222
        workbook.save(path)
        return path

    def _model_body(self, path: Path) -> str:
        chunks = parse_xlsx(path, max_words=5000)
        return "\n".join(
            chunk.body for chunk in chunks if "::Model!" in chunk.locator
        )

    def test_formula_free_label_and_input_rows_are_not_silently_dropped(self):
        with tempfile.TemporaryDirectory() as directory:
            body = self._model_body(self._workbook(directory))

        self.assertIn("A4=Revenue", body)
        self.assertIn("C4=90", body)
        self.assertIn("role=input", body)
        self.assertIn("A8=Working Capital", body)
        self.assertIn("A9=DSO", body)

    def test_merged_and_multirow_headers_form_a_full_cell_path(self):
        with tempfile.TemporaryDirectory() as directory:
            body = self._model_body(self._workbook(directory))

        self.assertIn("B1=Operating Model", body)
        self.assertIn("D2=Forecast", body)
        self.assertIn(
            "C4=90 [header_path=Operating Model > Forecast > FY2025E > Revenue > $m; role=input]",
            body,
        )
        dso_line = next(line for line in body.splitlines() if "A9=DSO" in line)
        self.assertIn("Working Capital", dso_line)
        self.assertIn("FY2025E", dso_line)

    def test_font_and_fill_colors_expose_financial_model_roles(self):
        with tempfile.TemporaryDirectory() as directory:
            body = self._model_body(self._workbook(directory))

        self.assertRegex(body, r"C5=FORMULA\(=C4/80-1\).*role=formula")
        self.assertRegex(body, r"C6=FORMULA\(='Inputs'!C2\).*role=cross_sheet_link")
        self.assertRegex(body, r"C7=FORMULA\(='\[external\.xlsx\]Sheet1'!C2\).*role=external_link")
        self.assertRegex(body, r"D9=62 .*role=input")

    def test_unlabeled_numeric_noise_remains_out_of_formula_sheet_chunks(self):
        with tempfile.TemporaryDirectory() as directory:
            body = self._model_body(self._workbook(directory))

        self.assertNotIn("C10=111", body)
        self.assertNotIn("D10=222", body)

    def test_output_is_deterministic(self):
        with tempfile.TemporaryDirectory() as directory:
            path = self._workbook(directory)
            first = [(chunk.locator, chunk.body, chunk.chunk_id) for chunk in parse_xlsx(path, 5000)]
            second = [(chunk.locator, chunk.body, chunk.chunk_id) for chunk in parse_xlsx(path, 5000)]

        self.assertEqual(first, second)


if __name__ == "__main__":
    unittest.main()
