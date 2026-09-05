"""Read original, case-scoped source bytes at a cited, immutable version.

The reader consumes the existing source-envelope/1.0. It never resolves a
client-provided path or substitutes another version with the same filename.
"""
from __future__ import annotations

import base64
import hashlib
import html
import io
import re
from dataclasses import dataclass
from pathlib import Path
from urllib.parse import urlencode

from fastapi import HTTPException

SCHEMA = "source-document/1.0"
MEDIA = {
    ".pptx": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
    ".eml": "message/rfc822",
    ".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".webp": "image/webp",
    ".pdf": "application/pdf",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".xlsm": "application/vnd.ms-excel.sheet.macroEnabled.12",
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".txt": "text/plain", ".md": "text/plain", ".csv": "text/csv",
    ".srt": "text/plain", ".vtt": "text/vtt",
    ".wav": "audio/wav", ".mp3": "audio/mpeg", ".m4a": "audio/mp4",
    ".mp4": "video/mp4", ".webm": "video/webm", ".ogg": "audio/ogg",
}
HEADERS = {"Cache-Control": "private, no-store", "X-Content-Type-Options": "nosniff"}


@dataclass(frozen=True)
class SourceDocument:
    case_id: str
    source_id: str
    version_id: str
    filename: str
    suffix: str
    data: bytes

    @property
    def media_type(self):
        return MEDIA.get(self.suffix, "application/octet-stream")


def resolve_document(vault: Path, records: list[dict], case_id: str,
                     source_id: str, version_id: str) -> SourceDocument:
    if not re.fullmatch(r"[A-Za-z0-9_-][A-Za-z0-9._-]*", case_id):
        raise HTTPException(400, "Invalid case reference.")
    if not re.fullmatch(r"sha256:[a-f0-9]{64}", version_id):
        raise HTTPException(422, "The original source version has no verifiable content hash.")
    candidates = []
    for record in records:
        if not isinstance(record, dict):
            continue
        envelope = record.get("source_envelope")
        if not isinstance(envelope, dict):
            continue
        if (record.get("case_id") != case_id or envelope.get("case_id") != case_id
                or envelope.get("source_id") != source_id
                or envelope.get("source_version_id") != version_id):
            continue
        filename = envelope.get("stored_filename")
        if not isinstance(filename, str) or not filename or Path(filename).name != filename:
            continue
        path = (vault / "inbox" / filename).resolve()
        if path.parent != (vault / "inbox").resolve() or not path.is_file():
            continue
        candidates.append((path, envelope))
    if not candidates:
        raise HTTPException(404, "The cited original file is not available for this case and version.")
    # Duplicate registry records are normal; each read verifies the actual bytes.
    for path, envelope in candidates:
        try:
            data = path.read_bytes()
        except OSError:
            continue
        if "sha256:" + hashlib.sha256(data).hexdigest() == version_id:
            original_name = str(envelope.get("original_filename") or path.name)
            return SourceDocument(case_id, source_id, version_id, original_name, path.suffix.lower(), data)
    raise HTTPException(409, "The original file no longer matches the cited version. No replacement was opened.")


def document_url(document: SourceDocument, action: str, locator: str = "", claim_id: str = "") -> str:
    query = {"source_id": document.source_id, "source_version_id": document.version_id}
    if locator:
        query["locator"] = locator
    if claim_id:
        query["claim_id"] = claim_id
    return f"/api/v20/cases/{document.case_id}/source-document/{action}?{urlencode(query)}"


def _text(document: SourceDocument) -> str:
    try:
        return document.data.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise HTTPException(422, "This source is not valid UTF-8 text. Download the original to inspect it.") from exc


def _timecode(locator: str) -> tuple[float, float | None] | None:
    address = re.sub(r"^cue:\d+:", "", locator.rsplit("::", 1)[-1])
    times = re.findall(r"(?<!\d)(\d{1,3}):([0-5]\d):([0-5]\d)([.,]\d+)?", address)
    if not times:
        return None
    parsed = [int(h) * 3600 + int(m) * 60 + int(s) + float("0" + (fraction or "").replace(",", "."))
              for h, m, s, fraction in times]
    if len(parsed) > 1 and parsed[1] < parsed[0]:
        raise HTTPException(422, "The cited end time precedes its start time.")
    return parsed[0], parsed[1] if len(parsed) > 1 else None


def workbook_positions(locator: str, sheets: dict[str, tuple[int, int]]) -> dict:
    """Resolve one or several explicit addresses against dimensions read from the file."""
    from openpyxl.utils.cell import range_boundaries
    address = locator.split("::", 1)[-1]
    # Split conjunctions only outside quoted sheet names.
    parts, beginning, quoted, index = [], 0, False, 0
    while index < len(address):
        if address[index] == "'":
            if quoted and index + 1 < len(address) and address[index + 1] == "'":
                index += 2
                continue
            quoted = not quoted
        separator = re.match(r"(?: and |;\s*)", address[index:]) if not quoted else None
        if separator:
            parts.append(address[beginning:index].strip())
            index += len(separator[0])
            beginning = index
        else:
            index += 1
    parts.append(address[beginning:].strip())
    selections = []
    for part in parts:
        match = re.fullmatch(r"(.+?)!(\$?[A-Za-z]{1,3}\$?[1-9]\d*(?::\$?[A-Za-z]{1,3}\$?[1-9]\d*)?|[1-9]\d*:[1-9]\d*)", part)
        if not match:
            return {"kind": "workbook", "status": "UNRESOLVED", "label": "Exact sheet and range could not be resolved"}
        sheet = match[1]
        if sheet.startswith("'") and sheet.endswith("'"):
            sheet = sheet[1:-1].replace("''", "'")
        if sheet not in sheets:
            raise HTTPException(422, "The cited sheet does not exist in this source version.")
        max_row, max_column = sheets[sheet]
        c1, r1, c2, r2 = range_boundaries(match[2].replace("$", "").upper())
        c1, c2 = c1 or 1, c2 or max(1, max_column)
        if r1 > r2 or c1 > c2 or r2 > max_row or c2 > max_column:
            raise HTTPException(422, "The cited range is outside this source version.")
        selections.append({"kind": "workbook", "status": "LOCATED", "label": f"{sheet}!{match[2]}",
                           "sheet": sheet, "bounds": [c1, r1, c2, r2],
                           "max_row": max_row, "max_column": max_column})
    if sum((p["bounds"][3] - p["bounds"][1] + 1) * (p["bounds"][2] - p["bounds"][0] + 1) for p in selections) > 5000:
        raise HTTPException(422, "This cited range is too large for the focused reader. Download the original workbook.")
    result = dict(selections[0])
    if len(selections) > 1:
        result.update(selections=selections, label="; ".join(p["label"] for p in selections))
    return result


def workbook_dimensions(book) -> dict:
    dimensions = {}
    for sheet in book:
        if sheet.max_row is None or sheet.max_column is None:
            sheet.calculate_dimension(force=True)
        dimensions[sheet.title] = (sheet.max_row or 0, sheet.max_column or 0)
    return dimensions


def _workbook_position(document: SourceDocument, locator: str) -> dict:
    from openpyxl import load_workbook
    book = load_workbook(io.BytesIO(document.data), read_only=True, keep_links=False)
    try:
        return workbook_positions(locator, workbook_dimensions(book))
    finally:
        book.close()


def _heading_spans(lines: list[str], locator: str) -> list[list[int]]:
    address = locator.split("::", 1)[-1].strip()
    headings = [(i + 1, len(m[1]), m[2].strip()) for i, line in enumerate(lines)
                if (m := re.match(r"^(#{1,6})\s+(.+?)\s*#*\s*$", line))]
    def match_name(name):
        name = re.sub(r"^#{1,6}\s+", "", name).strip()
        found = [(i, level) for i, level, title in headings if title == name]
        if len(found) != 1:
            return None
        start, level = found[0]
        end = next((i - 1 for i, next_level, _ in headings if i > start and next_level <= level), len(lines))
        return [start, end]
    if span := match_name(address):
        return [span]
    # Multiple explicitly named headings, never fuzzy topic matching.
    parts = re.split(r" / | and ", address)
    if len(parts) > 1 and all(spans := [match_name(part) for part in parts]):
        return spans
    return []


def _document_lines(document: SourceDocument) -> list[str]:
    if document.suffix == ".docx":
        from docx2python import docx2python
        # Share the extractor's physical block numbering, including tables.
        from tools.extract_v2_physical import _docx2python_render_section
        with docx2python(io.BytesIO(document.data)) as doc:
            return [text for section in doc.body
                    if (text := _docx2python_render_section(section)) is not None]
    return _text(document).splitlines()


def locate_document(document: SourceDocument, locator: str) -> dict:
    if document.suffix in {".pptx", ".eml", ".png", ".jpg", ".jpeg", ".webp"} or (document.suffix == ".docx" and locator.split("::", 1)[-1].startswith("section:")):
        from app.source_document_formats import extra_position
        return extra_position(document, locator)
    if document.suffix == ".pdf":
        import pdfplumber
        match = re.search(r"(?:^|::)p([1-9]\d*)(?=:|$)", locator)
        with pdfplumber.open(io.BytesIO(document.data)) as pdf:
            count = len(pdf.pages)
            page = int(match[1]) if match else 1
            from app.source_document_formats import rectangle
            box = rectangle(locator, pdf.pages[page - 1].width, pdf.pages[page - 1].height) if page <= count else None
        if page > count:
            raise HTTPException(422, "The cited page does not exist in this source version.")
        return {"kind": "pdf", "status": "LOCATED" if match else "UNRESOLVED",
                "page": page, "page_count": count, "box": box,
                "label": f"Page {page}" if match else "Exact page not supplied"}
    if document.suffix in {".xlsx", ".xlsm"}:
        return _workbook_position(document, locator)
    if document.media_type.startswith(("audio/", "video/")):
        times = _timecode(locator)
        return {"kind": "media", "status": "LOCATED" if times else "UNRESOLVED",
                "label": locator if times else "Exact time not supplied",
                "start": times[0] if times else 0, "end": times[1] if times else None}
    if document.suffix not in {".txt", ".md", ".csv", ".srt", ".vtt", ".docx"}:
        return {"kind": "download", "status": "UNRESOLVED",
                "label": "A focused reader is not available for this file type"}
    lines = _document_lines(document)
    span = re.search(r"(?:^|::)(?:lines|blocks):([1-9]\d*)-([1-9]\d*)(?=:|$)", locator)
    if not span:
        span = re.fullmatch(r"L([1-9]\d*)(?:-L?([1-9]\d*))?", locator)
    start, end = (int(span[1]), int(span[2] or span[1])) if span else (None, None)
    if document.suffix in {".srt", ".vtt"} and (times := _timecode(locator)):
        cue = re.search(r"::cue:([1-9]\d*):", locator)
        timing_lines = [(i, _timecode(line)) for i, line in enumerate(lines) if "-->" in line]
        matches = [(ordinal, i) for ordinal, (i, tc) in enumerate(timing_lines, 1)
                   if tc and tc[0] == times[0] and (times[1] is None or tc[1] == times[1])
                   and (not cue or int(cue[1]) == ordinal)]
        if len(matches) == 1:
            start = matches[0][1] + 1
            end = start
            while end < len(lines) and lines[end].strip():
                end += 1
    spans = []
    if document.suffix == ".md":
        # Legacy graph references can record literal line ranges or section names.
        numbered = re.search(r"(?:^|, )lines ([1-9]\d*)-([1-9]\d*)$", locator)
        if numbered:
            start, end = int(numbered[1]), int(numbered[2])
        elif not start:
            spans = _heading_spans(lines, locator)
            if spans:
                start, end = spans[0]
    if start is not None and (end < start or end > len(lines)):
        raise HTTPException(422, "The cited passage is outside this source version.")
    return {"kind": "text", "status": "LOCATED" if start else "UNRESOLVED",
            "label": locator if start else "Exact passage could not be located",
            "start": start, "end": end, "spans": spans or ([[start, end]] if start else []), "line_count": len(lines)}


def _pdf_body(document: SourceDocument, position: dict, quote: str) -> str:
    import pdfplumber
    from PIL import ImageDraw
    with pdfplumber.open(io.BytesIO(document.data)) as pdf:
        page = pdf.pages[position["page"] - 1]
        image = page.to_image(resolution=120).original.convert("RGB")
        highlighted = False
        selection_top = None
        if quote.strip():
            tokens = page.extract_words()
            needle = quote.split()
            words = [item["text"] for item in tokens]
            matches = [i for i in range(len(words) - len(needle) + 1) if words[i:i + len(needle)] == needle]
            # Ambiguous repeated text does not establish an exact highlight.
            if len(matches) == 1:
                drawing = ImageDraw.Draw(image, "RGBA")
                sx, sy = image.width / page.width, image.height / page.height
                for word in tokens[matches[0]:matches[0] + len(needle)]:
                    drawing.rectangle([(word["x0"] - page.bbox[0]) * sx, (word["top"] - page.bbox[1]) * sy,
                                       (word["x1"] - page.bbox[0]) * sx, (word["bottom"] - page.bbox[1]) * sy],
                                      fill=(255, 210, 70, 90))
                highlighted = True
                selection_top = (min(word['top'] for word in tokens[matches[0]:matches[0] + len(needle)]) - page.bbox[1]) / page.height * 100
        if position.get("box"):
            left, top, right, bottom = position["box"]
            sx, sy = image.width / page.width, image.height / page.height
            ImageDraw.Draw(image).rectangle([left*sx, top*sy, right*sx, bottom*sy], outline=(190, 145, 0), width=3)
            selection_top = top / page.height * 100
        buffer = io.BytesIO()
        image.save(buffer, format="PNG")
    encoded = base64.b64encode(buffer.getvalue()).decode("ascii")
    note = "Cited region outlined." if position.get("box") else "Cited passage highlighted." if highlighted else "Original page. No exact passage highlight is available."
    anchor = f'<span id="selection" style="position:absolute;top:{selection_top:.3f}%;scroll-margin-top:20px"></span>' if selection_top is not None else ""
    return f'<p>{note}</p><div style="position:relative">{anchor}<img class="pdf" alt="Original page {position["page"]}" src="data:image/png;base64,{encoded}"></div>'


def _workbook_body(document: SourceDocument, position: dict) -> str:
    if position["status"] != "LOCATED":
        return ""
    if position.get("selections"):
        return "".join(f'<h2>{html.escape(p["label"])}</h2>' + _workbook_body(document, p).replace('id="selection"', f'id="selection{index or ""}"')
                       for index, p in enumerate(position["selections"]))
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    formulas = load_workbook(io.BytesIO(document.data), read_only=True, data_only=False, keep_links=False)
    values = load_workbook(io.BytesIO(document.data), read_only=True, data_only=True, keep_links=False)
    try:
        c1, r1, c2, r2 = position["bounds"]
        left, right = max(1, c1 - 1), min(position["max_column"], c2 + 1)
        first, last = max(1, r1 - 2), min(position["max_row"], r2 + 2)
        ws, cached = formulas[position["sheet"]], values[position["sheet"]]
        rows = ['<tr><th scope="col">Row</th>' + "".join(
            f'<th scope="col">{get_column_letter(c)}</th>' for c in range(left, right + 1)) + "</tr>"]
        formula_rows = ws.iter_rows(min_row=first, max_row=last, min_col=left, max_col=right)
        cached_rows = cached.iter_rows(min_row=first, max_row=last, min_col=left, max_col=right)
        for row_index, (row, cache) in enumerate(zip(formula_rows, cached_rows), first):
            cells = []
            for column, (cell, cached_cell) in enumerate(zip(row, cache), left):
                selected = r1 <= row_index <= r2 and c1 <= column <= c2
                value = cell.value
                address = f"{get_column_letter(column)}{row_index}"
                text = html.escape(str(value)) if value is not None else "—"
                if cell.data_type == "f":
                    display = html.escape(str(cached_cell.value)) if cached_cell.value is not None else "No cached value"
                    text = f"<strong>{display}</strong><code>{text}</code>"
                cells.append(f'<td class="{"selected" if selected else ""}" aria-label="{address}">{text}</td>')
            anchor = ' id="selection"' if row_index == r1 else ""
            rows.append(f'<tr{anchor}><th scope="row">{row_index}</th>{"".join(cells)}</tr>')
        return '<p>Original worksheet values and formulas. Selected cells are highlighted; formulas are not recalculated.</p><div class="sheet"><table>' + "".join(rows) + "</table></div>"
    finally:
        formulas.close()
        values.close()


def render_document(document: SourceDocument, locator: str, position: dict, quote: str = "") -> str:
    file_url = document_url(document, "file")
    if position.get("native"):
        from app.source_document_formats import extra_body
        body = extra_body(document, position)
    elif position["kind"] == "pdf":
        body = _pdf_body(document, position, quote)
        page, count = position["page"], position["page_count"]
        navigation = []
        if page > 1:
            navigation.append(f'<a href="{html.escape(document_url(document, "view", f"p{page-1}"))}">Previous page</a>')
        if page < count:
            navigation.append(f'<a href="{html.escape(document_url(document, "view", f"p{page+1}"))}">Next page</a>')
        body = f'<nav>{" · ".join(navigation)} <span>Page {page} of {count}</span></nav>' + body
    elif position["kind"] == "workbook":
        body = _workbook_body(document, position)
    elif position["kind"] == "media":
        tag = "video" if document.media_type.startswith("video/") else "audio"
        end = f',{position["end"]}' if position["end"] is not None else ""
        body = f'<{tag} controls preload="metadata" src="{html.escape(file_url)}#t={position["start"]}{end}"></{tag}>'
    elif position["kind"] == "text":
        lines = _document_lines(document)
        body = '<div class="lines">' + "".join(
            f'<div class="text-line{" selected" if any(start <= i <= end for start, end in position.get("spans", [])) else ""}"'
            + (' id="selection"' if i == position["start"] else "")
            + f'><span>{i}</span><pre>{html.escape(line)}</pre></div>' for i, line in enumerate(lines, 1)) + "</div>"
    else:
        body = "<p>Download this source to inspect its original content.</p>"
    warning = "" if position["status"] == "LOCATED" else '<p class="notice">The exact location is unavailable. This is the original source, without a verified passage selection.</p>'
    return f"""<!doctype html><html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{html.escape(document.filename)}</title><style>
:root{{color-scheme:light dark}}body{{font:15px/1.5 system-ui,sans-serif;margin:0;padding:20px;background:light-dark(#fff,#151b17);color:light-dark(#20251f,#eff2e9)}}
header{{border-bottom:1px solid #7f887c;padding-bottom:16px;margin-bottom:18px}}h1{{font-size:18px;margin:0 0 8px;overflow-wrap:anywhere}}p{{margin:10px 0}}a{{color:inherit}}nav{{display:flex;gap:16px;flex-wrap:wrap}}small{{display:block;overflow-wrap:anywhere;opacity:.8}}.pdf{{width:100%;height:auto;display:block}}.sheet{{overflow:auto}}table{{border-collapse:collapse;min-width:100%}}td,th{{border:1px solid #7f887c;padding:10px;text-align:left;min-width:80px}}td{{max-width:350px;overflow-wrap:anywhere}}th{{font-weight:500}}code{{display:block;white-space:pre-wrap;margin-top:4px}}.selected{{background:light-dark(#fff0a3,#51481c)}}.text-line{{display:grid;grid-template-columns:40px minmax(0,1fr);padding:5px;scroll-margin:15px}}.text-line>span{{opacity:.6}}pre{{margin:0;white-space:pre-wrap;overflow-wrap:anywhere;font:inherit}}audio,video{{width:100%}}.notice{{border-left:3px solid #b39955;padding-left:12px}}
</style></head><body><header><h1>{html.escape(document.filename)}</h1><div>{html.escape(position["label"])}</div>
<small>Original version · {html.escape(document.version_id)}</small><a href="{html.escape(file_url + '&download=true')}" download>Download original</a></header>{warning}{body}</body></html>"""
